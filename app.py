import os
import re
import secrets
import sqlite3
import smtplib
import math
import hashlib
import csv
import json
from io import BytesIO, StringIO
import threading
import time
from html import escape
from datetime import datetime, timedelta, timezone
from email.message import EmailMessage
from functools import wraps
from pathlib import Path
from urllib.parse import urlencode, urlparse

import pandas as pd
import airportsdata
from dotenv import load_dotenv
from flask import (
    Flask,
    abort,
    flash,
    g,
    has_request_context,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.graphics.barcode import createBarcodeDrawing
from reportlab.pdfgen import canvas
from reportlab.graphics import renderPDF
from reportlab.graphics.shapes import Drawing
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename

load_dotenv()

BASE_DIR = Path(__file__).resolve().parent
DB_PATH = BASE_DIR / "safar24.db"
UPLOAD_DIR = BASE_DIR / "static" / "uploads" / "payment_proofs"
TICKETS_DIR = BASE_DIR / "static" / "tickets"
ALLOWED_PROOF_EXT = {"png", "jpg", "jpeg", "webp"}
REQUIRED_FLIGHT_COLUMNS = {
    "route_code",
    "from_city",
    "to_city",
    "departure_date",
    "return_date",
    "airline",
    "price_uzs",
    "seats",
    "status",
}
ALLOWED_FLIGHT_STATUSES = {"active", "inactive"}
EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
PHONE_RE = re.compile(r"^\+?\d{9,15}$")
PASSPORT_NUMBER_RE = re.compile(r"^[A-Z0-9]{7,15}$")
PASSPORT_SERIES_RE = re.compile(r"^[A-Z]{2}$")
LATIN_SEARCH_RE = re.compile(r"^[A-Za-z0-9\s(),.'-]+$")
PASSWORD_MIN_LENGTH = 8
LOGIN_MAX_ATTEMPTS = 5
LOGIN_WINDOW_SECONDS = 15 * 60
LOGIN_ATTEMPTS: dict[str, list[float]] = {}
AIRPORTS = airportsdata.load("IATA")
PASSPORT_ISSUE_MAX_AGE_YEARS = 10
PASSPORT_EXPIRY_MAX_YEARS = 20
PASSPORT_MIN_VALIDITY_DAYS = 180
BOOKING_HOLD_MINUTES = 30
PASSENGER_TYPE_LABELS = {
    "INF": "INF (0-2)",
    "CHD": "CHD (2-12)",
    "ADT": "ADT",
}
PASSENGER_FARE_RATIO = {
    "ADT": float(os.getenv("FARE_RATIO_ADT", "1.00")),
    "CHD": float(os.getenv("FARE_RATIO_CHD", "0.75")),
    "INF": float(os.getenv("FARE_RATIO_INF", "0.10")),
}
PAYMENT_METHOD_FEE_PERCENT = {
    "wallet": float(os.getenv("FEE_PCT_WALLET", "0.0")),
    "click": float(os.getenv("FEE_PCT_CLICK", "1.5")),
    "payme": float(os.getenv("FEE_PCT_PAYME", "1.8")),
}
MAX_BOOKING_TRAVELERS = int(os.getenv("MAX_BOOKING_TRAVELERS", "9"))
PASSWORD_RESET_TOKEN_MINUTES = int(os.getenv("PASSWORD_RESET_TOKEN_MINUTES", "20"))
PASSWORD_RESET_MAX_ATTEMPTS = int(os.getenv("PASSWORD_RESET_MAX_ATTEMPTS", "5"))
PASSWORD_RESET_WINDOW_SECONDS = int(os.getenv("PASSWORD_RESET_WINDOW_SECONDS", str(15 * 60)))
PASSWORD_RESET_ATTEMPTS: dict[str, list[float]] = {}
PASSWORD_RESET_RESEND_SECONDS = int(os.getenv("PASSWORD_RESET_RESEND_SECONDS", "120"))
SECURITY_CODE_MINUTES = int(os.getenv("SECURITY_CODE_MINUTES", "10"))
AUTH_SESSION_MAX_AGE_DAYS = int(os.getenv("AUTH_SESSION_MAX_AGE_DAYS", "30"))
SAFE_SQL_IDENTIFIER_RE = re.compile(r"^[A-Za-z_][A-Za-z0-9_]*$")
SUPER_DB_EXPORT_FORMATS = {"csv", "json", "xlsx"}
SUPER_DB_PAGE_SIZE = 30


def env_flag(name: str, default: str = "0") -> bool:
    return (os.getenv(name, default) or "").strip().lower() in {"1", "true", "yes", "on"}

app = Flask(__name__)
app.config["SECRET_KEY"] = os.getenv("SECRET_KEY", "safar24-dev-secret-change-me")
app.config["MAX_CONTENT_LENGTH"] = 5 * 1024 * 1024
app.config["SESSION_COOKIE_NAME"] = "safar24_session"
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_SECURE"] = os.getenv("SESSION_COOKIE_SECURE", "0") == "1"

MAIL_ENABLED = env_flag("MAIL_ENABLED", "0")
MAIL_SMTP_HOST = os.getenv("MAIL_SMTP_HOST", "smtp.gmail.com")
MAIL_SMTP_PORT = int(os.getenv("MAIL_SMTP_PORT", "587"))
MAIL_USERNAME = (os.getenv("MAIL_USERNAME") or "").strip()
MAIL_PASSWORD = os.getenv("MAIL_PASSWORD", "")
MAIL_FROM_EMAIL = (os.getenv("MAIL_FROM_EMAIL") or MAIL_USERNAME or "").strip()
MAIL_USE_TLS = env_flag("MAIL_USE_TLS", "1")
MAIL_USE_SSL = env_flag("MAIL_USE_SSL", "0")
USD_TO_UZS_RATE = float(os.getenv("USD_TO_UZS_RATE", "12700"))
APP_BASE_URL = (os.getenv("APP_BASE_URL", "http://127.0.0.1:5003") or "").rstrip("/")
REGISTRATION_CODE_MINUTES = int(os.getenv("REGISTRATION_CODE_MINUTES", "10"))
REGISTRATION_RESEND_SECONDS = int(os.getenv("REGISTRATION_RESEND_SECONDS", "120"))
UI_SUPPORTED_LANGUAGES = {"uz", "ru", "en"}
UI_DEFAULT_LANGUAGE = "uz"
RTL_LANGUAGES = {"ar", "fa", "he", "ur"}


def normalize_ui_language(raw_value: str | None) -> str:
    value = (raw_value or "").strip().lower()
    if not value:
        return UI_DEFAULT_LANGUAGE

    short = value[:2]
    if short in UI_SUPPORTED_LANGUAGES:
        return short

    return UI_DEFAULT_LANGUAGE


def detect_accept_language() -> str:
    raw_header = (request.headers.get("Accept-Language") or "").strip()
    if not raw_header:
        return UI_DEFAULT_LANGUAGE

    for token in raw_header.split(","):
        lang_token = token.split(";")[0].strip().lower()
        normalized = normalize_ui_language(lang_token)
        if normalized in UI_SUPPORTED_LANGUAGES:
            return normalized

    return UI_DEFAULT_LANGUAGE


@app.before_request
def resolve_ui_language_context():
    if request.endpoint == "static":
        selected = normalize_ui_language(str(session.get("ui_lang") or UI_DEFAULT_LANGUAGE))
        g.ui_lang = selected
        g.ui_dir = "rtl" if selected in RTL_LANGUAGES else "ltr"
        return None

    query_lang = (request.args.get("lang") or "").strip()

    if query_lang:
        selected = normalize_ui_language(query_lang)
        session["ui_lang"] = selected
    elif session.get("ui_lang"):
        selected = normalize_ui_language(str(session.get("ui_lang") or ""))
        session["ui_lang"] = selected
    else:
        selected = detect_accept_language()
        session["ui_lang"] = selected

    g.ui_lang = selected
    g.ui_dir = "rtl" if selected in RTL_LANGUAGES else "ltr"
    return None


@app.context_processor
def inject_ui_context():
    lang = normalize_ui_language(getattr(g, "ui_lang", session.get("ui_lang") or UI_DEFAULT_LANGUAGE))
    return {
        "ui_lang": lang,
        "ui_dir": "rtl" if lang in RTL_LANGUAGES else "ltr",
        "fare_ratio": {
            "ADT": fare_ratio_for_passenger_type("ADT"),
            "CHD": fare_ratio_for_passenger_type("CHD"),
            "INF": fare_ratio_for_passenger_type("INF"),
        },
        "payment_fee_options": {
            "wallet": payment_fee_percent("wallet"),
            "click": payment_fee_percent("click"),
            "payme": payment_fee_percent("payme"),
        },
        "max_booking_travelers": MAX_BOOKING_TRAVELERS,
    }


def normalize_phone_input(raw_phone: str | None) -> str:
    value = (raw_phone or "").strip()
    if not value:
        return ""

    normalized = re.sub(r"[^\d+]", "", value)
    if normalized.startswith("00"):
        normalized = f"+{normalized[2:]}"
    if normalized and not normalized.startswith("+"):
        normalized = f"+{normalized}"
    return normalized


# ----------------------------
# Database
# ----------------------------
def get_db() -> sqlite3.Connection:
    if "db" not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
    return g.db


@app.teardown_appcontext
def close_db(_error) -> None:
    db = g.pop("db", None)
    if db is not None:
        db.close()


@app.before_request
def cleanup_booking_holds_before_request():
    if request.endpoint == "static":
        return None

    db = get_db()
    released_count = cleanup_expired_booking_holds(db)
    if released_count:
        db.commit()
    return None


def init_db() -> None:
    db = get_db()
    db.executescript(
        """
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            email TEXT UNIQUE NOT NULL,
            full_name TEXT NOT NULL,
            phone TEXT,
            password_hash TEXT NOT NULL,
            account_id TEXT UNIQUE NOT NULL,
            balance_uzs INTEGER NOT NULL DEFAULT 0,
            balance_usd REAL NOT NULL DEFAULT 0,
            is_admin INTEGER NOT NULL DEFAULT 0,
            is_super_admin INTEGER NOT NULL DEFAULT 0,
            role_granted_by INTEGER,
            role_granted_at TEXT,
            created_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS user_profiles (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER UNIQUE NOT NULL,
            passport_number TEXT NOT NULL,
            passport_series TEXT,
            birth_date TEXT NOT NULL,
            nationality TEXT NOT NULL,
            passport_issue_date TEXT,
            passport_expiration_date TEXT,
            gender TEXT,
            notification_email TEXT,
            FOREIGN KEY(user_id) REFERENCES users(id)
        );

        CREATE TABLE IF NOT EXISTS flights (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            route_code TEXT,
            from_city TEXT NOT NULL,
            from_airport_code TEXT,
            to_city TEXT NOT NULL,
            to_airport_code TEXT,
            departure_date TEXT NOT NULL,
            departure_time TEXT,
            return_date TEXT,
            is_oneway INTEGER NOT NULL DEFAULT 0,
            airline TEXT NOT NULL,
            travel_class TEXT NOT NULL DEFAULT 'Economy',
            price_uzs INTEGER NOT NULL,
            price_value REAL NOT NULL,
            price_currency TEXT NOT NULL DEFAULT 'UZS',
            baggage_kg INTEGER NOT NULL DEFAULT 20,
            include_meal INTEGER NOT NULL DEFAULT 1,
            include_return INTEGER NOT NULL DEFAULT 1,
            seats INTEGER NOT NULL,
            status TEXT NOT NULL DEFAULT 'active',
            created_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS bookings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            flight_id INTEGER NOT NULL,
            baggage_option TEXT NOT NULL DEFAULT 'standard',
            seat_numbers TEXT,
            passenger_count INTEGER NOT NULL DEFAULT 1,
            adult_count INTEGER NOT NULL DEFAULT 1,
            child_count INTEGER NOT NULL DEFAULT 0,
            infant_count INTEGER NOT NULL DEFAULT 0,
            seat_count INTEGER NOT NULL DEFAULT 1,
            status TEXT NOT NULL DEFAULT 'pending_payment',
            reservation_expires_at TEXT,
            cancelled_at TEXT,
            expired_at TEXT,
            cancellation_reason TEXT,
            created_at TEXT NOT NULL,
            FOREIGN KEY(user_id) REFERENCES users(id),
            FOREIGN KEY(flight_id) REFERENCES flights(id)
        );

        CREATE TABLE IF NOT EXISTS passengers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            booking_id INTEGER NOT NULL,
            full_name TEXT NOT NULL,
            first_name TEXT,
            last_name TEXT,
            passenger_type TEXT,
            passport_number TEXT,
            passport_series TEXT,
            birth_date TEXT,
            nationality TEXT,
            gender TEXT,
            passport_issue_date TEXT,
            passport_expiration_date TEXT,
            phone TEXT,
            notification_email TEXT,
            travel_class TEXT NOT NULL DEFAULT 'Economy',
            seat_number TEXT,
            created_at TEXT NOT NULL,
            FOREIGN KEY(booking_id) REFERENCES bookings(id)
        );

        CREATE TABLE IF NOT EXISTS payments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            booking_id INTEGER UNIQUE NOT NULL,
            amount_uzs INTEGER NOT NULL,
            amount_value REAL,
            currency TEXT NOT NULL DEFAULT 'UZS',
            base_amount_uzs INTEGER,
            base_amount_value REAL,
            method_fee_pct REAL NOT NULL DEFAULT 0,
            method_fee_uzs INTEGER NOT NULL DEFAULT 0,
            method_fee_value REAL NOT NULL DEFAULT 0,
            proof_image TEXT,
            payment_method TEXT NOT NULL DEFAULT 'manual',
            payment_reference TEXT,
            status TEXT NOT NULL DEFAULT 'not_submitted',
            admin_note TEXT,
            created_at TEXT NOT NULL,
            reviewed_at TEXT,
            FOREIGN KEY(booking_id) REFERENCES bookings(id)
        );

        CREATE TABLE IF NOT EXISTS tickets (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            booking_id INTEGER UNIQUE NOT NULL,
            ticket_no TEXT UNIQUE NOT NULL,
            pdf_path TEXT NOT NULL,
            created_at TEXT NOT NULL,
            FOREIGN KEY(booking_id) REFERENCES bookings(id)
        );

        CREATE TABLE IF NOT EXISTS wallet_requests (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            amount_uzs INTEGER NOT NULL,
            amount_value REAL,
            currency TEXT NOT NULL DEFAULT 'UZS',
            payment_method TEXT NOT NULL,
            note TEXT,
            status TEXT NOT NULL DEFAULT 'pending',
            admin_note TEXT,
            created_at TEXT NOT NULL,
            reviewed_at TEXT,
            FOREIGN KEY(user_id) REFERENCES users(id)
        );

        CREATE TABLE IF NOT EXISTS reminder_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ticket_no TEXT NOT NULL,
            reminder_type TEXT NOT NULL,
            sent_to_email TEXT NOT NULL,
            created_at TEXT NOT NULL,
            UNIQUE(ticket_no, reminder_type)
        );

        CREATE TABLE IF NOT EXISTS password_reset_tokens (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            token_hash TEXT NOT NULL UNIQUE,
            expires_at TEXT NOT NULL,
            used_at TEXT,
            requested_ip TEXT,
            created_at TEXT NOT NULL,
            FOREIGN KEY(user_id) REFERENCES users(id)
        );

        CREATE TABLE IF NOT EXISTS user_device_sessions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            session_id TEXT NOT NULL UNIQUE,
            device_label TEXT,
            user_agent TEXT,
            ip_address TEXT,
            last_seen_at TEXT NOT NULL,
            created_at TEXT NOT NULL,
            revoked_at TEXT,
            FOREIGN KEY(user_id) REFERENCES users(id)
        );

        CREATE TABLE IF NOT EXISTS security_codes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            purpose TEXT NOT NULL,
            code_hash TEXT NOT NULL,
            expires_at TEXT NOT NULL,
            used_at TEXT,
            requested_ip TEXT,
            created_at TEXT NOT NULL,
            FOREIGN KEY(user_id) REFERENCES users(id)
        );
        """
    )

    admin_email = os.getenv("ADMIN_EMAIL", "admin@safar24.uz")
    admin_password = os.getenv("ADMIN_PASSWORD", "Admin@12345")
    existing_admin = db.execute("SELECT id FROM users WHERE is_admin = 1 LIMIT 1").fetchone()
    if not existing_admin:
        db.execute(
            """
            INSERT INTO users (email, full_name, phone, password_hash, account_id, is_admin, created_at)
            VALUES (?, ?, ?, ?, ?, 1, ?)
            """,
            (
                admin_email,
                "Safar24 Admin",
                "+998917230303",
                generate_password_hash(admin_password),
                "ADM000001",
                now_iso(),
            ),
        )

    has_flights = db.execute("SELECT COUNT(*) AS c FROM flights").fetchone()["c"]
    if has_flights == 0:
        db.executemany(
            """
            INSERT INTO flights
            (route_code, from_city, from_airport_code, to_city, to_airport_code, departure_date, departure_time, return_date, is_oneway, airline, travel_class, price_uzs, price_value, price_currency, baggage_kg, include_meal, include_return, seats, status, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                ("SF24001", "Tashkent", "TAS", "Istanbul", "IST", "2026-05-12", "10:30", "2026-05-20", 0, "Turkish Airlines", "Economy", 4200000, 4200000, "UZS", 20, 1, 1, 7, "active", now_iso()),
                ("SF24002", "Tashkent", "TAS", "Seoul", "ICN", "2026-05-15", "23:45", "2026-05-28", 0, "Asiana", "Economy", 6900000, 6900000, "UZS", 20, 1, 1, 5, "active", now_iso()),
                ("SF24003", "Tashkent", "TAS", "Dubai", "DXB", "2026-05-18", "08:10", "", 1, "Flydubai", "Business", 3600000, 3600000, "UZS", 25, 1, 0, 9, "active", now_iso()),
            ],
        )

    db.commit()


def table_has_column(db: sqlite3.Connection, table_name: str, column_name: str) -> bool:
    cols = db.execute(f"PRAGMA table_info({table_name})").fetchall()
    return any(c["name"] == column_name for c in cols)


def ensure_super_admin_user(db: sqlite3.Connection) -> None:
    super_admin_email = (os.getenv("SUPER_ADMIN_EMAIL", "superadmin@safar24.uz") or "").strip().lower()
    super_admin_password = os.getenv("SUPER_ADMIN_PASSWORD", "SuperAdmin@12345")

    existing_super = db.execute("SELECT id FROM users WHERE is_super_admin = 1 LIMIT 1").fetchone()
    if existing_super:
        return

    by_email = db.execute("SELECT id FROM users WHERE email = ?", (super_admin_email,)).fetchone()
    if by_email:
        db.execute(
            "UPDATE users SET is_admin = 1, is_super_admin = 1 WHERE id = ?",
            (by_email["id"],),
        )
        return

    db.execute(
        """
        INSERT INTO users (email, full_name, phone, password_hash, account_id, is_admin, is_super_admin, created_at)
        VALUES (?, ?, ?, ?, ?, 1, 1, ?)
        """,
        (
            super_admin_email,
            "Safar24 Super Admin",
            "+998917230303",
            generate_password_hash(super_admin_password),
            "SADM000001",
            now_iso(),
        ),
    )


def ensure_schema_updates() -> None:
    db = get_db()
    try:
        db.execute(
            """
            CREATE TABLE IF NOT EXISTS passengers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                booking_id INTEGER NOT NULL,
                full_name TEXT NOT NULL,
                passport_number TEXT,
                passport_series TEXT,
                birth_date TEXT,
                nationality TEXT,
                gender TEXT,
                passport_issue_date TEXT,
                passport_expiration_date TEXT,
                phone TEXT,
                notification_email TEXT,
                travel_class TEXT,
                seat_number TEXT,
                created_at TEXT NOT NULL,
                FOREIGN KEY(booking_id) REFERENCES bookings(id)
            )
            """
        )
    except Exception:
        pass

    db.execute(
        """
        CREATE TABLE IF NOT EXISTS password_reset_tokens (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            token_hash TEXT NOT NULL UNIQUE,
            expires_at TEXT NOT NULL,
            used_at TEXT,
            requested_ip TEXT,
            created_at TEXT NOT NULL,
            FOREIGN KEY(user_id) REFERENCES users(id)
        )
        """
    )

    db.execute(
        """
        CREATE TABLE IF NOT EXISTS user_device_sessions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            session_id TEXT NOT NULL UNIQUE,
            device_label TEXT,
            user_agent TEXT,
            ip_address TEXT,
            last_seen_at TEXT NOT NULL,
            created_at TEXT NOT NULL,
            revoked_at TEXT,
            FOREIGN KEY(user_id) REFERENCES users(id)
        )
        """
    )

    db.execute(
        """
        CREATE TABLE IF NOT EXISTS security_codes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            purpose TEXT NOT NULL,
            code_hash TEXT NOT NULL,
            expires_at TEXT NOT NULL,
            used_at TEXT,
            requested_ip TEXT,
            created_at TEXT NOT NULL,
            FOREIGN KEY(user_id) REFERENCES users(id)
        )
        """
    )

    migration_sql = []
    if not table_has_column(db, "users", "balance_uzs"):
        migration_sql.append("ALTER TABLE users ADD COLUMN balance_uzs INTEGER NOT NULL DEFAULT 0")
    if not table_has_column(db, "users", "balance_usd"):
        migration_sql.append("ALTER TABLE users ADD COLUMN balance_usd REAL NOT NULL DEFAULT 0")
    if not table_has_column(db, "users", "is_super_admin"):
        migration_sql.append("ALTER TABLE users ADD COLUMN is_super_admin INTEGER NOT NULL DEFAULT 0")
    if not table_has_column(db, "users", "role_granted_by"):
        migration_sql.append("ALTER TABLE users ADD COLUMN role_granted_by INTEGER")
    if not table_has_column(db, "users", "role_granted_at"):
        migration_sql.append("ALTER TABLE users ADD COLUMN role_granted_at TEXT")
    if not table_has_column(db, "payments", "payment_method"):
        migration_sql.append("ALTER TABLE payments ADD COLUMN payment_method TEXT NOT NULL DEFAULT 'manual'")
    if not table_has_column(db, "payments", "payment_reference"):
        migration_sql.append("ALTER TABLE payments ADD COLUMN payment_reference TEXT")
    if not table_has_column(db, "payments", "amount_value"):
        migration_sql.append("ALTER TABLE payments ADD COLUMN amount_value REAL")
    if not table_has_column(db, "payments", "currency"):
        migration_sql.append("ALTER TABLE payments ADD COLUMN currency TEXT NOT NULL DEFAULT 'UZS'")
    if not table_has_column(db, "flights", "from_airport_code"):
        migration_sql.append("ALTER TABLE flights ADD COLUMN from_airport_code TEXT")
    if not table_has_column(db, "flights", "to_airport_code"):
        migration_sql.append("ALTER TABLE flights ADD COLUMN to_airport_code TEXT")
    if not table_has_column(db, "flights", "departure_time"):
        migration_sql.append("ALTER TABLE flights ADD COLUMN departure_time TEXT")
    if not table_has_column(db, "flights", "travel_class"):
        migration_sql.append("ALTER TABLE flights ADD COLUMN travel_class TEXT NOT NULL DEFAULT 'Economy'")
    if not table_has_column(db, "flights", "is_oneway"):
        migration_sql.append("ALTER TABLE flights ADD COLUMN is_oneway INTEGER NOT NULL DEFAULT 0")
    if not table_has_column(db, "flights", "baggage_kg"):
        migration_sql.append("ALTER TABLE flights ADD COLUMN baggage_kg INTEGER NOT NULL DEFAULT 20")
    if not table_has_column(db, "flights", "include_meal"):
        migration_sql.append("ALTER TABLE flights ADD COLUMN include_meal INTEGER NOT NULL DEFAULT 1")
    if not table_has_column(db, "flights", "include_return"):
        migration_sql.append("ALTER TABLE flights ADD COLUMN include_return INTEGER NOT NULL DEFAULT 1")
    if not table_has_column(db, "flights", "price_value"):
        migration_sql.append("ALTER TABLE flights ADD COLUMN price_value REAL")
    if not table_has_column(db, "flights", "price_currency"):
        migration_sql.append("ALTER TABLE flights ADD COLUMN price_currency TEXT NOT NULL DEFAULT 'UZS'")
    if not table_has_column(db, "bookings", "baggage_option"):
        migration_sql.append("ALTER TABLE bookings ADD COLUMN baggage_option TEXT NOT NULL DEFAULT 'standard'")
    if not table_has_column(db, "bookings", "passenger_count"):
        migration_sql.append("ALTER TABLE bookings ADD COLUMN passenger_count INTEGER NOT NULL DEFAULT 1")
    if not table_has_column(db, "bookings", "adult_count"):
        migration_sql.append("ALTER TABLE bookings ADD COLUMN adult_count INTEGER NOT NULL DEFAULT 1")
    if not table_has_column(db, "bookings", "child_count"):
        migration_sql.append("ALTER TABLE bookings ADD COLUMN child_count INTEGER NOT NULL DEFAULT 0")
    if not table_has_column(db, "bookings", "infant_count"):
        migration_sql.append("ALTER TABLE bookings ADD COLUMN infant_count INTEGER NOT NULL DEFAULT 0")
    if not table_has_column(db, "bookings", "seat_count"):
        migration_sql.append("ALTER TABLE bookings ADD COLUMN seat_count INTEGER NOT NULL DEFAULT 1")
    if not table_has_column(db, "bookings", "reservation_expires_at"):
        migration_sql.append("ALTER TABLE bookings ADD COLUMN reservation_expires_at TEXT")
    if not table_has_column(db, "bookings", "cancelled_at"):
        migration_sql.append("ALTER TABLE bookings ADD COLUMN cancelled_at TEXT")
    if not table_has_column(db, "bookings", "expired_at"):
        migration_sql.append("ALTER TABLE bookings ADD COLUMN expired_at TEXT")
    if not table_has_column(db, "bookings", "cancellation_reason"):
        migration_sql.append("ALTER TABLE bookings ADD COLUMN cancellation_reason TEXT")
    if not table_has_column(db, "user_profiles", "passport_series"):
        migration_sql.append("ALTER TABLE user_profiles ADD COLUMN passport_series TEXT")
    if not table_has_column(db, "user_profiles", "passport_issue_date"):
        migration_sql.append("ALTER TABLE user_profiles ADD COLUMN passport_issue_date TEXT")
    if not table_has_column(db, "user_profiles", "passport_expiration_date"):
        migration_sql.append("ALTER TABLE user_profiles ADD COLUMN passport_expiration_date TEXT")
    if not table_has_column(db, "user_profiles", "gender"):
        migration_sql.append("ALTER TABLE user_profiles ADD COLUMN gender TEXT")
    if not table_has_column(db, "user_profiles", "notification_email"):
        migration_sql.append("ALTER TABLE user_profiles ADD COLUMN notification_email TEXT")
    if not table_has_column(db, "passengers", "passport_number"):
        migration_sql.append("ALTER TABLE passengers ADD COLUMN passport_number TEXT")
    if not table_has_column(db, "passengers", "first_name"):
        migration_sql.append("ALTER TABLE passengers ADD COLUMN first_name TEXT")
    if not table_has_column(db, "passengers", "last_name"):
        migration_sql.append("ALTER TABLE passengers ADD COLUMN last_name TEXT")
    if not table_has_column(db, "passengers", "passenger_type"):
        migration_sql.append("ALTER TABLE passengers ADD COLUMN passenger_type TEXT")
    if not table_has_column(db, "passengers", "passport_series"):
        migration_sql.append("ALTER TABLE passengers ADD COLUMN passport_series TEXT")
    if not table_has_column(db, "passengers", "birth_date"):
        migration_sql.append("ALTER TABLE passengers ADD COLUMN birth_date TEXT")
    if not table_has_column(db, "passengers", "nationality"):
        migration_sql.append("ALTER TABLE passengers ADD COLUMN nationality TEXT")
    if not table_has_column(db, "passengers", "gender"):
        migration_sql.append("ALTER TABLE passengers ADD COLUMN gender TEXT")
    if not table_has_column(db, "passengers", "passport_issue_date"):
        migration_sql.append("ALTER TABLE passengers ADD COLUMN passport_issue_date TEXT")
    if not table_has_column(db, "passengers", "passport_expiration_date"):
        migration_sql.append("ALTER TABLE passengers ADD COLUMN passport_expiration_date TEXT")
    if not table_has_column(db, "passengers", "phone"):
        migration_sql.append("ALTER TABLE passengers ADD COLUMN phone TEXT")
    if not table_has_column(db, "passengers", "notification_email"):
        migration_sql.append("ALTER TABLE passengers ADD COLUMN notification_email TEXT")
    if not table_has_column(db, "wallet_requests", "currency"):
        migration_sql.append("ALTER TABLE wallet_requests ADD COLUMN currency TEXT NOT NULL DEFAULT 'UZS'")
    if not table_has_column(db, "wallet_requests", "amount_value"):
        migration_sql.append("ALTER TABLE wallet_requests ADD COLUMN amount_value REAL")
    if not table_has_column(db, "payments", "base_amount_uzs"):
        migration_sql.append("ALTER TABLE payments ADD COLUMN base_amount_uzs INTEGER")
    if not table_has_column(db, "payments", "base_amount_value"):
        migration_sql.append("ALTER TABLE payments ADD COLUMN base_amount_value REAL")
    if not table_has_column(db, "payments", "method_fee_pct"):
        migration_sql.append("ALTER TABLE payments ADD COLUMN method_fee_pct REAL NOT NULL DEFAULT 0")
    if not table_has_column(db, "payments", "method_fee_uzs"):
        migration_sql.append("ALTER TABLE payments ADD COLUMN method_fee_uzs INTEGER NOT NULL DEFAULT 0")
    if not table_has_column(db, "payments", "method_fee_value"):
        migration_sql.append("ALTER TABLE payments ADD COLUMN method_fee_value REAL NOT NULL DEFAULT 0")

    for stmt in migration_sql:
        try:
            db.execute(stmt)
        except sqlite3.OperationalError as ex:
            if "duplicate column name" not in str(ex).lower():
                raise
    if migration_sql:
        flights = db.execute("SELECT id, from_city, to_city FROM flights").fetchall()
        for flight in flights:
            from_code = city_to_airport_code(flight["from_city"])
            to_code = city_to_airport_code(flight["to_city"])
            db.execute(
                """
                UPDATE flights
                SET from_airport_code = COALESCE(NULLIF(from_airport_code, ''), ?),
                    to_airport_code = COALESCE(NULLIF(to_airport_code, ''), ?)
                WHERE id = ?
                """,
                (from_code, to_code, flight["id"]),
            )
        db.commit()

    db.execute("UPDATE flights SET departure_time = COALESCE(departure_time, '')")
    db.execute("UPDATE flights SET travel_class = COALESCE(NULLIF(travel_class, ''), 'Economy')")
    db.execute("UPDATE flights SET is_oneway = COALESCE(is_oneway, CASE WHEN return_date IS NOT NULL AND return_date != '' THEN 0 ELSE 1 END)")
    db.execute("UPDATE flights SET baggage_kg = COALESCE(NULLIF(baggage_kg, 0), 20)")
    db.execute("UPDATE flights SET include_meal = COALESCE(include_meal, 1)")
    db.execute("UPDATE flights SET include_return = COALESCE(include_return, 1)")
    db.execute("UPDATE flights SET price_currency = COALESCE(NULLIF(upper(price_currency), ''), 'UZS')")
    db.execute(
        """
        UPDATE flights
        SET price_value = COALESCE(
            price_value,
            CASE
                WHEN upper(price_currency) = 'USD' THEN ROUND(CAST(price_uzs AS REAL) / ?, 2)
                ELSE CAST(price_uzs AS REAL)
            END
        )
        """,
        (USD_TO_UZS_RATE,),
    )
    db.execute("UPDATE bookings SET baggage_option = COALESCE(NULLIF(baggage_option, ''), 'standard')")
    db.execute("UPDATE bookings SET passenger_count = COALESCE(NULLIF(passenger_count, 0), 1)")
    db.execute("UPDATE bookings SET adult_count = COALESCE(NULLIF(adult_count, 0), passenger_count)")
    db.execute("UPDATE bookings SET child_count = COALESCE(child_count, 0)")
    db.execute("UPDATE bookings SET infant_count = COALESCE(infant_count, 0)")
    db.execute(
        """
        UPDATE bookings
        SET seat_count = COALESCE(
            NULLIF(seat_count, 0),
            CASE
                WHEN (adult_count + child_count) > 0 THEN (adult_count + child_count)
                ELSE passenger_count
            END
        )
        """
    )
    db.execute("UPDATE payments SET currency = COALESCE(NULLIF(upper(currency), ''), 'UZS')")
    db.execute("UPDATE payments SET amount_value = COALESCE(amount_value, amount_uzs)")
    db.execute("UPDATE payments SET base_amount_value = COALESCE(base_amount_value, amount_value, amount_uzs)")
    db.execute("UPDATE payments SET base_amount_uzs = COALESCE(base_amount_uzs, amount_uzs)")
    db.execute("UPDATE payments SET method_fee_pct = COALESCE(method_fee_pct, 0)")
    db.execute("UPDATE payments SET method_fee_uzs = COALESCE(method_fee_uzs, 0)")
    db.execute("UPDATE payments SET method_fee_value = COALESCE(method_fee_value, 0)")
    db.execute("UPDATE users SET balance_usd = COALESCE(balance_usd, 0)")
    db.execute("UPDATE users SET is_super_admin = COALESCE(is_super_admin, 0)")
    db.execute(
        "UPDATE users SET role_granted_by = NULL, role_granted_at = NULL WHERE COALESCE(is_admin, 0) = 0"
    )
    db.execute("UPDATE users SET role_granted_by = NULL WHERE role_granted_by = id")
    db.execute("UPDATE wallet_requests SET currency = COALESCE(NULLIF(currency, ''), 'UZS')")
    db.execute("UPDATE wallet_requests SET amount_value = COALESCE(amount_value, amount_uzs)")
    db.execute(
        """
        DELETE FROM user_device_sessions
        WHERE revoked_at IS NOT NULL
          AND COALESCE(last_seen_at, created_at, '') < ?
        """,
        ((datetime.now(timezone.utc) - timedelta(days=AUTH_SESSION_MAX_AGE_DAYS)).replace(microsecond=0).isoformat(),),
    )
    db.execute(
        """
        DELETE FROM security_codes
        WHERE used_at IS NOT NULL
           OR expires_at <= ?
        """,
        (now_iso(),),
    )
    sync_passenger_name_and_type_columns(db)
    ensure_super_admin_user(db)

    # Normalize long account IDs to short A0001 style for non-admin users.
    users_for_account = db.execute("SELECT id, account_id, is_admin FROM users ORDER BY id ASC").fetchall()
    for user_row in users_for_account:
        if int(user_row["is_admin"]) == 1:
            continue
        normalized_id = generate_account_id(int(user_row["id"]))
        if user_row["account_id"] != normalized_id:
            db.execute("UPDATE users SET account_id = ? WHERE id = ?", (normalized_id, user_row["id"]))
    db.commit()


def now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def parse_iso_datetime(value: str | None) -> datetime | None:
    raw = (value or "").strip()
    if not raw:
        return None
    if raw.endswith("Z"):
        raw = f"{raw[:-1]}+00:00"
    try:
        parsed = datetime.fromisoformat(raw)
    except ValueError:
        return None
    if parsed.tzinfo is None:
        return parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(timezone.utc)


def booking_hold_expires_at(booking: sqlite3.Row | None) -> datetime | None:
    if not booking:
        return None
    return parse_iso_datetime(str(booking["reservation_expires_at"] or ""))


def booking_hold_seconds_left(booking: sqlite3.Row | None) -> int | None:
    expires_at = booking_hold_expires_at(booking)
    if not expires_at:
        return None
    remaining = int((expires_at - datetime.now(timezone.utc)).total_seconds())
    return max(0, remaining)


def booking_hold_is_active(booking: sqlite3.Row | None) -> bool:
    if not booking:
        return False
    if booking["status"] not in {"pending_payment", "payment_review", "payment_rejected"}:
        return False
    expires_at = booking_hold_expires_at(booking)
    if not expires_at:
        return True
    return expires_at > datetime.now(timezone.utc)


def release_booking_seats(
    db: sqlite3.Connection,
    booking: sqlite3.Row,
    final_status: str,
    timestamp_field: str | None = None,
    cancellation_reason: str | None = None,
) -> None:
    seats_to_release = int(booking["seat_count"] or booking["passenger_count"] or 1)
    db.execute("UPDATE flights SET seats = seats + ? WHERE id = ?", (seats_to_release, booking["flight_id"]))

    assignment_sql = ["status = ?"]
    assignment_values: list[object] = [final_status]
    if timestamp_field:
        assignment_sql.append(f"{timestamp_field} = ?")
        assignment_values.append(now_iso())
    if cancellation_reason is not None:
        assignment_sql.append("cancellation_reason = ?")
        assignment_values.append(cancellation_reason)
    assignment_values.append(booking["id"])

    db.execute(
        f"UPDATE bookings SET {', '.join(assignment_sql)} WHERE id = ?",
        assignment_values,
    )


def cleanup_expired_booking_holds(db: sqlite3.Connection) -> int:
    expired_rows = db.execute(
        """
        SELECT b.*
        FROM bookings b
        WHERE b.status IN ('pending_payment', 'payment_review', 'payment_rejected')
          AND b.reservation_expires_at IS NOT NULL
          AND b.reservation_expires_at <= ?
        ORDER BY b.id ASC
        """,
        (now_iso(),),
    ).fetchall()

    released_count = 0
    for booking in expired_rows:
        release_booking_seats(db, booking, "expired", "expired_at")
        released_count += 1
    return released_count


def expire_booking_hold_if_needed(db: sqlite3.Connection, booking: sqlite3.Row | None) -> bool:
    if not booking:
        return False
    if not booking_hold_is_active(booking):
        if booking["status"] in {"pending_payment", "payment_review", "payment_rejected"} and booking_hold_expires_at(booking):
            release_booking_seats(db, booking, "expired", "expired_at")
        return True
    return False


def now_ts() -> float:
    return datetime.now(timezone.utc).timestamp()


def normalize_login_buckets() -> None:
    cutoff = now_ts() - LOGIN_WINDOW_SECONDS
    to_remove = []
    for key, timestamps in LOGIN_ATTEMPTS.items():
        recent = [t for t in timestamps if t >= cutoff]
        if recent:
            LOGIN_ATTEMPTS[key] = recent
        else:
            to_remove.append(key)
    for key in to_remove:
        LOGIN_ATTEMPTS.pop(key, None)


def register_failed_login(bucket: str) -> None:
    normalize_login_buckets()
    LOGIN_ATTEMPTS.setdefault(bucket, []).append(now_ts())


def clear_failed_logins(bucket: str) -> None:
    LOGIN_ATTEMPTS.pop(bucket, None)


def is_login_rate_limited(bucket: str) -> bool:
    normalize_login_buckets()
    return len(LOGIN_ATTEMPTS.get(bucket, [])) >= LOGIN_MAX_ATTEMPTS


def normalize_password_reset_buckets() -> None:
    cutoff = now_ts() - PASSWORD_RESET_WINDOW_SECONDS
    to_remove = []
    for key, timestamps in PASSWORD_RESET_ATTEMPTS.items():
        recent = [t for t in timestamps if t >= cutoff]
        if recent:
            PASSWORD_RESET_ATTEMPTS[key] = recent
        else:
            to_remove.append(key)
    for key in to_remove:
        PASSWORD_RESET_ATTEMPTS.pop(key, None)


def register_password_reset_attempt(bucket: str) -> None:
    normalize_password_reset_buckets()
    PASSWORD_RESET_ATTEMPTS.setdefault(bucket, []).append(now_ts())


def is_password_reset_rate_limited(bucket: str) -> bool:
    normalize_password_reset_buckets()
    return len(PASSWORD_RESET_ATTEMPTS.get(bucket, [])) >= PASSWORD_RESET_MAX_ATTEMPTS


def client_bucket(email: str) -> str:
    ip = request.remote_addr or "unknown"
    ua = request.headers.get("User-Agent", "na")[:80]
    return f"{ip}|{ua}|{(email or '').strip().lower()}"


def password_reset_bucket(email: str) -> str:
    ip = request.remote_addr or "unknown"
    normalized_email = (email or "anonymous").strip().lower()
    return f"{ip}|{normalized_email}"


def is_valid_email(value: str) -> bool:
    return bool(EMAIL_RE.match(value))


def is_valid_gmail(value: str) -> bool:
    value = (value or "").strip().lower()
    return bool(value) and is_valid_email(value) and value.endswith("@gmail.com")


def is_valid_latin_search_term(value: str) -> bool:
    term = (value or "").strip()
    return bool(term) and bool(LATIN_SEARCH_RE.fullmatch(term))


def is_safe_internal_path(target: str) -> bool:
    parsed = urlparse(target or "")
    return bool(target) and target.startswith("/") and not target.startswith("//") and not parsed.scheme and not parsed.netloc


def parse_date_ymd(value: str) -> datetime.date | None:
    raw = (value or "").strip()
    if not raw:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%d.%m.%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return None


def parse_time_hhmm(value: str | None) -> tuple[int, int] | None:
    raw = (value or "").strip()
    if not raw:
        return None

    match = re.match(r"^(\d{1,2}):(\d{2})", raw)
    if not match:
        return None

    hour = int(match.group(1))
    minute = int(match.group(2))
    if 0 <= hour <= 23 and 0 <= minute <= 59:
        return hour, minute
    return None


def parse_flight_departure_datetime(departure_date: str | None, departure_time: str | None = None) -> datetime | None:
    raw_departure = (departure_date or "").strip()
    if not raw_departure:
        return None

    date_token = raw_departure
    inferred_time = ""
    if "T" in raw_departure:
        date_token, inferred_time = raw_departure.split("T", 1)
    elif " " in raw_departure:
        date_token, inferred_time = raw_departure.split(" ", 1)

    dep_date = parse_date_ymd(date_token)
    if not dep_date:
        return None

    time_candidate = (departure_time or "").strip() or inferred_time.strip() or "00:00"
    parsed_time = parse_time_hhmm(time_candidate) or (0, 0)

    return datetime.combine(dep_date, datetime.min.time()).replace(hour=parsed_time[0], minute=parsed_time[1])


def is_flight_upcoming(departure_date: str | None, departure_time: str | None = None, reference_dt: datetime | None = None) -> bool:
    dep_dt = parse_flight_departure_datetime(departure_date, departure_time)
    if dep_dt is None:
        return True
    now_dt = reference_dt or datetime.now()
    return dep_dt >= now_dt


def normalize_date_ymd(value: str) -> str:
    dt = parse_date_ymd(value)
    return dt.strftime("%Y-%m-%d") if dt else (value or "").strip()


def split_name_parts(full_name: str) -> tuple[str, str]:
    normalized = " ".join((full_name or "").split())
    if not normalized:
        return "", ""
    parts = normalized.split(" ")
    if len(parts) == 1:
        return parts[0], ""
    return " ".join(parts[:-1]), parts[-1]


def build_full_name(first_name: str, last_name: str) -> str:
    return " ".join(part for part in [first_name.strip(), last_name.strip()] if part).strip()


def passenger_name_parts(row: sqlite3.Row | dict[str, object]) -> tuple[str, str]:
    first_name = str(row["first_name"] or "").strip()
    last_name = str(row["last_name"] or "").strip()
    if first_name and last_name:
        return first_name, last_name
    derived_first, derived_last = split_name_parts(str(row["full_name"] or ""))
    return first_name or derived_first, last_name or derived_last


def calculate_age_years(birth_date: str | None, reference_date: str | None = None) -> int | None:
    dob = parse_date_ymd(birth_date or "")
    if not dob:
        return None

    ref_date = parse_date_ymd(reference_date or "") if reference_date else None
    if not ref_date:
        ref_date = datetime.now(timezone.utc).date()

    years = ref_date.year - dob.year
    if (ref_date.month, ref_date.day) < (dob.month, dob.day):
        years -= 1
    return years


def passenger_type_from_birth_date(birth_date: str | None, reference_date: str | None = None) -> str:
    age_years = calculate_age_years(birth_date, reference_date)
    if age_years is None:
        return ""
    if age_years < 2:
        return "INF"
    if age_years < 12:
        return "CHD"
    return "ADT"


def normalize_passenger_type(passenger_type: str | None) -> str:
    code = (passenger_type or "").strip().upper()
    return code if code in PASSENGER_TYPE_LABELS else ""


def passenger_type_label(passenger_type: str | None) -> str:
    code = normalize_passenger_type(passenger_type)
    return PASSENGER_TYPE_LABELS.get(code, "-")


def normalize_currency_code(currency: str | None) -> str:
    code = (currency or "UZS").strip().upper()
    return "USD" if code == "USD" else "UZS"


def round_currency_amount(amount: float | int, currency: str | None) -> float:
    code = normalize_currency_code(currency)
    value = float(amount or 0)
    if code == "USD":
        return round(value, 2)
    return float(int(round(value)))


def fare_ratio_for_passenger_type(passenger_type: str | None) -> float:
    code = normalize_passenger_type(passenger_type) or "ADT"
    raw_ratio = PASSENGER_FARE_RATIO.get(code, 1.0)
    try:
        ratio = float(raw_ratio)
    except (TypeError, ValueError):
        ratio = 1.0
    return max(0.0, min(ratio, 2.0))


def payment_fee_percent(payment_method: str | None) -> float:
    method = (payment_method or "wallet").strip().lower()
    raw_percent = PAYMENT_METHOD_FEE_PERCENT.get(method, 0.0)
    try:
        percent = float(raw_percent)
    except (TypeError, ValueError):
        percent = 0.0
    return max(0.0, min(percent, 25.0))


def calculate_booking_fare_components(
    base_price_value: float,
    currency: str,
    baggage_option: str,
    adult_count: int,
    child_count: int,
    infant_count: int,
) -> dict[str, float]:
    code = normalize_currency_code(currency)
    base_unit_value = round_currency_amount(base_price_value * baggage_option_multiplier(baggage_option), code)
    adult_unit_value = round_currency_amount(base_unit_value * fare_ratio_for_passenger_type("ADT"), code)
    child_unit_value = round_currency_amount(base_unit_value * fare_ratio_for_passenger_type("CHD"), code)
    infant_unit_value = round_currency_amount(base_unit_value * fare_ratio_for_passenger_type("INF"), code)
    base_total_value = round_currency_amount(
        (adult_unit_value * adult_count) + (child_unit_value * child_count) + (infant_unit_value * infant_count),
        code,
    )

    return {
        "base_unit_value": base_unit_value,
        "adult_unit_value": adult_unit_value,
        "child_unit_value": child_unit_value,
        "infant_unit_value": infant_unit_value,
        "base_total_value": base_total_value,
    }


def calculate_payment_breakdown(base_amount_value: float, currency: str, payment_method: str) -> dict[str, float | int]:
    code = normalize_currency_code(currency)
    safe_base_value = round_currency_amount(base_amount_value, code)
    fee_pct = payment_fee_percent(payment_method)
    fee_value = round_currency_amount((safe_base_value * fee_pct) / 100.0, code)
    final_value = round_currency_amount(safe_base_value + fee_value, code)

    base_uzs = int(round(convert_currency(safe_base_value, code, "UZS")))
    fee_uzs = int(round(convert_currency(fee_value, code, "UZS")))
    final_uzs = int(round(convert_currency(final_value, code, "UZS")))

    return {
        "currency": code,
        "base_amount_value": safe_base_value,
        "base_amount_uzs": base_uzs,
        "method_fee_pct": fee_pct,
        "method_fee_value": fee_value,
        "method_fee_uzs": fee_uzs,
        "amount_value": final_value,
        "amount_uzs": final_uzs,
    }


def booking_age_counts(booking: sqlite3.Row | None, passenger_rows: list[sqlite3.Row] | None = None) -> tuple[int, int, int]:
    if booking:
        adult_count = int(booking["adult_count"] or 0)
        child_count = int(booking["child_count"] or 0)
        infant_count = int(booking["infant_count"] or 0)
    else:
        adult_count = 0
        child_count = 0
        infant_count = 0

    if passenger_rows and (adult_count + child_count + infant_count) == 0:
        counts = {"ADT": 0, "CHD": 0, "INF": 0}
        for row in passenger_rows:
            code = passenger_type_from_birth_date(str(row["birth_date"] or "")) or normalize_passenger_type(row["passenger_type"])
            counts[code if code in counts else "ADT"] += 1
        adult_count = counts["ADT"]
        child_count = counts["CHD"]
        infant_count = counts["INF"]

    if booking and (adult_count + child_count + infant_count) == 0:
        adult_count = int(booking["passenger_count"] or 1)

    return adult_count, child_count, infant_count


def passenger_mix_label(adult_count: int, child_count: int, infant_count: int) -> str:
    return f"ADT {adult_count} • CHD {child_count} • INF {infant_count}"


def validate_password_strength(password: str, current_password_hash: str | None = None) -> list[str]:
    errors: list[str] = []
    raw = password or ""

    if len(raw) < PASSWORD_MIN_LENGTH:
        errors.append(f"Parol kamida {PASSWORD_MIN_LENGTH} ta belgidan iborat bo'lsin.")
    if re.search(r"\s", raw):
        errors.append("Parolda bo'sh joy ishlatmang.")
    if not re.search(r"[A-Z]", raw):
        errors.append("Parolda kamida bitta katta harf bo'lishi kerak.")
    if not re.search(r"[a-z]", raw):
        errors.append("Parolda kamida bitta kichik harf bo'lishi kerak.")
    if not re.search(r"\d", raw):
        errors.append("Parolda kamida bitta raqam bo'lishi kerak.")
    if not re.search(r"[^A-Za-z0-9]", raw):
        errors.append("Parolda kamida bitta maxsus belgi bo'lishi kerak.")

    if current_password_hash and raw and check_password_hash(current_password_hash, raw):
        errors.append("Yangi parol joriy parol bilan bir xil bo'lmasligi kerak.")

    return errors


def hash_secure_token(raw_token: str) -> str:
    return hashlib.sha256((raw_token or "").encode("utf-8")).hexdigest()


def build_password_reset_token() -> str:
    return secrets.token_urlsafe(32)


def build_external_url(endpoint: str, **values) -> str:
    if APP_BASE_URL:
        return f"{APP_BASE_URL}{url_for(endpoint, _external=False, **values)}"
    if has_request_context():
        return url_for(endpoint, _external=True, **values)
    return url_for(endpoint, _external=False, **values)


def send_password_reset_email(to_email: str, full_name: str, reset_url: str) -> tuple[bool, str]:
    return send_plain_email(
        to_email,
        "Safar24 parolni tiklash",
        (
            f"Assalomu alaykum, {full_name}!\n\n"
            "Parolni tiklash so'rovi qabul qilindi.\n"
            f"Quyidagi havola orqali yangi parol o'rnating ({PASSWORD_RESET_TOKEN_MINUTES} daqiqa amal qiladi):\n"
            f"{reset_url}\n\n"
            "Agar bu so'rovni siz yubormagan bo'lsangiz, ushbu xatni e'tiborsiz qoldiring.\n\n"
            "Safar24"
        ),
    )


def current_request_ip() -> str:
    forwarded = (request.headers.get("X-Forwarded-For") or "").split(",")[0].strip()
    return forwarded or (request.remote_addr or "unknown")


def describe_device(user_agent: str) -> str:
    raw = (user_agent or "").strip().lower()
    os_label = "Unknown OS"
    browser_label = "Unknown browser"

    if "iphone" in raw or "ipad" in raw:
        os_label = "iOS"
    elif "android" in raw:
        os_label = "Android"
    elif "windows" in raw:
        os_label = "Windows"
    elif "mac os" in raw or "macintosh" in raw:
        os_label = "macOS"
    elif "linux" in raw:
        os_label = "Linux"

    if "edg/" in raw:
        browser_label = "Edge"
    elif "opr/" in raw or "opera" in raw:
        browser_label = "Opera"
    elif "chrome/" in raw and "chromium" not in raw and "edg/" not in raw:
        browser_label = "Chrome"
    elif "safari/" in raw and "chrome/" not in raw:
        browser_label = "Safari"
    elif "firefox/" in raw:
        browser_label = "Firefox"

    return f"{os_label} / {browser_label}"


def security_code_hash(user_id: int, purpose: str, code: str) -> str:
    return hash_secure_token(f"{purpose}:{user_id}:{(code or '').strip()}")


def issue_security_code(
    db: sqlite3.Connection,
    user_id: int,
    purpose: str,
    requested_ip: str | None = None,
) -> str:
    now = now_iso()
    expires_at = (datetime.now(timezone.utc) + timedelta(minutes=SECURITY_CODE_MINUTES)).replace(microsecond=0).isoformat()

    db.execute(
        """
        UPDATE security_codes
        SET used_at = ?
        WHERE user_id = ? AND purpose = ? AND used_at IS NULL
        """,
        (now, user_id, purpose),
    )

    for _ in range(6):
        code = build_registration_code()
        db.execute(
            """
            INSERT INTO security_codes (user_id, purpose, code_hash, expires_at, used_at, requested_ip, created_at)
            VALUES (?, ?, ?, ?, NULL, ?, ?)
            """,
            (user_id, purpose, security_code_hash(user_id, purpose, code), expires_at, requested_ip or "", now),
        )
        return code

    raise ValueError("Tasdiqlash kodi yaratilmadi.")


def consume_security_code(db: sqlite3.Connection, user_id: int, purpose: str, code: str) -> sqlite3.Row | None:
    code_hash = security_code_hash(user_id, purpose, code)
    row = db.execute(
        """
        SELECT *
        FROM security_codes
        WHERE user_id = ?
          AND purpose = ?
          AND code_hash = ?
          AND used_at IS NULL
          AND expires_at > ?
        ORDER BY id DESC
        LIMIT 1
        """,
        (user_id, purpose, code_hash, now_iso()),
    ).fetchone()
    if not row:
        return None

    db.execute("UPDATE security_codes SET used_at = ? WHERE id = ?", (now_iso(), row["id"]))
    return row


def send_password_change_code_email(to_email: str, full_name: str, code: str) -> tuple[bool, str]:
    return send_plain_email(
        to_email,
        "Safar24 xavfsizlik kodi",
        (
            f"Assalomu alaykum, {full_name}!\n\n"
            f"Parolni o'zgartirish uchun tasdiqlash kodi: {code}\n"
            f"Kod {SECURITY_CODE_MINUTES} daqiqa amal qiladi.\n\n"
            "Agar bu amalni siz bajarmagan bo'lsangiz, ushbu xatni e'tiborsiz qoldiring.\n\n"
            "Safar24"
        ),
    )


def create_user_device_session(db: sqlite3.Connection, user_id: int) -> str:
    now = now_iso()
    user_agent = (request.headers.get("User-Agent") or "")[:300]
    device_label = describe_device(user_agent)
    ip_address = current_request_ip()

    for _ in range(6):
        session_id = secrets.token_urlsafe(24)
        try:
            db.execute(
                """
                INSERT INTO user_device_sessions (
                    user_id,
                    session_id,
                    device_label,
                    user_agent,
                    ip_address,
                    last_seen_at,
                    created_at,
                    revoked_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, NULL)
                """,
                (user_id, session_id, device_label, user_agent, ip_address, now, now),
            )
            return session_id
        except sqlite3.IntegrityError:
            continue

    raise ValueError("Qurilma sessiyasi yaratilmagan.")


def ensure_active_user_device_session(db: sqlite3.Connection, user_id: int) -> bool:
    session_id = str(session.get("_auth_session_id") or "").strip()
    if not session_id:
        session["_auth_session_id"] = create_user_device_session(db, user_id)
        return True

    row = db.execute(
        """
        SELECT id, last_seen_at
        FROM user_device_sessions
        WHERE user_id = ? AND session_id = ? AND revoked_at IS NULL
        LIMIT 1
        """,
        (user_id, session_id),
    ).fetchone()
    if not row:
        return False

    last_seen = parse_iso_datetime(str(row["last_seen_at"] or ""))
    if last_seen and last_seen < (datetime.now(timezone.utc) - timedelta(days=AUTH_SESSION_MAX_AGE_DAYS)):
        db.execute(
            "UPDATE user_device_sessions SET revoked_at = ? WHERE id = ?",
            (now_iso(), row["id"]),
        )
        return False

    user_agent = (request.headers.get("User-Agent") or "")[:300]
    db.execute(
        """
        UPDATE user_device_sessions
        SET last_seen_at = ?,
            ip_address = ?,
            user_agent = ?,
            device_label = COALESCE(NULLIF(device_label, ''), ?)
        WHERE id = ?
        """,
        (now_iso(), current_request_ip(), user_agent, describe_device(user_agent), row["id"]),
    )
    return True


def revoke_other_user_device_sessions(db: sqlite3.Connection, user_id: int, keep_session_id: str) -> int:
    cursor = db.execute(
        """
        UPDATE user_device_sessions
        SET revoked_at = ?
        WHERE user_id = ?
          AND session_id != ?
          AND revoked_at IS NULL
        """,
        (now_iso(), user_id, keep_session_id),
    )
    return int(cursor.rowcount or 0)


def issue_password_reset_token(db: sqlite3.Connection, user_id: int, requested_ip: str | None = None) -> str:
    now = now_iso()
    expires_at = (datetime.now(timezone.utc) + timedelta(minutes=PASSWORD_RESET_TOKEN_MINUTES)).replace(microsecond=0).isoformat()

    db.execute(
        """
        UPDATE password_reset_tokens
        SET used_at = ?
        WHERE user_id = ? AND used_at IS NULL
        """,
        (now, user_id),
    )

    for _ in range(5):
        raw_token = build_password_reset_token()
        token_hash = hash_secure_token(raw_token)
        try:
            db.execute(
                """
                INSERT INTO password_reset_tokens (user_id, token_hash, expires_at, used_at, requested_ip, created_at)
                VALUES (?, ?, ?, NULL, ?, ?)
                """,
                (user_id, token_hash, expires_at, requested_ip or "", now),
            )
            return raw_token
        except sqlite3.IntegrityError:
            continue
    raise ValueError("Reset token yaratib bo'lmadi.")


def get_active_password_reset_token(db: sqlite3.Connection, raw_token: str) -> sqlite3.Row | None:
    token_hash = hash_secure_token(raw_token)
    return db.execute(
        """
        SELECT prt.*, u.id AS user_id, u.email, u.full_name, u.password_hash
        FROM password_reset_tokens prt
        JOIN users u ON u.id = prt.user_id
        WHERE prt.token_hash = ?
          AND prt.used_at IS NULL
          AND prt.expires_at > ?
        LIMIT 1
        """,
        (token_hash, now_iso()),
    ).fetchone()


def consume_password_reset_token(db: sqlite3.Connection, raw_token: str) -> sqlite3.Row | None:
    token_row = get_active_password_reset_token(db, raw_token)

    if not token_row:
        return None

    db.execute("UPDATE password_reset_tokens SET used_at = ? WHERE id = ?", (now_iso(), token_row["id"]))
    return token_row


def sync_passenger_name_and_type_columns(db: sqlite3.Connection) -> None:
    passenger_rows = db.execute(
        """
        SELECT p.id, p.full_name, p.first_name, p.last_name, p.birth_date, p.passenger_type, f.departure_date
        FROM passengers p
        JOIN bookings b ON b.id = p.booking_id
        JOIN flights f ON f.id = b.flight_id
        ORDER BY p.id ASC
        """
    ).fetchall()

    for row in passenger_rows:
        first_name, last_name = passenger_name_parts(row)
        reference_date = str(row["departure_date"] or "")
        computed_type = passenger_type_from_birth_date(str(row["birth_date"] or ""), reference_date)
        stored_type = normalize_passenger_type(row["passenger_type"])
        passenger_type = computed_type or stored_type or "ADT"
        full_name = build_full_name(first_name, last_name) or str(row["full_name"] or "").strip()
        db.execute(
            """
            UPDATE passengers
            SET first_name = ?,
                last_name = ?,
                passenger_type = ?,
                full_name = ?
            WHERE id = ?
            """,
            (first_name, last_name, passenger_type, full_name, row["id"]),
        )


def validate_name_parts(first_name: str, last_name: str) -> list[str]:
    if len(first_name.strip()) < 2 or len(last_name.strip()) < 2:
        return ["Ism va familiyani alohida kiriting."]
    return []


def estimate_flight_duration_minutes(from_code: str, to_code: str) -> int | None:
    source = AIRPORTS.get((from_code or "").upper())
    target = AIRPORTS.get((to_code or "").upper())
    if not source or not target:
        return None

    lat1 = math.radians(float(source["lat"]))
    lon1 = math.radians(float(source["lon"]))
    lat2 = math.radians(float(target["lat"]))
    lon2 = math.radians(float(target["lon"]))
    dlat = lat2 - lat1
    dlon = lon2 - lon1
    a = math.sin(dlat / 2) ** 2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon / 2) ** 2
    distance_km = 6371.0 * 2 * math.asin(min(1.0, math.sqrt(a)))
    estimated = max(55, int(round((distance_km / 820.0) * 60 + 35)))
    return int(round(estimated / 5.0) * 5)


def format_duration_label(total_minutes: int | None) -> str:
    if not total_minutes:
        return "Direct"
    hours, minutes = divmod(total_minutes, 60)
    if not hours:
        return f"{minutes} min"
    if not minutes:
        return f"{hours}h"
    return f"{hours}h {minutes}m"


def validate_profile_payload(
    phone: str,
    passport_number: str,
    passport_series: str,
    birth_date: str,
    nationality: str,
    passport_issue_date: str,
    passport_expiration_date: str,
    gender: str,
    notification_email: str,
) -> list[str]:
    errors: list[str] = []
    today = datetime.now(timezone.utc).date()

    if not PHONE_RE.match(phone):
        errors.append("Telefon raqam formati noto'g'ri. Masalan: +998901234567")

    if not PASSPORT_NUMBER_RE.match(passport_number):
        errors.append("Passport raqami 7-15 ta A-Z/0-9 belgidan iborat bo'lishi kerak.")

    if passport_series and not PASSPORT_SERIES_RE.match(passport_series):
        errors.append("Passport seriyasi 2 ta lotin harfi bo'lishi kerak (masalan: AA).")

    dob = parse_date_ymd(birth_date)
    if not dob:
        errors.append("Tug'ilgan sana noto'g'ri. Format: DD.MM.YYYY.")
    else:
        if dob > today:
            errors.append("Tug'ilgan sana kelajakda bo'lishi mumkin emas.")
        if dob.year < 1900:
            errors.append("Tug'ilgan sana juda eski ko'rinadi, qayta tekshiring.")

    if not (nationality or "").strip():
        errors.append("Millat/fuqarolik majburiy.")

    if gender not in {"male", "female"}:
        errors.append("Jins maydoni noto'g'ri.")

    if not is_valid_gmail(notification_email):
        errors.append("Ticket yuborish uchun Gmail majburiy (masalan: user@gmail.com).")

    issue_dt = parse_date_ymd(passport_issue_date)
    exp_dt = parse_date_ymd(passport_expiration_date)
    if not issue_dt:
        errors.append("Passport berilgan sanasi majburiy. Format: DD.MM.YYYY.")
    if not exp_dt:
        errors.append("Passport amal qilish muddati majburiy. Format: DD.MM.YYYY.")

    if issue_dt and exp_dt:
        if issue_dt > today:
            errors.append("Passport berilgan sanasi kelajakda bo'lishi mumkin emas.")
        oldest_issue = today - timedelta(days=PASSPORT_ISSUE_MAX_AGE_YEARS * 366)
        if issue_dt < oldest_issue:
            errors.append("Passport berilgan sana oxirgi 10 yil ichida bo'lishi kerak.")
        if exp_dt <= issue_dt:
            errors.append("Passport tugash sanasi berilgan sanadan keyin bo'lishi kerak.")
        max_exp = issue_dt + timedelta(days=PASSPORT_EXPIRY_MAX_YEARS * 366)
        if exp_dt > max_exp:
            errors.append("Passport amal muddati berilgan sanadan boshlab 20 yildan oshmasligi kerak.")

    return errors


def validate_passenger_payload(
    full_name: str,
    phone: str,
    passport_number: str,
    passport_series: str,
    birth_date: str,
    nationality: str,
    passport_issue_date: str,
    passport_expiration_date: str,
    gender: str,
    notification_email: str,
    passenger_type: str = "",
    travel_date: str | None = None,
) -> list[str]:
    errors: list[str] = []
    if len((full_name or "").strip()) < 3:
        errors.append("Ism-familiya kamida 3 ta belgidan iborat bo'lishi kerak.")

    expected_passenger_type = passenger_type_from_birth_date(birth_date, travel_date)
    supplied_passenger_type = (passenger_type or "").strip().upper()
    if supplied_passenger_type:
        if supplied_passenger_type not in PASSENGER_TYPE_LABELS:
            errors.append("Yosh kategoriyasi noto'g'ri.")
        elif expected_passenger_type and supplied_passenger_type != expected_passenger_type:
            errors.append(
                f"Yosh kategoriyasi tug'ilgan sana bo'yicha {passenger_type_label(expected_passenger_type)} bo'lishi kerak."
            )

    errors.extend(
        validate_profile_payload(
            phone,
            passport_number,
            passport_series,
            birth_date,
            nationality,
            passport_issue_date,
            passport_expiration_date,
            gender,
            notification_email,
        )
    )
    return errors


def passport_travel_valid(passport_expiration_date: str) -> tuple[bool, str]:
    exp_dt = parse_date_ymd(passport_expiration_date)
    if not exp_dt:
        return False, "Passport amal muddati kiritilmagan yoki noto'g'ri."
    min_valid_date = datetime.now(timezone.utc).date() + timedelta(days=PASSPORT_MIN_VALIDITY_DAYS)
    if exp_dt < min_valid_date:
        return False, "Passport safar uchun kamida 6 oy amal qilishi kerak."
    return True, ""


def send_plain_email(to_email: str, subject: str, body: str) -> tuple[bool, str]:
    if not MAIL_ENABLED:
        return False, "MAIL_ENABLED=0"
    if not MAIL_FROM_EMAIL or not MAIL_SMTP_HOST:
        return False, "mail config yetarli emas"
    if not is_valid_email(to_email):
        return False, "email noto'g'ri"

    message = EmailMessage()
    message["Subject"] = subject
    message["From"] = MAIL_FROM_EMAIL
    message["To"] = to_email
    message.set_content(body)

    try:
        deliver_email_message(message)
        return True, f"email yuborildi: {to_email}"
    except Exception as ex:
        app.logger.exception("Email yuborishda xatolik")
        return False, f"email xatosi: {ex}"


def build_registration_code() -> str:
    return f"{secrets.randbelow(900000) + 100000:06d}"


def registration_pending_expired(pending: dict | None) -> bool:
    if not pending:
        return True
    expires_at = pending.get("expires_at")
    if not expires_at:
        return True
    try:
        return datetime.fromisoformat(expires_at) <= datetime.now(timezone.utc)
    except ValueError:
        return True


def registration_resend_remaining_seconds(pending: dict | None) -> int:
    if not pending:
        return 0
    last_sent_at = pending.get("last_sent_at")
    if not last_sent_at:
        return 0
    try:
        sent_at = datetime.fromisoformat(last_sent_at)
    except ValueError:
        return 0
    remaining = REGISTRATION_RESEND_SECONDS - int((datetime.now(timezone.utc) - sent_at).total_seconds())
    return max(0, remaining)


def password_reset_pending_expired(pending: dict | None) -> bool:
    if not pending:
        return True
    expires_at = pending.get("expires_at")
    if not expires_at:
        return True
    try:
        return datetime.fromisoformat(expires_at) <= datetime.now(timezone.utc)
    except ValueError:
        return True


def password_reset_resend_remaining_seconds(pending: dict | None) -> int:
    if not pending:
        return 0
    last_sent_at = pending.get("last_sent_at")
    if not last_sent_at:
        return 0
    try:
        sent_at = datetime.fromisoformat(last_sent_at)
    except ValueError:
        return 0
    remaining = PASSWORD_RESET_RESEND_SECONDS - int((datetime.now(timezone.utc) - sent_at).total_seconds())
    return max(0, remaining)


def send_registration_code_email(email: str, full_name: str, otp_code: str) -> tuple[bool, str]:
    return send_plain_email(
        email,
        "Safar24 tasdiqlash kodi",
        (
            f"Assalomu alaykum, {full_name}!\n\n"
            f"Sizning tasdiqlash kodingiz: {otp_code}\n"
            f"Kodni {REGISTRATION_CODE_MINUTES} daqiqa ichida kiriting.\n"
            f"Agar kod kelmasa, {REGISTRATION_RESEND_SECONDS // 60} daqiqadan keyin qayta yuborishingiz mumkin.\n\n"
            "Safar24"
        ),
    )


def send_welcome_email(email: str, full_name: str, account_id: str) -> tuple[bool, str]:
    return send_plain_email(
        email,
        "Safar24 ga xush kelibsiz",
        (
            f"Assalomu alaykum, {full_name}!\n\n"
            "Safar24 tizimida ro'yxatdan o'tganingiz bilan tabriklaymiz.\n"
            f"Sizning account ID: {account_id}\n\n"
            "Endi kabinetga kirib profilingizni to'ldirishingiz, reys qidirishingiz va bron yaratishingiz mumkin.\n\n"
            "Yaxshi safarlar tilaymiz,\n"
            "Safar24 jamoasi"
        ),
    )


def airport_label(code: str, airport: dict) -> str:
    city = airport.get("city") or airport.get("name") or ""
    country = airport.get("country") or ""
    city_part = f" ({city})" if city else ""
    country_part = f", {country}" if country else ""
    return f"{code} - {airport.get('name', 'Airport')}{city_part}{country_part}"


def airport_search_results(query: str, limit: int = 10) -> list[dict]:
    term = (query or "").strip().lower()
    if not term:
        return []

    scored: list[tuple[int, str, dict]] = []
    for code, airport in AIRPORTS.items():
        name = str(airport.get("name", "")).lower()
        city = str(airport.get("city", "")).lower()
        country = str(airport.get("country", "")).lower()
        icao = str(airport.get("icao", "")).lower()

        haystack = f"{code.lower()} {icao} {name} {city} {country}"
        if term not in haystack:
            continue

        score = 5
        if code.lower().startswith(term):
            score = 0
        elif icao.startswith(term):
            score = 1
        elif city.startswith(term):
            score = 2
        elif name.startswith(term):
            score = 3
        scored.append((score, code, airport))

    scored.sort(key=lambda item: (item[0], item[1]))

    results: list[dict] = []
    for _, code, airport in scored[:limit]:
        results.append(
            {
                "code": code,
                "icao": airport.get("icao", ""),
                "name": airport.get("name", ""),
                "city": airport.get("city", ""),
                "country": airport.get("country", ""),
                "label": airport_label(code, airport),
            }
        )
    return results


def resolve_airport_to_city(value: str) -> str:
    term = (value or "").strip()
    if not term:
        return ""

    # Handle "City(CODE)" format from autocomplete
    if "(" in term and ")" in term:
        city_part = term.split("(")[0].strip()
        return city_part if city_part else term

    upper = term.upper()
    if upper in AIRPORTS:
        airport = AIRPORTS[upper]
        return str(airport.get("city") or airport.get("name") or term)

    search = airport_search_results(term, limit=1)
    if search:
        return search[0].get("city") or search[0].get("name") or term
    return term


def resolve_airport_code(value: str) -> str:
    term = (value or "").strip()
    if not term:
        return ""

    # Handle "City(CODE)" format from autocomplete
    if "(" in term and ")" in term:
        code_part = term.split("(")[1].split(")")[0].strip().upper()
        if code_part and len(code_part) == 3:
            return code_part

    upper = term.upper()
    if upper in AIRPORTS:
        return upper

    search = airport_search_results(term, limit=1)
    if search:
        return search[0]["code"]
    return ""


def city_to_airport_code(city: str) -> str:
    term = (city or "").strip().lower()
    if not term:
        return ""

    exact_matches = []
    for code, airport in AIRPORTS.items():
        airport_city = str(airport.get("city", "")).strip().lower()
        airport_name = str(airport.get("name", "")).strip().lower()
        if airport_city == term or airport_name == term:
            exact_matches.append(code)

    if exact_matches:
        exact_matches.sort()
        return exact_matches[0]

    search = airport_search_results(city, limit=1)
    if search:
        return search[0]["code"]
    return ""


def export_route_short_code(flight: sqlite3.Row) -> str:
    from_code = str(flight["from_airport_code"] or "").strip().upper()
    to_code = str(flight["to_airport_code"] or "").strip().upper()

    if not from_code:
        from_code = city_to_airport_code(str(flight["from_city"] or ""))
    if not to_code:
        to_code = city_to_airport_code(str(flight["to_city"] or ""))

    if not from_code:
        from_code = re.sub(r"[^A-Za-z0-9]+", "", str(flight["from_city"] or "").upper())[:3]
    if not to_code:
        to_code = re.sub(r"[^A-Za-z0-9]+", "", str(flight["to_city"] or "").upper())[:3]

    return f"{from_code or 'ORG'}-{to_code or 'DST'}"


def ticket_route_slug(flight: sqlite3.Row) -> str:
    base_code = export_route_short_code(flight)
    slug = re.sub(r"[^A-Za-z0-9-]+", "", base_code).strip("-").lower()
    return slug or "route"


# ----------------------------
# Auth helpers
# ----------------------------
def login_required(view_func):
    @wraps(view_func)
    def wrapper(*args, **kwargs):
        if not session.get("user_id"):
            flash("Davom etish uchun tizimga kiring.", "warning")
            return redirect(url_for("login", next=request.path))
        return view_func(*args, **kwargs)

    return wrapper


def admin_required(view_func):
    @wraps(view_func)
    def wrapper(*args, **kwargs):
        if not session.get("user_id") or not session.get("is_admin"):
            return redirect(url_for("admin_login"))
        return view_func(*args, **kwargs)

    return wrapper


def super_admin_required(view_func):
    @wraps(view_func)
    def wrapper(*args, **kwargs):
        if not session.get("user_id") or not session.get("is_super_admin"):
            flash("Ushbu bo'lim faqat super admin uchun.", "danger")
            return redirect(url_for("admin_dashboard"))
        return view_func(*args, **kwargs)

    return wrapper


@app.before_request
def load_user() -> None:
    g.user = None
    if request.endpoint == "static":
        return

    uid = session.get("user_id")
    if not uid:
        return

    db = get_db()
    user = db.execute("SELECT * FROM users WHERE id = ?", (uid,)).fetchone()
    if not user:
        session.clear()
        return

    if not ensure_active_user_device_session(db, int(user["id"])):
        session.clear()
        return

    g.user = user
    fresh_is_admin = bool(user["is_admin"])
    fresh_is_super_admin = bool(user["is_super_admin"])
    if session.get("is_admin") != fresh_is_admin:
        session["is_admin"] = fresh_is_admin
    if session.get("is_super_admin") != fresh_is_super_admin:
        session["is_super_admin"] = fresh_is_super_admin
    db.commit()


@app.after_request
def set_security_headers(response):
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    response.headers["Content-Security-Policy"] = (
        "default-src 'self'; "
        "img-src 'self' data:; "
        "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com; "
        "font-src 'self' https://fonts.gstatic.com; "
        "script-src 'self';"
    )
    return response


@app.context_processor
def inject_csrf_token():
    def csrf_token() -> str:
        token = session.get("_csrf_token")
        if not token:
            token = secrets.token_urlsafe(32)
            session["_csrf_token"] = token
        return token

    return {"csrf_token": csrf_token}


@app.before_request
def validate_csrf() -> None:
    if request.method not in {"POST", "PUT", "PATCH", "DELETE"}:
        return
    token = request.form.get("csrf_token") or request.headers.get("X-CSRF-Token")
    session_token = session.get("_csrf_token")
    if not token or not session_token or token != session_token:
        abort(400, description="Noto'g'ri so'rov (CSRF).")


# ----------------------------
# Utility
# ----------------------------
def generate_account_id(user_id: int) -> str:
    return f"A{user_id:04d}"


def parse_positive_int(raw_value: str | None, default: int = 1, minimum: int = 1) -> int:
    try:
        value = int(str(raw_value or "").strip())
    except (TypeError, ValueError):
        return default
    return value if value >= minimum else default


def normalize_booking_mix_for_ui(adult_count: int, child_count: int, infant_count: int) -> tuple[int, int, int]:
    adults = max(1, int(adult_count or 1))
    children = max(0, int(child_count or 0))
    infants = max(0, min(int(infant_count or 0), adults))

    total = adults + children + infants
    if total <= MAX_BOOKING_TRAVELERS:
        return adults, children, infants

    overflow = total - MAX_BOOKING_TRAVELERS
    if children > 0:
        cut_child = min(children, overflow)
        children -= cut_child
        overflow -= cut_child
    if overflow > 0 and infants > 0:
        infants = max(0, infants - overflow)

    return adults, children, infants


def is_safe_sql_identifier(name: str) -> bool:
    return bool(name and SAFE_SQL_IDENTIFIER_RE.fullmatch(name))


def quote_sql_identifier(name: str) -> str:
    if not is_safe_sql_identifier(name):
        raise ValueError("Unsafe SQL identifier")
    return f'"{name}"'


def list_user_table_names(db: sqlite3.Connection) -> list[str]:
    rows = db.execute(
        """
        SELECT name
        FROM sqlite_master
        WHERE type = 'table'
          AND name NOT LIKE 'sqlite_%'
        ORDER BY name ASC
        """
    ).fetchall()
    names: list[str] = []
    for row in rows:
        table_name = str(row["name"] or "").strip()
        if is_safe_sql_identifier(table_name):
            names.append(table_name)
    return names


def get_table_columns(db: sqlite3.Connection, table_name: str) -> list[sqlite3.Row]:
    quoted = quote_sql_identifier(table_name)
    return db.execute(f"PRAGMA table_info({quoted})").fetchall()


def resolve_table_record_key(db: sqlite3.Connection, table_name: str) -> dict[str, str]:
    columns = get_table_columns(db, table_name)
    pk_columns = [str(col["name"] or "").strip() for col in columns if int(col["pk"] or 0) > 0]
    if len(pk_columns) == 1 and is_safe_sql_identifier(pk_columns[0]):
        return {
            "mode": "pk",
            "column": pk_columns[0],
            "select_prefix": "",
            "select_suffix": f", {quote_sql_identifier(pk_columns[0])} AS __record_key",
            "order_by": quote_sql_identifier(pk_columns[0]),
            "where_expr": quote_sql_identifier(pk_columns[0]),
        }

    return {
        "mode": "rowid",
        "column": "rowid",
        "select_prefix": "rowid AS __record_key, ",
        "select_suffix": "",
        "order_by": "rowid",
        "where_expr": "rowid",
    }


def build_table_search_clause(column_names: list[str], query: str) -> tuple[str, list[str]]:
    keyword = (query or "").strip()
    safe_columns = [name for name in column_names if is_safe_sql_identifier(name)]
    if not keyword or not safe_columns:
        return "", []

    like_value = f"%{keyword}%"
    filters = [f"CAST({quote_sql_identifier(name)} AS TEXT) LIKE ?" for name in safe_columns[:16]]
    return f" WHERE {' OR '.join(filters)}", [like_value] * len(filters)


def sanitize_record_keys(raw_keys: list[str], key_mode: str) -> list[str | int]:
    cleaned: list[str | int] = []
    for raw in raw_keys:
        token = (raw or "").strip()
        if not token:
            continue
        if key_mode == "rowid":
            try:
                cleaned.append(int(token))
            except (TypeError, ValueError):
                continue
        else:
            cleaned.append(token)
    return cleaned


def fetch_table_preview(
    db: sqlite3.Connection,
    table_name: str,
    query: str,
    page: int,
    page_size: int,
) -> dict[str, object]:
    quoted_table = quote_sql_identifier(table_name)
    columns_info = get_table_columns(db, table_name)
    column_names = [str(col["name"] or "").strip() for col in columns_info if str(col["name"] or "").strip()]
    key_info = resolve_table_record_key(db, table_name)
    where_clause, where_params = build_table_search_clause(column_names, query)

    total_rows = db.execute(
        f"SELECT COUNT(*) AS c FROM {quoted_table}{where_clause}",
        tuple(where_params),
    ).fetchone()["c"]
    total_pages = max(1, math.ceil(total_rows / page_size)) if page_size > 0 else 1
    page = min(max(1, page), total_pages)
    offset = (page - 1) * page_size

    select_sql = (
        f"SELECT {key_info['select_prefix']}*{key_info['select_suffix']} "
        f"FROM {quoted_table}{where_clause} "
        f"ORDER BY {key_info['order_by']} DESC LIMIT ? OFFSET ?"
    )
    preview_rows = db.execute(select_sql, (*where_params, page_size, offset)).fetchall()

    rows = [dict(row) for row in preview_rows]
    table_columns = column_names.copy()
    if rows and "__record_key" in rows[0]:
        table_columns = ["__record_key", *column_names]

    return {
        "columns": table_columns,
        "rows": rows,
        "page": page,
        "page_size": page_size,
        "total_rows": int(total_rows),
        "total_pages": int(total_pages),
        "query": (query or "").strip(),
        "can_delete": bool(rows and "__record_key" in rows[0]),
        "key_info": key_info,
    }


def fetch_rows_for_export(
    db: sqlite3.Connection,
    table_name: str,
    query: str,
    selected_keys: list[str],
) -> tuple[list[str], list[dict[str, object]], dict[str, str]]:
    quoted_table = quote_sql_identifier(table_name)
    columns_info = get_table_columns(db, table_name)
    column_names = [str(col["name"] or "").strip() for col in columns_info if str(col["name"] or "").strip()]
    key_info = resolve_table_record_key(db, table_name)

    where_parts: list[str] = []
    where_params: list[str | int] = []

    search_clause, search_params = build_table_search_clause(column_names, query)
    if search_clause:
        where_parts.append(search_clause.replace(" WHERE ", "", 1))
        where_params.extend(search_params)

    cleaned_keys = sanitize_record_keys(selected_keys, key_info["mode"])
    if cleaned_keys:
        placeholders = ",".join("?" for _ in cleaned_keys)
        where_parts.append(f"{key_info['where_expr']} IN ({placeholders})")
        where_params.extend(cleaned_keys)

    where_sql = f" WHERE {' AND '.join(where_parts)}" if where_parts else ""
    select_sql = (
        f"SELECT {key_info['select_prefix']}*{key_info['select_suffix']} "
        f"FROM {quoted_table}{where_sql} "
        f"ORDER BY {key_info['order_by']} DESC"
    )
    records = [dict(row) for row in db.execute(select_sql, tuple(where_params)).fetchall()]

    export_columns = column_names.copy()
    if records and "__record_key" in records[0]:
        export_columns = ["__record_key", *column_names]

    return export_columns, records, key_info


def redirect_super_admin_with_db_state(
    q: str = "",
    db_table: str = "",
    db_q: str = "",
    db_page: int = 1,
):
    params: dict[str, object] = {}
    if q:
        params["q"] = q
    if db_table:
        params["db_table"] = db_table
    if db_q:
        params["db_q"] = db_q
    if int(db_page) > 1:
        params["db_page"] = int(db_page)
    return redirect(url_for("super_admin_panel", **params))


def user_wallet_account_no(user_id: int, currency: str) -> str:
    code = (currency or "UZS").upper()
    return f"{code}{user_id:03d}"


def convert_currency(amount: float, from_currency: str, to_currency: str) -> float:
    src = (from_currency or "").upper()
    dst = (to_currency or "").upper()
    if src == dst:
        return amount
    if src == "USD" and dst == "UZS":
        return amount * USD_TO_UZS_RATE
    if src == "UZS" and dst == "USD":
        return amount / USD_TO_UZS_RATE
    raise ValueError("Qo'llab-quvvatlanmagan valyuta.")


def generate_temporary_password(length: int = 12) -> str:
    alphabet = "ABCDEFGHJKLMNPQRSTUVWXYZabcdefghijkmnopqrstuvwxyz23456789"
    return "".join(secrets.choice(alphabet) for _ in range(max(8, length)))


def generate_next_ticket_no(db: sqlite3.Connection) -> str:
    row = db.execute("SELECT ticket_no FROM tickets ORDER BY id DESC LIMIT 1").fetchone()
    if row:
        next_int = int(row["ticket_no"]) + 1
    else:
        next_int = 1
    return f"{next_int:010d}"


def allowed_proof(filename: str) -> bool:
    if "." not in filename:
        return False
    ext = filename.rsplit(".", 1)[1].lower()
    return ext in ALLOWED_PROOF_EXT


def require_profile(user_id: int) -> bool:
    row = get_db().execute(
        """
        SELECT up.passport_number, up.birth_date, up.nationality, up.gender, up.notification_email,
               up.passport_series, up.passport_issue_date, up.passport_expiration_date,
               u.phone
        FROM user_profiles up
        JOIN users u ON u.id = up.user_id
        WHERE up.user_id = ?
        """,
        (user_id,),
    ).fetchone()
    if not row:
        return False

    required_values = [
        row["passport_number"],
        row["birth_date"],
        row["nationality"],
        row["gender"],
        row["notification_email"],
        row["passport_issue_date"],
        row["passport_expiration_date"],
        row["phone"],
    ]
    if any(not str(v or "").strip() for v in required_values):
        return False
    if not is_valid_gmail((row["notification_email"] or "").strip().lower()):
        return False
    validation_errors = validate_profile_payload(
        (row["phone"] or "").strip(),
        (row["passport_number"] or "").strip().upper(),
        (row["passport_series"] or "").strip().upper(),
        (row["birth_date"] or "").strip(),
        (row["nationality"] or "").strip(),
        (row["passport_issue_date"] or "").strip(),
        (row["passport_expiration_date"] or "").strip(),
        (row["gender"] or "").strip().lower(),
        (row["notification_email"] or "").strip().lower(),
    )
    if validation_errors:
        return False
    travel_ok, _ = passport_travel_valid((row["passport_expiration_date"] or "").strip())
    return travel_ok


def format_currency_display(amount_value: float | int | None, currency: str, amount_uzs: int | None = None) -> str:
    code = (currency or "UZS").upper()
    if code == "USD":
        return f"{float(amount_value or 0):,.2f} USD"
    base = amount_uzs if amount_uzs is not None else int(round(float(amount_value or 0)))
    return f"{base:,} UZS"


def format_date_dot(value: str | None) -> str:
    dt = parse_date_ymd(value or "")
    if not dt:
        return (value or "").strip() or "-"
    return dt.strftime("%d.%m.%Y")


MANIFEST_NATIONALITY_ALIASES = (
    ("uzbekistan", "UZB", "UZBEKISTAN"),
    ("uzbek", "UZB", "UZBEKISTAN"),
    ("ozbek", "UZB", "UZBEKISTAN"),
    ("kyrgyzstan", "KGZ", "KYRGYZ"),
    ("kyrgyz", "KGZ", "KYRGYZ"),
    ("qirgiz", "KGZ", "KYRGYZ"),
    ("qirgiziston", "KGZ", "KYRGYZ"),
    ("qazaq", "KAZ", "KAZAKISTAN"),
    ("qozoq", "KAZ", "KAZAKISTAN"),
    ("kazakhstan", "KAZ", "KAZAKISTAN"),
    ("kazakh", "KAZ", "KAZAKISTAN"),
    ("tajikistan", "TJK", "TAJIKISTAN"),
    ("tajik", "TJK", "TAJIKISTAN"),
    ("turkmenistan", "TKM", "TURKMENISTAN"),
    ("turkmen", "TKM", "TURKMENISTAN"),
    ("russia", "RUS", "RUSSIAN"),
    ("russian", "RUS", "RUSSIAN"),
    ("azerbaijan", "AZE", "AZERBAIJAN"),
    ("azeri", "AZE", "AZERBAIJAN"),
    ("turkey", "TUR", "TURKISH"),
    ("turkish", "TUR", "TURKISH"),
)

MANIFEST_NATIONALITY_DISPLAY_BY_CODE = {
    "AZE": "AZERBAIJAN",
    "KGZ": "KYRGYZ",
    "KAZ": "KAZAKISTAN",
    "RUS": "RUSSIAN",
    "TJK": "TAJIKISTAN",
    "TKM": "TURKMENISTAN",
    "TUR": "TURKISH",
    "UZB": "UZBEKISTAN",
}


def manifest_nationality_fields(value: str | None) -> tuple[str, str]:
    raw = " ".join((value or "").split()).strip()
    if not raw:
        return "", ""

    normalized = re.sub(r"[^a-z0-9]+", "", raw.lower())
    for alias, code, display in MANIFEST_NATIONALITY_ALIASES:
        if alias in normalized:
            return code, display

    if len(normalized) == 3 and normalized.isalpha():
        code = normalized.upper()
        return code, MANIFEST_NATIONALITY_DISPLAY_BY_CODE.get(code, code)

    code = re.sub(r"[^A-Za-z]", "", raw).upper()[:3]
    display = raw.upper()
    return code or display, display


def build_centrum_export_workbook(flight: sqlite3.Row, passengers: list[sqlite3.Row]):
    from openpyxl import load_workbook

    template_path = BASE_DIR / "Centrum.xlsx"
    if not template_path.exists():
        template_path = BASE_DIR / "19HK.xlsx"
    if not template_path.exists():
        raise FileNotFoundError("Centrum.xlsx shabloni topilmadi.")

    wb = load_workbook(template_path)
    ws = wb["Pax Data"] if "Pax Data" in wb.sheetnames else wb.active

    max_row = ws.max_row
    max_col = ws.max_column
    for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.value = None

    travel_date = str(flight["departure_date"] or "")
    previous_booking_id = None
    for row_num, pax in enumerate(passengers, start=2):
        first_name, last_name = passenger_name_parts(pax)
        passenger_type_code = passenger_type_from_birth_date(str(pax["birth_date"] or ""), travel_date) or normalize_passenger_type(pax["passenger_type"])
        if not passenger_type_code:
            passenger_type_code = "ADT"

        gender_value = str(pax["gender"] or "").strip().lower()
        gender_label = "MALE" if gender_value == "male" else ("FEMALE" if gender_value == "female" else "-")
        nationality_code, passport_country = manifest_nationality_fields(str(pax["nationality"] or ""))
        passport_number = str(pax["passport_number"] or "").strip().upper()
        passport_series = str(pax["passport_series"] or "").strip().upper()
        display_passport_number = passport_number
        if passport_series and passport_number and not passport_number.startswith(passport_series):
            display_passport_number = f"{passport_series}{passport_number}"

        is_first_booking_row = pax["booking_id"] != previous_booking_id
        previous_booking_id = pax["booking_id"]
        phone_value = str(pax["phone"] or "").strip() if is_first_booking_row else ""
        email_value = str((pax["notification_email"] or pax["email"] or "")).strip() if is_first_booking_row else ""

        row_values = [
            row_num - 1,
            passenger_type_code,
            gender_label,
            (last_name.upper() if last_name else "-"),
            (first_name.upper() if first_name else "-"),
            gender_label,
            format_date_dot(pax["birth_date"]),
            nationality_code or "-",
            phone_value or None,
            email_value or None,
            None,
            display_passport_number or None,
            format_date_dot(pax["passport_expiration_date"]),
            passport_country or "-",
            passport_country or "-",
        ]

        for col_num, value in enumerate(row_values, start=1):
            ws.cell(row=row_num, column=col_num, value=value)

    return wb


def format_datetime_dot(date_value: str | None, time_value: str | None = None) -> str:
    date_part = format_date_dot(date_value)
    time_part = (time_value or "").strip()
    return f"{date_part} {time_part}".strip()


def baggage_option_multiplier(option: str | None) -> float:
    normalized = (option or "standard").strip().lower()
    mapping = {
        "cabin_only": 0.9,
        "standard": 1.0,
        "premium": 1.25,
    }
    return mapping.get(normalized, 1.0)


def baggage_option_label(option: str | None, flight: sqlite3.Row | None = None) -> str:
    normalized = (option or "standard").strip().lower()
    baggage_kg = int(flight["baggage_kg"] or 20) if flight else 20
    labels = {
        "cabin_only": "Bez bagaj",
        "standard": f"Bagaj bilan ({baggage_kg} kg)",
        "premium": f"Premium bagaj ({baggage_kg + 10} kg)",
    }
    return labels.get(normalized, "Bagaj bilan")


@app.template_filter("date_dot")
def date_dot_filter(value: str | None) -> str:
    return format_date_dot(value)


@app.template_filter("datetime_dot")
def datetime_dot_filter(value: str | None, time_value: str | None = None) -> str:
    return format_datetime_dot(value, time_value)


def payment_method_label(method: str | None) -> str:
    labels = {
        "wallet": "Wallet",
        "click": "Click",
        "payme": "Payme",
        "manual": "Manual",
    }
    return labels.get((method or "").strip().lower(), (method or "Unknown").strip().title() or "Unknown")


def booking_passenger_names(db: sqlite3.Connection, booking_id: int) -> list[str]:
    rows = db.execute(
        "SELECT full_name FROM passengers WHERE booking_id = ? ORDER BY id ASC",
        (booking_id,),
    ).fetchall()
    return [str(row["full_name"] or "").strip() for row in rows if str(row["full_name"] or "").strip()]


def booking_passenger_rows(db: sqlite3.Connection, booking_id: int) -> list[sqlite3.Row]:
    return db.execute(
        "SELECT * FROM passengers WHERE booking_id = ? ORDER BY id ASC",
        (booking_id,),
    ).fetchall()


def notification_recipients_from_passengers(
    passenger_rows: list[sqlite3.Row],
    user: sqlite3.Row,
    profile: sqlite3.Row | None,
) -> list[str]:
    for row in passenger_rows:
        email = (row["notification_email"] or "").strip().lower()
        if email and is_valid_email(email):
            return [email]

    fallback = resolve_notification_email(user, profile)
    return [fallback] if fallback else []


def ticket_public_url(ticket_no: str) -> str:
    base = (APP_BASE_URL or "").rstrip("/")
    if base:
        return f"{base}/ticket/{ticket_no}"
    if has_request_context():
        return url_for("ticket_public_view", ticket_no=ticket_no, _external=True)
    return f"/ticket/{ticket_no}"


def deliver_email_message(message: EmailMessage) -> None:
    if MAIL_USE_SSL:
        with smtplib.SMTP_SSL(MAIL_SMTP_HOST, MAIL_SMTP_PORT, timeout=25) as smtp:
            if MAIL_USERNAME:
                smtp.login(MAIL_USERNAME, MAIL_PASSWORD)
            smtp.send_message(message)
        return

    with smtplib.SMTP(MAIL_SMTP_HOST, MAIL_SMTP_PORT, timeout=25) as smtp:
        smtp.ehlo()
        if MAIL_USE_TLS:
            smtp.starttls()
            smtp.ehlo()
        if MAIL_USERNAME:
            smtp.login(MAIL_USERNAME, MAIL_PASSWORD)
        smtp.send_message(message)


def build_ticket_email_html(
    user: sqlite3.Row,
    profile: sqlite3.Row | None,
    flight: sqlite3.Row,
    ticket_no: str,
    public_url: str,
    passenger_names: list[str],
    payment_amount: str,
    payment_method: str,
) -> str:
    route = f"{flight['from_city']} ({flight['from_airport_code'] or '-'}) → {flight['to_city']} ({flight['to_airport_code'] or '-'})"
    passenger_items = "".join(
        f"<li style=\"margin:0 0 6px;\">{escape(name)}</li>"
        for name in (passenger_names or [str(user['full_name'] or 'Yo‘lovchi')])
    )
    passport = escape((profile["passport_number"] if profile and profile["passport_number"] else "-"))
    safe_name = escape(str(user["full_name"] or "Safar24 mijoz"))
    safe_airline = escape(str(flight["airline"] or "Safar24"))
    safe_route = escape(route)
    safe_departure = escape(f"{flight['departure_date']} {flight['departure_time'] or ''}".strip())
    safe_ticket = escape(ticket_no)
    safe_payment_amount = escape(payment_amount)
    safe_payment_method = escape(payment_method)
    safe_public_url = escape(public_url, quote=True)

    return f"""
    <html>
      <body style="margin:0;padding:24px;background:#eef5ff;font-family:Arial,sans-serif;color:#10243d;">
        <div style="max-width:680px;margin:0 auto;background:#ffffff;border-radius:28px;overflow:hidden;border:1px solid #d4e2f5;box-shadow:0 24px 48px rgba(16,36,61,0.12);">
          <div style="padding:28px 32px;background:linear-gradient(135deg,#06111f,#123a63);color:#ffffff;">
            <div style="font-size:12px;letter-spacing:2px;font-weight:700;color:#7fe7ff;margin-bottom:10px;">SAFAR24 E-TICKET</div>
            <h1 style="margin:0 0 10px;font-size:30px;line-height:1.1;">Ticket tayyor</h1>
            <p style="margin:0;font-size:15px;line-height:1.6;color:#dbe9ff;">{safe_name}, safaringiz uchun PDF ticket ilova qilindi. Web ko‘rinish va tez tekshiruv havolasi ham tayyor.</p>
          </div>

          <div style="padding:28px 32px;">
            <div style="display:inline-block;padding:8px 14px;border-radius:999px;background:#e8f6ff;border:1px solid #b5eaff;color:#0e6da5;font-size:12px;font-weight:700;margin-bottom:18px;">
              Ticket #{safe_ticket}
            </div>

            <div style="padding:20px;border-radius:22px;background:linear-gradient(180deg,#f7fbff,#eef6ff);border:1px solid #d7e7f8;margin-bottom:18px;">
              <div style="font-size:24px;font-weight:800;color:#10243d;margin-bottom:8px;">{safe_route}</div>
              <div style="font-size:14px;color:#4b6078;line-height:1.8;">
                <div><strong>Sana/Vaqt:</strong> {safe_departure}</div>
                <div><strong>Aviakompaniya:</strong> {safe_airline}</div>
                <div><strong>To‘lov:</strong> {safe_payment_amount} via {safe_payment_method}</div>
                <div><strong>Passport:</strong> {passport}</div>
              </div>
            </div>

            <div style="margin-bottom:18px;">
              <div style="font-size:15px;font-weight:700;color:#10243d;margin-bottom:8px;">Yo‘lovchilar</div>
              <ol style="margin:0;padding-left:20px;color:#435973;line-height:1.7;">
                {passenger_items}
              </ol>
            </div>

            <div style="margin:26px 0;">
              <a href="{safe_public_url}" style="display:inline-block;padding:14px 22px;border-radius:999px;background:linear-gradient(135deg,#33d6ff,#2f8cff);color:#041321;text-decoration:none;font-weight:800;">
                Ticketni webda ko‘rish
              </a>
            </div>

            <div style="padding:18px 20px;border-radius:18px;background:#081727;color:#dce9fb;">
              <div style="font-size:14px;font-weight:700;color:#7fe7ff;margin-bottom:8px;">Muhim eslatma</div>
              <div style="font-size:13px;line-height:1.7;">
                Aeroportga oldinroq yetib boring, pasport va ushbu PDF ticketni o‘zingiz bilan olib yuring.
                Agar web ko‘rinish kerak bo‘lsa yuqoridagi tugma orqali ochishingiz mumkin.
              </div>
            </div>
          </div>
        </div>
      </body>
    </html>
    """


def save_ticket_pdf(
    ticket_no: str,
    booking: sqlite3.Row,
    user: sqlite3.Row,
    profile: sqlite3.Row | None,
    flight: sqlite3.Row,
    payment: sqlite3.Row,
    payment_method: str,
    passenger_rows: list[sqlite3.Row],
) -> str:
    route_slug = ticket_route_slug(flight)
    pdf_name = f"{route_slug}-{ticket_no}.pdf"
    pdf_path = TICKETS_DIR / pdf_name

    c = canvas.Canvas(str(pdf_path), pagesize=A4)
    w, h = A4

    ticket_display = f"Safar{ticket_no}"
    from_code = (flight["from_airport_code"] or "").upper()
    to_code = (flight["to_airport_code"] or "").upper()
    from_city = str(flight["from_city"] or "")
    to_city = str(flight["to_city"] or "")
    route_left = from_code or from_city[:3].upper()
    route_right = to_code or to_city[:3].upper()
    travel_class = str(flight["travel_class"] or "Economy")
    payment_amount = format_currency_display(
        payment["amount_value"] if payment["amount_value"] is not None else payment["amount_uzs"],
        str(payment["currency"] or "UZS"),
        int(payment["amount_uzs"] or 0),
    )
    payment_currency = normalize_currency_code(str(payment["currency"] or "UZS"))
    payment_base_amount = format_currency_display(
        payment["base_amount_value"] if payment["base_amount_value"] is not None else payment["amount_value"],
        payment_currency,
        int(payment["base_amount_uzs"] or payment["amount_uzs"] or 0),
    )
    payment_fee_amount = format_currency_display(
        payment["method_fee_value"] if payment["method_fee_value"] is not None else 0,
        payment_currency,
        int(payment["method_fee_uzs"] or 0),
    )
    payment_fee_percent = float(payment["method_fee_pct"] or 0)
    adult_count, child_count, infant_count = booking_age_counts(booking, passenger_rows)
    passenger_mix = passenger_mix_label(adult_count, child_count, infant_count)
    passenger_names = [str(row["full_name"] or "").strip() for row in passenger_rows if str(row["full_name"] or "").strip()]
    if not passenger_names:
        passenger_names = [str(user["full_name"] or "Yo'lovchi")]
    primary_passenger = passenger_rows[0] if passenger_rows else None
    issue_email = (
        (primary_passenger["notification_email"] if primary_passenger and primary_passenger["notification_email"] else None)
        or resolve_notification_email(user, profile)
        or str(user["email"] or "-")
    )
    issue_phone = (
        (primary_passenger["phone"] if primary_passenger and primary_passenger["phone"] else None)
        or str(user["phone"] or "-")
    )
    baggage_label = baggage_option_label(booking["baggage_option"], flight)
    qr_value = ticket_public_url(ticket_no)
    route_duration = format_duration_label(estimate_flight_duration_minutes(from_code, to_code))

    def draw_panel(x: float, y: float, width: float, height: float, fill_color: str, stroke_color: str | None = None, radius: float = 5 * mm) -> None:
        c.setFillColor(colors.HexColor(fill_color))
        if stroke_color:
            c.setStrokeColor(colors.HexColor(stroke_color))
            c.roundRect(x, y, width, height, radius, stroke=1, fill=1)
        else:
            c.roundRect(x, y, width, height, radius, stroke=0, fill=1)

    def draw_glass_panel(
        x: float,
        y: float,
        width: float,
        height: float,
        fill_color: str = "#FFFFFF",
        stroke_color: str = "#D5E5F6",
        radius: float = 6 * mm,
        alpha: float = 0.68,
    ) -> None:
        c.saveState()
        if hasattr(c, "setFillAlpha"):
            c.setFillAlpha(alpha)
        if hasattr(c, "setStrokeAlpha"):
            c.setStrokeAlpha(min(1.0, alpha + 0.18))
        c.setFillColor(colors.HexColor(fill_color))
        c.setStrokeColor(colors.HexColor(stroke_color))
        c.roundRect(x, y, width, height, radius, stroke=1, fill=1)
        c.restoreState()

        c.saveState()
        if hasattr(c, "setFillAlpha"):
            c.setFillAlpha(0.12)
        c.setFillColor(colors.white)
        c.roundRect(x + 1.5 * mm, y + height * 0.5, width - 3 * mm, max(3.2 * mm, height * 0.22), radius * 0.65, stroke=0, fill=1)
        c.restoreState()

    def draw_plane_icon(cx: float, cy: float, scale: float = 1.0, color: str = "#1F5F93") -> None:
        body = 6.2 * mm * scale
        nose = 2.2 * mm * scale
        wing = 3.6 * mm * scale
        tail = 2.4 * mm * scale
        stabilizer = 1.55 * mm * scale
        path = c.beginPath()
        path.moveTo(cx - body - tail, cy - 0.18 * mm * scale)
        path.lineTo(cx - body + 0.5 * mm * scale, cy - 0.18 * mm * scale)
        path.lineTo(cx - body + stabilizer, cy - 1.55 * mm * scale)
        path.lineTo(cx - body + stabilizer + 0.95 * mm * scale, cy - 1.55 * mm * scale)
        path.lineTo(cx - body + 1.0 * mm * scale, cy - 0.18 * mm * scale)
        path.lineTo(cx - 0.95 * mm * scale, cy - 0.18 * mm * scale)
        path.lineTo(cx + 1.15 * mm * scale, cy - wing)
        path.lineTo(cx + 2.25 * mm * scale, cy - wing)
        path.lineTo(cx + 0.95 * mm * scale, cy - 0.18 * mm * scale)
        path.lineTo(cx + body, cy - 0.18 * mm * scale)
        path.lineTo(cx + body + nose, cy + 0.58 * mm * scale)
        path.lineTo(cx + body + nose + 0.7 * mm * scale, cy + 0.08 * mm * scale)
        path.lineTo(cx + body + nose + 0.12 * mm * scale, cy - 0.42 * mm * scale)
        path.lineTo(cx + body, cy - 0.42 * mm * scale)
        path.lineTo(cx + 0.95 * mm * scale, cy - 0.42 * mm * scale)
        path.lineTo(cx + 2.25 * mm * scale, cy + wing)
        path.lineTo(cx + 1.15 * mm * scale, cy + wing)
        path.lineTo(cx - 0.95 * mm * scale, cy + 0.42 * mm * scale)
        path.lineTo(cx - body + 1.0 * mm * scale, cy + 0.42 * mm * scale)
        path.lineTo(cx - body + stabilizer + 0.95 * mm * scale, cy + 1.55 * mm * scale)
        path.lineTo(cx - body + stabilizer, cy + 1.55 * mm * scale)
        path.lineTo(cx - body + 0.5 * mm * scale, cy + 0.42 * mm * scale)
        path.lineTo(cx - body - tail, cy + 0.42 * mm * scale)
        path.close()
        c.saveState()
        c.setFillColor(colors.HexColor(color))
        c.setStrokeColor(colors.HexColor(color))
        c.drawPath(path, stroke=0, fill=1)
        c.restoreState()

    def draw_flight_path(x1: float, x2: float, y: float, color: str = "#9CC7EA") -> None:
        c.saveState()
        c.setStrokeColor(colors.HexColor(color))
        c.setLineWidth(1.05)
        c.setDash(2.8, 2.4)
        c.line(x1, y, x2, y)
        c.restoreState()

        c.saveState()
        c.setFillColor(colors.HexColor(color))
        c.circle(x1, y, 1.0 * mm, stroke=0, fill=1)
        c.circle(x2, y, 1.0 * mm, stroke=0, fill=1)
        c.restoreState()

    def draw_label_value(x: float, y: float, label: str, value: str, value_color: str = "#0D2238") -> None:
        c.setFont("Helvetica", 8)
        c.setFillColor(colors.HexColor("#6E87A2"))
        c.drawString(x, y, label.upper())
        c.setFont("Helvetica-Bold", 10)
        c.setFillColor(colors.HexColor(value_color))
        c.drawString(x, y - 5.5 * mm, value)

    def fit_text(value: str, font_name: str, font_size: float, max_width: float) -> str:
        raw = (value or "").strip()
        if not raw or c.stringWidth(raw, font_name, font_size) <= max_width:
            return raw
        ellipsis = "..."
        while raw and c.stringWidth(f"{raw}{ellipsis}", font_name, font_size) > max_width:
            raw = raw[:-1].rstrip()
        return f"{raw}{ellipsis}" if raw else ellipsis

    def draw_label_value_fit(
        x: float,
        y: float,
        label: str,
        value: str,
        max_width: float,
        value_color: str = "#0D2238",
        label_color: str = "#6E87A2",
        value_font_size: float = 10,
    ) -> None:
        c.setFont("Helvetica", 8)
        c.setFillColor(colors.HexColor(label_color))
        c.drawString(x, y, label.upper())
        c.setFont("Helvetica-Bold", value_font_size)
        c.setFillColor(colors.HexColor(value_color))
        c.drawString(x, y - 5.5 * mm, fit_text(value, "Helvetica-Bold", value_font_size, max_width))

    def compact(value: str, limit: int = 28) -> str:
        raw = (value or "").strip()
        if len(raw) <= limit:
            return raw
        return f"{raw[:limit - 1]}…"

    # Background
    c.setFillColor(colors.HexColor("#F2F7FF"))
    c.rect(0, 0, w, h, fill=1, stroke=0)
    c.setFillColor(colors.HexColor("#0A1730"))
    c.circle(w - 20 * mm, h - 18 * mm, 28 * mm, fill=1, stroke=0)
    c.setFillColor(colors.HexColor("#123E66"))
    c.circle(18 * mm, h - 30 * mm, 22 * mm, fill=1, stroke=0)

    # Main shell
    draw_panel(12 * mm, 12 * mm, w - 24 * mm, h - 24 * mm, "#07111F", "#0E3150", 9 * mm)
    draw_panel(16 * mm, 16 * mm, w - 32 * mm, h - 32 * mm, "#F8FBFF", "#D4E4F6", 8 * mm)

    # Hero header
    draw_panel(20 * mm, h - 60 * mm, w - 40 * mm, 30 * mm, "#081B31", "#0D3458", 7 * mm)
    c.setFillColor(colors.HexColor("#33D6FF"))
    c.roundRect(20 * mm, h - 62.5 * mm, 58 * mm, 3 * mm, 1.5 * mm, stroke=0, fill=1)

    logo_path = BASE_DIR / "static" / "logo.PNG"
    if logo_path.exists():
        draw_panel(23 * mm, h - 54 * mm, 32 * mm, 14 * mm, "#FFFFFF", None, 3 * mm)
        c.drawImage(ImageReader(str(logo_path)), 24 * mm, h - 52.8 * mm, 30 * mm, 11.4 * mm, mask="auto", preserveAspectRatio=True)

    c.setFillColor(colors.HexColor("#FFFFFF"))
    c.setFont("Helvetica-Bold", 18)
    c.drawString(60 * mm, h - 38 * mm, "Safar24 Group Travel Ticket")
    c.setFont("Helvetica", 8.5)
    c.setFillColor(colors.HexColor("#D8E9FF"))
    c.drawString(
        60 * mm,
        h - 43.5 * mm,
        fit_text("Guruh ticket, passenger ro'yxati va QR tekshiruv", "Helvetica", 8.5, 80 * mm),
    )
    c.setFont("Helvetica-Bold", 11)
    c.setFillColor(colors.HexColor("#7FE7FF"))
    c.drawRightString(w - 25 * mm, h - 37 * mm, fit_text(f"#{ticket_display}", "Helvetica-Bold", 11, 42 * mm))
    c.setFont("Helvetica", 8)
    c.setFillColor(colors.HexColor("#D8E9FF"))
    c.drawRightString(
        w - 25 * mm,
        h - 43 * mm,
        fit_text(
            f"Issued {format_date_dot(now_iso()[:10])} {datetime.now(timezone.utc).strftime('%H:%M')} UTC",
            "Helvetica",
            8,
            42 * mm,
        ),
    )

    # Route spotlight
    route_panel_y = h - 100 * mm
    route_panel_h = 30 * mm
    draw_glass_panel(20 * mm, route_panel_y, w - 40 * mm, route_panel_h, "#EEF7FF", "#B9D8F8", 7 * mm, 0.82)
    c.setFillColor(colors.HexColor("#6D87A3"))
    c.setFont("Helvetica-Bold", 9)
    c.drawString(26 * mm, route_panel_y + 22 * mm, from_city.upper())
    c.drawRightString(w - 26 * mm, route_panel_y + 22 * mm, to_city.upper())
    c.setFont("Helvetica-Bold", 22)
    c.setFillColor(colors.HexColor("#0C2038"))
    c.drawString(26 * mm, route_panel_y + 10 * mm, route_left)
    c.drawRightString(w - 26 * mm, route_panel_y + 10 * mm, route_right)
    c.setFont("Helvetica", 8.5)
    c.setFillColor(colors.HexColor("#38556F"))
    c.drawString(26 * mm, route_panel_y + 4.5 * mm, f"Flight {flight['route_code'] or '-'} • {flight['airline']}")
    c.drawRightString(w - 26 * mm, route_panel_y + 4.5 * mm, format_datetime_dot(flight["departure_date"], flight["departure_time"] or "--:--"))

    route_mid_y = route_panel_y + 11 * mm
    route_mid_x = w / 2
    draw_flight_path(82 * mm, route_mid_x - 24 * mm, route_mid_y)
    draw_flight_path(route_mid_x + 24 * mm, w - 82 * mm, route_mid_y)
    draw_glass_panel(route_mid_x - 19 * mm, route_panel_y + 6.6 * mm, 38 * mm, 13.2 * mm, "#FFFFFF", "#C5DEF5", 6.4 * mm, 0.74)
    draw_plane_icon(route_mid_x, route_panel_y + 16.1 * mm, 0.7, "#1C5C92")
    c.setFont("Helvetica-Bold", 7.9)
    c.setFillColor(colors.HexColor("#184D7A"))
    c.drawCentredString(route_mid_x, route_panel_y + 10.55 * mm, route_duration)
    c.setFont("Helvetica", 6.5)
    c.setFillColor(colors.HexColor("#6D87A3"))
    c.drawCentredString(route_mid_x, route_panel_y + 7.55 * mm, "parvoz yo'li")

    # Summary mini cards
    card_y = h - 124 * mm
    card_w = (w - 58 * mm) / 4
    for idx, (label, value, tone) in enumerate([
        ("Class", travel_class, "#135DD8"),
        ("Passengers", passenger_mix, "#0D2238"),
        ("Total", payment_amount, "#0D2238"),
        ("Method", payment_method_label(payment_method), "#0D2238"),
    ]):
        x = 20 * mm + idx * (card_w + 6 * mm)
        draw_panel(x, card_y, card_w, 18 * mm, "#FFFFFF", "#D5E5F6", 5 * mm)
        draw_label_value_fit(x + 3.5 * mm, card_y + 13.5 * mm, label, value, card_w - 7 * mm, tone)

    # Main detail panels
    left_x = 20 * mm
    right_x = 122 * mm
    top_y = 84 * mm
    detail_panel_h = 80 * mm
    left_w = 96 * mm
    right_w = w - right_x - 20 * mm
    draw_glass_panel(left_x, top_y, left_w, detail_panel_h, "#FFFFFF", "#D5E5F6", 6 * mm, 0.78)
    draw_panel(right_x, top_y, right_w, detail_panel_h, "#0B1B32", "#13395E", 6 * mm)

    c.setFont("Helvetica-Bold", 12)
    c.setFillColor(colors.HexColor("#123A63"))
    c.drawString(left_x + 5 * mm, top_y + 68 * mm, "Passenger Manifest")
    c.setFont("Helvetica", 8)
    c.setFillColor(colors.HexColor("#6C84A0"))
    c.drawString(left_x + 5 * mm, top_y + 62 * mm, "Ism passport bilan bir xil bo'lishi kerak")
    c.setFont("Helvetica-Bold", 8.2)
    c.setFillColor(colors.HexColor("#0D2238"))
    visible_passengers = passenger_rows[:5]
    line_y = top_y + 53 * mm
    for idx, row in enumerate(visible_passengers, start=1):
        full_name = fit_text(str(row["full_name"] or "-"), "Helvetica-Bold", 8.2, left_w - 14 * mm)
        passport = compact(f"{row['passport_series'] or ''}{row['passport_number'] or '-'}", 16)
        gender = "M" if (row["gender"] or "").lower() == "male" else ("F" if (row["gender"] or "").lower() == "female" else "-")
        c.drawString(left_x + 5 * mm, line_y, f"{idx:02d}. {full_name}")
        c.setFont("Helvetica", 7.2)
        c.setFillColor(colors.HexColor("#56718F"))
        c.drawString(left_x + 7 * mm, line_y - 4.5 * mm, f"Passport: {passport}  •  {gender}  •  {format_date_dot(row['birth_date'])}")
        c.setFont("Helvetica-Bold", 8.2)
        c.setFillColor(colors.HexColor("#0D2238"))
        line_y -= 9.5 * mm

    if len(passenger_rows) > len(visible_passengers):
        c.setFont("Helvetica", 7.2)
        c.setFillColor(colors.HexColor("#6C84A0"))
        c.drawString(left_x + 5 * mm, top_y + 16 * mm, f"+{len(passenger_rows) - len(visible_passengers)} yo'lovchi yana mavjud")

    draw_label_value_fit(left_x + 5 * mm, top_y + 12 * mm, "Lead contact", issue_phone, 40 * mm)
    draw_label_value_fit(left_x + 51 * mm, top_y + 12 * mm, "Group email", issue_email, 38 * mm)

    c.setFont("Helvetica-Bold", 12)
    c.setFillColor(colors.HexColor("#FFFFFF"))
    c.drawString(right_x + 5 * mm, top_y + 68 * mm, "Travel Snapshot")
    c.setFont("Helvetica", 8)
    c.setFillColor(colors.HexColor("#B8CDE6"))
    c.drawString(right_x + 5 * mm, top_y + 62 * mm, "Group booking uchun zarur ma'lumotlar")
    seat_count = int(booking["seat_count"] or booking["passenger_count"] or 1)
    snapshot_col_x = right_x + 33 * mm
    draw_label_value_fit(right_x + 5 * mm, top_y + 52 * mm, "Travelers", str(booking["passenger_count"] or 1), 18 * mm, "#7FE7FF", "#B8CDE6")
    draw_label_value_fit(snapshot_col_x, top_y + 52 * mm, "Mix", passenger_mix, right_w - 38 * mm, "#7FE7FF", "#B8CDE6")
    draw_label_value_fit(right_x + 5 * mm, top_y + 39 * mm, "Seats", str(seat_count), 18 * mm, "#FFFFFF", "#B8CDE6")
    draw_label_value_fit(snapshot_col_x, top_y + 39 * mm, "Baggage", baggage_label, right_w - 38 * mm, "#FFFFFF", "#B8CDE6")
    draw_label_value_fit(
        right_x + 5 * mm,
        top_y + 26 * mm,
        "Base fare",
        payment_base_amount,
        right_w - 10 * mm,
        "#FFFFFF",
        "#B8CDE6",
    )
    draw_label_value_fit(
        right_x + 5 * mm,
        top_y + 13 * mm,
        "Fee",
        f"{payment_fee_percent:.2f}% ({payment_fee_amount})",
        28 * mm,
        "#FFFFFF",
        "#B8CDE6",
    )
    draw_label_value_fit(snapshot_col_x, top_y + 13 * mm, "Total", payment_amount, right_w - 38 * mm, "#FFFFFF", "#B8CDE6")

    # Bottom section
    reminder_y = 47 * mm
    qr_y = 41 * mm
    qr_h = 35 * mm
    draw_glass_panel(20 * mm, reminder_y, 116 * mm, 24 * mm, "#F4FAFF", "#D5E5F6", 6 * mm, 0.8)
    draw_glass_panel(140 * mm, qr_y, w - 160 * mm, qr_h, "#FFFFFF", "#D5E5F6", 6 * mm, 0.82)
    draw_glass_panel(20 * mm, 20 * mm, w - 40 * mm, 14 * mm, "#F4FAFF", "#D5E5F6", 6 * mm, 0.78)

    c.setFont("Helvetica-Bold", 11)
    c.setFillColor(colors.HexColor("#123A63"))
    c.drawString(25 * mm, reminder_y + 16 * mm, "Safar eslatmasi")
    c.setFont("Helvetica", 8)
    c.setFillColor(colors.HexColor("#52708E"))
    c.drawString(25 * mm, reminder_y + 9 * mm, "Passportning original nusxasini olib boring va aeroportga oldindan yetib boring.")
    c.drawString(25 * mm, reminder_y + 4 * mm, f"Yo'nalish: {from_city} → {to_city}  •  Bilet: {ticket_display}")

    c.setFont("Helvetica-Bold", 11)
    c.setFillColor(colors.HexColor("#123A63"))
    c.drawString(148 * mm, qr_y + 28.4 * mm, "QR tekshiruv")
    qr_drawing = createBarcodeDrawing("QR", value=qr_value, width=22 * mm, height=22 * mm, barLevel="M")
    renderPDF.draw(qr_drawing, c, 150.5 * mm, qr_y + 6.8 * mm)
    c.setFont("Helvetica", 7.5)
    c.setFillColor(colors.HexColor("#6C84A0"))
    c.drawCentredString(161.5 * mm, qr_y + 2.9 * mm, "Bilet sahifasini ochish")

    c.setFont("Helvetica-Bold", 10.5)
    c.setFillColor(colors.HexColor("#123A63"))
    c.drawString(25 * mm, 29 * mm, "Verification Barcode")
    barcode_value = f"{ticket_no}|{user['account_id']}|{flight['route_code'] or ''}|{flight['departure_date']}"
    barcode = createBarcodeDrawing("Code128", value=barcode_value, barHeight=6.5 * mm, barWidth=0.28 * mm)
    renderPDF.draw(barcode, c, 24 * mm, 22 * mm)
    c.setFont("Helvetica", 7.5)
    c.setFillColor(colors.HexColor("#6C84A0"))
    c.drawRightString(w - 25 * mm, 26.5 * mm, f"Booking #{booking['id']} • {payment_method_label(payment_method)}")

    c.showPage()
    c.save()

    return f"tickets/{pdf_name}"


def resolve_notification_email(user: sqlite3.Row, profile: sqlite3.Row | None) -> str:
    preferred = (profile["notification_email"] if profile and profile["notification_email"] else "").strip().lower()
    if preferred and is_valid_email(preferred):
        return preferred
    fallback = (user["email"] or "").strip().lower()
    if fallback and is_valid_email(fallback):
        return fallback
    return ""


def send_ticket_email_with_attachment(
    user: sqlite3.Row,
    profile: sqlite3.Row | None,
    flight: sqlite3.Row,
    ticket_no: str,
    pdf_rel_path: str,
    payment: sqlite3.Row,
    payment_method: str,
    passenger_rows: list[sqlite3.Row],
) -> tuple[bool, str]:
    if not MAIL_ENABLED:
        return False, "MAIL_ENABLED=0"
    if not MAIL_FROM_EMAIL or not MAIL_SMTP_HOST:
        return False, "mail config yetarli emas"

    passenger_names = [str(row["full_name"] or "").strip() for row in passenger_rows if str(row["full_name"] or "").strip()]
    primary_passenger = passenger_rows[0] if passenger_rows else None
    passport = (
        (primary_passenger["passport_number"] if primary_passenger and primary_passenger["passport_number"] else None)
        or (profile["passport_number"] if profile else None)
        or "-"
    )
    recipients = notification_recipients_from_passengers(passenger_rows, user, profile)
    if not recipients:
        return False, "foydalanuvchi email topilmadi"

    pdf_abs_path = BASE_DIR / "static" / pdf_rel_path
    if not pdf_abs_path.exists():
        return False, "ticket PDF topilmadi"

    subject = f"Safar24 Ticket #{ticket_no} - {flight['from_city']} -> {flight['to_city']}"
    public_url = ticket_public_url(ticket_no)
    payment_amount = format_currency_display(
        payment["amount_value"] if payment["amount_value"] is not None else payment["amount_uzs"],
        str(payment["currency"] or "UZS"),
        int(payment["amount_uzs"] or 0),
    )
    payment_label = payment_method_label(payment_method)
    body = (
        f"Assalomu alaykum, {user['full_name']}!\n\n"
        f"Sizning bilet tayyor: {ticket_no}\n"
        f"Yo'nalish: {flight['from_city']} -> {flight['to_city']}\n"
        f"Sana: {flight['departure_date']} {flight['departure_time'] or ''}\n"
        f"To'lov: {payment_amount} ({payment_label})\n"
        f"Passport: {passport}\n\n"
        "PDF ticket ushbu xatga ilova qilindi.\n"
        f"Web ko'rinish: {public_url}\n\n"
        "Safaringiz bexatar bo'lsin!\n\n"
        "Safar24"
    )

    message = EmailMessage()
    message["Subject"] = subject
    message["From"] = MAIL_FROM_EMAIL
    message["To"] = ", ".join(recipients)
    message.set_content(body)
    message.add_alternative(
        build_ticket_email_html(
            user,
            profile,
            flight,
            ticket_no,
            public_url,
            passenger_names,
            payment_amount,
            payment_label,
        ),
        subtype="html",
    )

    with open(pdf_abs_path, "rb") as f:
        pdf_bytes = f.read()
    message.add_attachment(
        pdf_bytes,
        maintype="application",
        subtype="pdf",
        filename=pdf_abs_path.name,
    )

    try:
        deliver_email_message(message)
        return True, f"ticket yuborildi: {', '.join(recipients)}"
    except Exception as ex:
        app.logger.exception("Ticket email yuborishda xatolik")
        return False, f"email xatosi: {ex}"


def send_flight_reminder_email(
    to_email: str,
    full_name: str,
    ticket_no: str,
    flight: sqlite3.Row,
    hours_left: int,
) -> tuple[bool, str]:
    if not MAIL_ENABLED:
        return False, "MAIL_ENABLED=0"
    if not MAIL_FROM_EMAIL or not MAIL_SMTP_HOST:
        return False, "mail config yetarli emas"
    if not is_valid_email(to_email):
        return False, "email noto'g'ri"

    subject = f"Safar24 eslatma: {hours_left} soatdan keyin reys ({ticket_no})"
    body = (
        f"Assalomu alaykum, {full_name}!\n\n"
        f"Bu Safar24 tomonidan eslatma xati.\n"
        f"Sizning reysingiz {hours_left} soatdan keyin:\n"
        f"Yo'nalish: {flight['from_city']} -> {flight['to_city']}\n"
        f"Sana: {flight['departure_date']} {flight['departure_time'] or ''}\n"
        f"Ticket: {ticket_no}\n\n"
        "Iltimos, aeroportga oldindan yetib boring.\n\n"
        "Safar24"
    )

    message = EmailMessage()
    message["Subject"] = subject
    message["From"] = MAIL_FROM_EMAIL
    message["To"] = to_email
    message.set_content(body)

    try:
        deliver_email_message(message)
        return True, f"eslatma yuborildi: {to_email}"
    except Exception as ex:
        app.logger.exception("Eslatma email yuborishda xatolik")
        return False, f"email xatosi: {ex}"


def approve_payment_and_issue_ticket(
    db: sqlite3.Connection,
    payment_id: int,
    note: str,
    payment_method: str,
    payment_reference: str | None = None,
) -> tuple[str, bool, str]:
    payment = db.execute(
        """
        SELECT p.*, b.id AS booking_id, b.user_id, b.flight_id, b.passenger_count
        FROM payments p
        JOIN bookings b ON b.id = p.booking_id
        WHERE p.id = ?
        """,
        (payment_id,),
    ).fetchone()
    if not payment:
        raise ValueError("Payment topilmadi.")
    if payment["status"] == "approved":
        raise ValueError("Bu payment allaqachon tasdiqlangan.")

    booking = db.execute("SELECT * FROM bookings WHERE id = ?", (payment["booking_id"],)).fetchone()
    user = db.execute("SELECT * FROM users WHERE id = ?", (payment["user_id"],)).fetchone()
    profile = db.execute("SELECT * FROM user_profiles WHERE user_id = ?", (payment["user_id"],)).fetchone()
    flight = db.execute("SELECT * FROM flights WHERE id = ?", (payment["flight_id"],)).fetchone()
    passenger_rows = booking_passenger_rows(db, int(payment["booking_id"]))

    ticket_no = generate_next_ticket_no(db)
    pdf_rel_path = save_ticket_pdf(ticket_no, booking, user, profile, flight, payment, payment_method, passenger_rows)

    db.execute(
        """
        UPDATE payments
        SET status = 'approved', admin_note = ?, reviewed_at = ?, payment_method = ?, payment_reference = ?
        WHERE id = ?
        """,
        (note, now_iso(), payment_method, payment_reference or "", payment_id),
    )
    db.execute("UPDATE bookings SET status = 'ticketed' WHERE id = ?", (payment["booking_id"],))
    db.execute(
        """
        INSERT INTO tickets (booking_id, ticket_no, pdf_path, created_at)
        VALUES (?, ?, ?, ?)
        """,
        (payment["booking_id"], ticket_no, pdf_rel_path, now_iso()),
    )

    email_sent, email_note = send_ticket_email_with_attachment(
        user,
        profile,
        flight,
        ticket_no,
        pdf_rel_path,
        payment,
        payment_method,
        passenger_rows,
    )
    return ticket_no, email_sent, email_note


def parse_uploaded_flights(file_path: Path) -> pd.DataFrame:
    suffix = file_path.suffix.lower()
    if suffix == ".csv":
        df = pd.read_csv(file_path)
    elif suffix in {".xlsx", ".xls"}:
        df = pd.read_excel(file_path)
    else:
        raise ValueError("Faqat CSV yoki Excel (xlsx/xls) fayllar qabul qilinadi.")

    df.columns = [str(c).strip() for c in df.columns]
    missing = REQUIRED_FLIGHT_COLUMNS - set(df.columns)
    if missing:
        raise ValueError(f"Yetishmayotgan ustunlar: {', '.join(sorted(missing))}")

    if "from_airport_code" not in df.columns:
        df["from_airport_code"] = ""
    if "to_airport_code" not in df.columns:
        df["to_airport_code"] = ""
    if "departure_time" not in df.columns:
        df["departure_time"] = ""
    if "travel_class" not in df.columns:
        df["travel_class"] = "Economy"

    for field in ["price_uzs", "seats"]:
        df[field] = pd.to_numeric(df[field], errors="raise")

    for field in ["departure_date", "return_date"]:
        df[field] = df[field].astype(str)

    df["status"] = df["status"].astype(str).str.strip().str.lower()
    if not set(df["status"]).issubset(ALLOWED_FLIGHT_STATUSES):
        raise ValueError("status qiymati faqat active yoki inactive bo'lishi mumkin.")

    if (df["price_uzs"] <= 0).any():
        raise ValueError("price_uzs musbat bo'lishi shart.")
    if (df["seats"] < 0).any():
        raise ValueError("seats manfiy bo'lishi mumkin emas.")

    for field in ["departure_date", "return_date"]:
        for value in df[field]:
            value = str(value).strip()
            if value and value != "nan":
                datetime.strptime(value, "%Y-%m-%d")

    df["from_airport_code"] = df["from_airport_code"].fillna("").astype(str).str.strip().str.upper()
    df["to_airport_code"] = df["to_airport_code"].fillna("").astype(str).str.strip().str.upper()
    df["departure_time"] = df["departure_time"].fillna("").astype(str).str.strip()
    df["travel_class"] = df["travel_class"].fillna("Economy").astype(str).str.strip()

    return df


# ----------------------------
# Public pages
# ----------------------------
@app.route("/")
def index():
    flights = get_db().execute(
        """
        SELECT * FROM flights
        WHERE status = 'active' AND seats > 0
        ORDER BY departure_date ASC
        LIMIT 24
        """
    ).fetchall()
    now_dt = datetime.now()
    visible_flights = [
        flight for flight in flights
        if is_flight_upcoming(flight["departure_date"], flight["departure_time"], now_dt)
    ][:6]
    return render_template("index.html", flights=visible_flights)


@app.route("/dark.jpg")
def dark_background_image():
    target = BASE_DIR / "dark.jpg"
    if not target.exists():
        target = BASE_DIR / "static" / "dark.jpg"
    if not target.exists():
        abort(404)
    return send_file(target, mimetype="image/jpeg", max_age=3600)


@app.route("/light.jpg")
def light_background_image():
    target = BASE_DIR / "light.jpg"
    if not target.exists():
        target = BASE_DIR / "static" / "light.jpg"
    if not target.exists():
        abort(404)
    return send_file(target, mimetype="image/jpeg", max_age=3600)


@app.route("/search")
def search():
    from_city = (request.args.get("from_airport") or request.args.get("from_city") or "").strip()
    to_city = (request.args.get("to_airport") or request.args.get("to_city") or "").strip()
    depart = (request.args.get("departure_date") or "").strip()
    search_adult_count, search_child_count, search_infant_count = normalize_booking_mix_for_ui(
        parse_positive_int(request.args.get("adult_count"), default=1, minimum=1),
        parse_positive_int(request.args.get("child_count"), default=0, minimum=0),
        parse_positive_int(request.args.get("infant_count"), default=0, minimum=0),
    )

    passenger_defaults = {
        "adult_count": search_adult_count,
        "child_count": search_child_count,
        "infant_count": search_infant_count,
    }
    today_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    has_query_intent = bool(from_city or to_city or depart)
    if has_query_intent and (not from_city or not to_city):
        flash("Qayerdan va Qayerga maydonlarini to'ldiring.", "warning")
        return render_template("search.html", flights=[], today_date=today_date, passenger_defaults=passenger_defaults)

    if from_city and not is_valid_latin_search_term(from_city):
        flash("Qayerdan/Qayerga maydonlarida faqat lotin harflari ishlatiladi.", "danger")
        return render_template("search.html", flights=[], today_date=today_date, passenger_defaults=passenger_defaults)

    if to_city and not is_valid_latin_search_term(to_city):
        flash("Qayerdan/Qayerga maydonlarida faqat lotin harflari ishlatiladi.", "danger")
        return render_template("search.html", flights=[], today_date=today_date, passenger_defaults=passenger_defaults)

    normalized_from = resolve_airport_to_city(from_city)
    normalized_to = resolve_airport_to_city(to_city)
    from_code = resolve_airport_code(from_city)
    to_code = resolve_airport_code(to_city)

    clauses = [
        "SELECT * FROM flights",
        "WHERE status = 'active' AND seats > 0",
    ]
    args = []

    if from_city:
        from_parts = ["lower(from_city) LIKE ?"]
        args.append(f"%{normalized_from.lower()}%")
        if from_code:
            from_parts.append("lower(COALESCE(from_airport_code, '')) = ?")
            args.append(from_code.lower())
        clauses.append(f"AND ({' OR '.join(from_parts)})")

    if to_city:
        to_parts = ["lower(to_city) LIKE ?"]
        args.append(f"%{normalized_to.lower()}%")
        if to_code:
            to_parts.append("lower(COALESCE(to_airport_code, '')) = ?")
            args.append(to_code.lower())
        clauses.append(f"AND ({' OR '.join(to_parts)})")

    if depart:
        clauses.append("AND departure_date = ?")
        args.append(depart)

    clauses.append("ORDER BY departure_date ASC, COALESCE(departure_time, '') ASC, id DESC")

    flights = get_db().execute("\n".join(clauses), args).fetchall()
    now_dt = datetime.now()
    visible_flights = [
        flight for flight in flights
        if is_flight_upcoming(flight["departure_date"], flight["departure_time"], now_dt)
    ]
    return render_template("search.html", flights=visible_flights, today_date=today_date, passenger_defaults=passenger_defaults)


@app.route("/api/airports")
def api_airports():
    query = request.args.get("q", "")
    results = airport_search_results(query, limit=12)
    return {"results": results}


@app.route("/health")
def health():
    return {"status": "ok", "service": "safar24"}


@app.route("/<lang>/")
@app.route("/<lang>/<path:subpath>")
def localized_entry(lang: str, subpath: str = ""):
    normalized = (lang or "").strip().lower()
    if normalized not in UI_SUPPORTED_LANGUAGES:
        abort(404)

    session["ui_lang"] = normalize_ui_language(normalized)

    target_path = f"/{subpath}" if subpath else "/"
    params = request.args.to_dict(flat=True)
    params["lang"] = normalized
    query_string = urlencode(params)

    redirect_to = f"{target_path}?{query_string}" if query_string else target_path
    return redirect(redirect_to)


# ----------------------------
# User auth/profile
# ----------------------------
@app.route("/register", methods=["GET", "POST"])
def register():
    db = get_db()
    pending = session.get("pending_registration")
    if pending and registration_pending_expired(pending):
        session.pop("pending_registration", None)
        pending = None
        flash("Tasdiqlash kodi muddati tugadi. Yangi kod so'rang.", "warning")

    if request.method == "POST":
        action = (request.form.get("action") or "send_code").strip().lower()

        if action == "reset":
            session.pop("pending_registration", None)
            return redirect(url_for("register"))

        verification_code = (request.form.get("verification_code") or "").strip()

        if action == "verify" or (pending and verification_code and action not in {"resend", "reset"}):
            if not pending:
                flash("Avval tasdiqlash kodini so'rang.", "warning")
                return redirect(url_for("register"))

            if verification_code != pending.get("otp_code"):
                flash("Tasdiqlash kodi noto'g'ri.", "danger")
                return render_template(
                    "register.html",
                    verification_pending=True,
                    pending_email=pending.get("email", ""),
                    pending_full_name=pending.get("full_name", ""),
                    pending_phone=pending.get("phone", ""),
                    pending_resend_seconds=registration_resend_remaining_seconds(pending),
                )

            db = get_db()
            email = pending["email"]
            full_name = pending["full_name"]
            phone = pending["phone"]
            password_hash = pending["password_hash"]

            try:
                cur = db.execute(
                    """
                    INSERT INTO users (email, full_name, phone, password_hash, account_id, is_admin, created_at)
                    VALUES (?, ?, ?, ?, '', 0, ?)
                    """,
                    (email, full_name, phone, password_hash, now_iso()),
                )
                user_id = cur.lastrowid
                account_id = generate_account_id(user_id)
                db.execute("UPDATE users SET account_id = ? WHERE id = ?", (account_id, user_id))
                db.commit()
                session.pop("pending_registration", None)
                welcome_ok, welcome_note = send_welcome_email(email, full_name, account_id)
                if not welcome_ok:
                    app.logger.warning("Welcome email yuborilmadi: %s", welcome_note)
                flash("Ro'yxatdan o'tish muvaffaqiyatli. Endi login qiling.", "success")
                return redirect(url_for("login"))
            except sqlite3.IntegrityError:
                db.rollback()
                session.pop("pending_registration", None)
                flash("Bunday email allaqachon mavjud.", "danger")
                return redirect(url_for("register"))

        if action == "resend":
            if not pending:
                flash("Yangi kod yuborish uchun avval ma'lumot kiriting.", "warning")
                return redirect(url_for("register"))

            resend_remaining = registration_resend_remaining_seconds(pending)
            if resend_remaining > 0:
                flash(f"Kodni qayta yuborish uchun {resend_remaining} soniya kuting.", "warning")
                return render_template(
                    "register.html",
                    verification_pending=True,
                    pending_email=pending.get("email", ""),
                    pending_full_name=pending.get("full_name", ""),
                    pending_phone=pending.get("phone", ""),
                    pending_resend_seconds=resend_remaining,
                )

            otp_code = build_registration_code()
            pending["otp_code"] = otp_code
            pending["expires_at"] = (datetime.now(timezone.utc) + timedelta(minutes=REGISTRATION_CODE_MINUTES)).isoformat(timespec="seconds")
            pending["last_sent_at"] = datetime.now(timezone.utc).isoformat(timespec="seconds")
            session["pending_registration"] = pending
            ok, note = send_registration_code_email(pending["email"], pending["full_name"], otp_code)
            flash(note if ok else note, "success" if ok else "danger")
            return render_template(
                "register.html",
                verification_pending=True,
                pending_email=pending.get("email", ""),
                pending_full_name=pending.get("full_name", ""),
                pending_phone=pending.get("phone", ""),
                pending_resend_seconds=registration_resend_remaining_seconds(pending),
            )

        email = (request.form.get("email") or "").strip().lower()
        first_name = (request.form.get("first_name") or "").strip()
        last_name = (request.form.get("last_name") or "").strip()
        if not first_name and not last_name and request.form.get("full_name"):
            first_name, last_name = split_name_parts(str(request.form.get("full_name") or ""))
        full_name = build_full_name(first_name, last_name)
        phone = normalize_phone_input(request.form.get("phone"))
        password = request.form.get("password") or ""

        if not is_valid_email(email):
            flash("Email formatini to'g'ri kiriting.", "danger")
            return render_template("register.html")

        password_errors = validate_password_strength(password)
        if password_errors:
            flash(password_errors[0], "danger")
            return render_template("register.html")

        name_errors = validate_name_parts(first_name, last_name)
        if name_errors or len(full_name) < 3:
            flash(name_errors[0] if name_errors else "Ism-familiya kamida 3 ta belgidan iborat bo'lsin.", "danger")
            return render_template("register.html")

        if not PHONE_RE.match(phone):
            flash("Telefon raqam formati noto'g'ri. Masalan: +998901234567", "danger")
            return render_template("register.html")

        if db.execute("SELECT id FROM users WHERE email = ?", (email,)).fetchone():
            flash("Bunday email allaqachon mavjud. Login qiling yoki boshqa email kiriting.", "danger")
            return render_template("register.html")

        otp_code = build_registration_code()
        session["pending_registration"] = {
            "email": email,
            "full_name": full_name,
            "phone": phone,
            "password_hash": generate_password_hash(password),
            "otp_code": otp_code,
            "expires_at": (datetime.now(timezone.utc) + timedelta(minutes=REGISTRATION_CODE_MINUTES)).isoformat(timespec="seconds"),
            "last_sent_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        }

        ok, note = send_registration_code_email(email, full_name, otp_code)
        if not ok:
            session.pop("pending_registration", None)
            flash(note, "danger")
            return render_template("register.html")

        flash("Tasdiqlash kodi emailingizga yuborildi. Kodni kiriting.", "success")
        return render_template(
            "register.html",
            verification_pending=True,
            pending_email=email,
            pending_full_name=full_name,
            pending_phone=phone,
            pending_resend_seconds=registration_resend_remaining_seconds(session.get("pending_registration")),
        )

    return render_template(
        "register.html",
        verification_pending=bool(pending),
        pending_email=(pending.get("email") if pending else ""),
        pending_full_name=(pending.get("full_name") if pending else ""),
        pending_phone=(pending.get("phone") if pending else ""),
        pending_resend_seconds=registration_resend_remaining_seconds(pending),
    )


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = (request.form.get("email") or "").strip().lower()
        password = request.form.get("password") or ""
        bucket = client_bucket(email)
        nxt = (request.form.get("next") or request.args.get("next") or "").strip()

        if is_login_rate_limited(bucket):
            flash("Juda ko'p urinish bo'ldi. 15 daqiqadan keyin qayta urinib ko'ring.", "danger")
            return render_template("login.html")

        db = get_db()
        user = db.execute("SELECT * FROM users WHERE email = ?", (email,)).fetchone()

        if not user or not check_password_hash(user["password_hash"], password):
            register_failed_login(bucket)
            flash("Email yoki parol noto'g'ri.", "danger")
            return render_template("login.html")

        clear_failed_logins(bucket)
        session.clear()
        session["user_id"] = user["id"]
        session["is_admin"] = bool(user["is_admin"])
        session["is_super_admin"] = bool(user["is_super_admin"])
        session["_csrf_token"] = secrets.token_urlsafe(32)
        session["_auth_session_id"] = create_user_device_session(db, int(user["id"]))
        db.commit()

        if session["is_admin"]:
            return redirect(url_for("admin_dashboard"))
        return redirect(nxt or url_for("dashboard"))

    return render_template("login.html")


@app.route("/forgot-password", methods=["GET", "POST"])
def forgot_password():
    pending = session.get("pending_password_reset")
    if password_reset_pending_expired(pending):
        session.pop("pending_password_reset", None)
        pending = None

    if request.method == "POST":
        action = (request.form.get("action") or "send_code").strip().lower()

        if action == "reset":
            session.pop("pending_password_reset", None)
            flash("Ma'lumotlar tozalandi. Emailni qayta kiriting.", "info")
            return redirect(url_for("forgot_password"))

        if action in {"verify", "resend"} and not pending:
            flash("Avval emailingizni kiriting.", "warning")
            return redirect(url_for("forgot_password"))

        if action == "verify":
            verification_code = (request.form.get("verification_code") or "").strip()
            new_password = request.form.get("new_password") or ""
            confirm_password = request.form.get("confirm_new_password") or ""

            if not verification_code:
                flash("Tasdiqlash kodini kiriting.", "danger")
                return render_template(
                    "forgot_password.html",
                    verification_pending=True,
                    pending_email=pending.get("email", ""),
                    pending_resend_seconds=password_reset_resend_remaining_seconds(pending),
                )

            db = get_db()
            user = db.execute(
                "SELECT id, email, full_name, password_hash FROM users WHERE id = ?",
                (int(pending.get("user_id") or 0),),
            ).fetchone()
            if not user:
                session.pop("pending_password_reset", None)
                flash("Parolni tiklash sessiyasi topilmadi. Qaytadan urinib ko'ring.", "warning")
                return redirect(url_for("forgot_password"))

            errors = validate_password_strength(new_password, user["password_hash"])
            if new_password != confirm_password:
                errors.append("Yangi parollar mos emas.")

            if errors:
                flash(errors[0], "danger")
                return render_template(
                    "forgot_password.html",
                    verification_pending=True,
                    pending_email=pending.get("email", ""),
                    pending_resend_seconds=password_reset_resend_remaining_seconds(pending),
                )

            code_row = consume_security_code(db, int(user["id"]), "password_reset", verification_code)
            if not code_row:
                flash("Tasdiqlash kodi noto'g'ri yoki muddati tugagan.", "danger")
                return render_template(
                    "forgot_password.html",
                    verification_pending=True,
                    pending_email=pending.get("email", ""),
                    pending_resend_seconds=password_reset_resend_remaining_seconds(pending),
                )

            db.execute(
                "UPDATE users SET password_hash = ? WHERE id = ?",
                (generate_password_hash(new_password), user["id"]),
            )
            db.execute(
                "UPDATE user_device_sessions SET revoked_at = ? WHERE user_id = ? AND revoked_at IS NULL",
                (now_iso(), user["id"]),
            )
            db.commit()
            session.pop("pending_password_reset", None)
            flash("Parol yangilandi. Endi yangi parol bilan kiring.", "success")
            return redirect(url_for("login"))

        if action == "resend":
            wait_seconds = password_reset_resend_remaining_seconds(pending)
            if wait_seconds > 0:
                flash(f"Kodni qayta yuborish uchun {wait_seconds} soniya kuting.", "warning")
                return render_template(
                    "forgot_password.html",
                    verification_pending=True,
                    pending_email=pending.get("email", ""),
                    pending_resend_seconds=wait_seconds,
                )

            db = get_db()
            user = db.execute(
                "SELECT id, email, full_name FROM users WHERE id = ?",
                (int(pending.get("user_id") or 0),),
            ).fetchone()
            if not user:
                session.pop("pending_password_reset", None)
                flash("Parolni tiklash sessiyasi topilmadi. Qaytadan urinib ko'ring.", "warning")
                return redirect(url_for("forgot_password"))

            bucket = password_reset_bucket(str(user["email"]))
            if is_password_reset_rate_limited(bucket):
                flash("Juda ko'p urinish bo'ldi. 15 daqiqadan keyin qayta urinib ko'ring.", "danger")
                return render_template(
                    "forgot_password.html",
                    verification_pending=True,
                    pending_email=pending.get("email", ""),
                    pending_resend_seconds=0,
                )

            register_password_reset_attempt(bucket)
            code = issue_security_code(db, int(user["id"]), "password_reset", current_request_ip())
            sent, note = send_password_change_code_email(str(user["email"]), str(user["full_name"]), code)

            pending["last_sent_at"] = now_iso()
            pending["expires_at"] = (
                datetime.now(timezone.utc) + timedelta(minutes=SECURITY_CODE_MINUTES)
            ).replace(microsecond=0).isoformat()
            session["pending_password_reset"] = pending

            if sent:
                db.commit()
                flash("Tasdiqlash kodi qayta yuborildi.", "success")
            else:
                if not MAIL_ENABLED:
                    db.commit()
                    flash(f"Email xizmati o'chirilgan. Test kodi: {code}", "warning")
                else:
                    db.rollback()
                    flash("Tasdiqlash kodini yuborib bo'lmadi.", "danger")
                    if note:
                        flash(note, "warning")

            return render_template(
                "forgot_password.html",
                verification_pending=True,
                pending_email=pending.get("email", ""),
                pending_resend_seconds=password_reset_resend_remaining_seconds(pending),
            )

        email = (request.form.get("email") or "").strip().lower()
        if not is_valid_email(email):
            flash("Email formatini to'g'ri kiriting.", "danger")
            return render_template("forgot_password.html")

        bucket = password_reset_bucket(email)
        if is_password_reset_rate_limited(bucket):
            flash("Juda ko'p urinish bo'ldi. 15 daqiqadan keyin qayta urinib ko'ring.", "danger")
            return render_template("forgot_password.html")

        register_password_reset_attempt(bucket)

        db = get_db()
        user = db.execute("SELECT id, email, full_name FROM users WHERE email = ?", (email,)).fetchone()
        if not user:
            flash("Agar email tizimda mavjud bo'lsa, tasdiqlash kodi yuborildi.", "info")
            return render_template("forgot_password.html")

        code = issue_security_code(db, int(user["id"]), "password_reset", current_request_ip())
        sent, note = send_password_change_code_email(str(user["email"]), str(user["full_name"]), code)
        pending_payload = {
            "user_id": int(user["id"]),
            "email": str(user["email"]),
            "last_sent_at": now_iso(),
            "expires_at": (
                datetime.now(timezone.utc) + timedelta(minutes=SECURITY_CODE_MINUTES)
            ).replace(microsecond=0).isoformat(),
        }
        session["pending_password_reset"] = pending_payload

        if sent:
            db.commit()
            flash("Tasdiqlash kodi emailingizga yuborildi.", "success")
        else:
            if not MAIL_ENABLED:
                db.commit()
                flash(f"Email xizmati o'chirilgan. Test kodi: {code}", "warning")
            else:
                db.rollback()
                session.pop("pending_password_reset", None)
                flash("Tasdiqlash kodini yuborib bo'lmadi.", "danger")
                if note:
                    flash(note, "warning")
                return render_template("forgot_password.html")

        return render_template(
            "forgot_password.html",
            verification_pending=True,
            pending_email=pending_payload["email"],
            pending_resend_seconds=password_reset_resend_remaining_seconds(pending_payload),
        )

    return render_template(
        "forgot_password.html",
        verification_pending=bool(pending),
        pending_email=(pending.get("email") if pending else ""),
        pending_resend_seconds=password_reset_resend_remaining_seconds(pending),
    )


@app.route("/reset-password", methods=["GET", "POST"])
def reset_password():
    flash("Parolni tiklash endi emailga yuborilgan tasdiqlash kodi orqali bajariladi.", "info")
    return redirect(url_for("forgot_password"))


@app.route("/logout")
def logout():
    uid = int(session.get("user_id") or 0)
    sid = str(session.get("_auth_session_id") or "").strip()
    if uid and sid:
        db = get_db()
        db.execute(
            "UPDATE user_device_sessions SET revoked_at = COALESCE(revoked_at, ?) WHERE user_id = ? AND session_id = ?",
            (now_iso(), uid, sid),
        )
        db.commit()
    session.clear()
    return redirect(url_for("index"))


@app.route("/dashboard")
@login_required
def dashboard():
    db = get_db()
    bookings = db.execute(
        """
        SELECT b.*, f.from_city, f.from_airport_code, f.to_city, f.to_airport_code, f.departure_date, f.airline, f.price_uzs,
               f.price_value, f.price_currency, f.route_code, f.departure_time, f.travel_class,
               p.status AS payment_status,
               t.ticket_no
        FROM bookings b
        JOIN flights f ON f.id = b.flight_id
        LEFT JOIN payments p ON p.booking_id = b.id
        LEFT JOIN tickets t ON t.booking_id = b.id
        WHERE b.user_id = ?
        ORDER BY b.id DESC
        """,
        (session["user_id"],),
    ).fetchall()
    wallet_requests = db.execute(
        """
        SELECT * FROM wallet_requests
        WHERE user_id = ?
        ORDER BY id DESC
        LIMIT 10
        """,
        (session["user_id"],),
    ).fetchall()
    profile_ok = require_profile(session["user_id"])
    user_profile = db.execute("SELECT * FROM user_profiles WHERE user_id = ?", (session["user_id"],)).fetchone()
    profile_fields = [
        "passport_number",
        "passport_series",
        "birth_date",
        "nationality",
        "gender",
        "passport_issue_date",
        "passport_expiration_date",
        "notification_email",
    ]
    filled_count = sum(1 for field in profile_fields if user_profile and user_profile[field])
    if db.execute("SELECT phone FROM users WHERE id = ?", (session["user_id"],)).fetchone()["phone"]:
        filled_count += 1
    profile_completion = int((filled_count / (len(profile_fields) + 1)) * 100)

    balance_row = db.execute("SELECT balance_uzs, balance_usd FROM users WHERE id = ?", (session["user_id"],)).fetchone()
    balance_uzs = int(balance_row["balance_uzs"] if balance_row else 0)
    balance_usd = float(balance_row["balance_usd"] if balance_row else 0)
    uzs_account_no = user_wallet_account_no(session["user_id"], "UZS")
    usd_account_no = user_wallet_account_no(session["user_id"], "USD")
    return render_template(
        "dashboard.html",
        bookings=bookings,
        profile_ok=profile_ok,
        profile_completion=profile_completion,
        wallet_requests=wallet_requests,
        balance_uzs=balance_uzs,
        balance_usd=balance_usd,
        uzs_account_no=uzs_account_no,
        usd_account_no=usd_account_no,
    )


@app.route("/wallet", methods=["GET", "POST"])
@login_required
def wallet():
    db = get_db()
    user_row = db.execute("SELECT balance_uzs, balance_usd FROM users WHERE id = ?", (session["user_id"],)).fetchone()
    balance_uzs = int(user_row["balance_uzs"] if user_row else 0)
    balance_usd = float(user_row["balance_usd"] if user_row else 0)
    rate = USD_TO_UZS_RATE

    if request.method == "POST":
        amount_raw = (request.form.get("amount") or "").strip().replace(" ", "")
        currency = (request.form.get("currency") or "UZS").strip().upper()
        method = (request.form.get("payment_method") or "").strip().lower()
        note = (request.form.get("note") or "").strip()

        if method not in {"click", "payme", "manual"}:
            flash("To'lov usuli noto'g'ri.", "danger")
            return redirect(url_for("wallet"))
        if currency not in {"UZS", "USD"}:
            flash("Valyuta noto'g'ri.", "danger")
            return redirect(url_for("wallet"))

        try:
            amount_value = float(amount_raw)
        except ValueError:
            flash("Summani to'g'ri kiriting.", "danger")
            return redirect(url_for("wallet"))

        if amount_value <= 0:
            flash("Summa musbat bo'lishi kerak.", "warning")
            return redirect(url_for("wallet"))
        if currency == "UZS" and amount_value < 10000:
            flash("Minimal to'ldirish summasi 10,000 UZS.", "warning")
            return redirect(url_for("wallet"))
        if currency == "USD" and amount_value < 1:
            flash("Minimal to'ldirish summasi 1 USD.", "warning")
            return redirect(url_for("wallet"))

        amount_uzs = int(round(convert_currency(amount_value, currency, "UZS")))

        db.execute(
            """
            INSERT INTO wallet_requests (user_id, amount_uzs, amount_value, currency, payment_method, note, status, created_at)
            VALUES (?, ?, ?, ?, ?, ?, 'pending', ?)
            """,
            (session["user_id"], amount_uzs, amount_value, currency, method, note, now_iso()),
        )
        db.commit()
        flash("So'rov yuborildi. Click/Payme to'liq integratsiyasi tez kunda ishga tushadi.", "success")
        return redirect(url_for("dashboard"))

    requests = db.execute(
        """
        SELECT * FROM wallet_requests
        WHERE user_id = ?
        ORDER BY id DESC
        """,
        (session["user_id"],),
    ).fetchall()
    return render_template(
        "wallet.html",
        requests=requests,
        balance_uzs=balance_uzs,
        balance_usd=balance_usd,
        rate=rate,
        uzs_account_no=user_wallet_account_no(session["user_id"], "UZS"),
        usd_account_no=user_wallet_account_no(session["user_id"], "USD"),
    )


@app.route("/wallet/convert", methods=["POST"])
@login_required
def wallet_convert():
    db = get_db()
    from_currency = (request.form.get("from_currency") or "").strip().upper()
    to_currency = (request.form.get("to_currency") or "").strip().upper()
    amount_raw = (request.form.get("amount") or "").strip()

    if from_currency not in {"UZS", "USD"} or to_currency not in {"UZS", "USD"} or from_currency == to_currency:
        flash("Konvertatsiya valyutalari noto'g'ri tanlangan.", "danger")
        return redirect(url_for("wallet"))

    try:
        amount = float(amount_raw)
    except ValueError:
        flash("Konvertatsiya summasi noto'g'ri.", "danger")
        return redirect(url_for("wallet"))

    if amount <= 0:
        flash("Summa musbat bo'lishi kerak.", "danger")
        return redirect(url_for("wallet"))

    user = db.execute("SELECT balance_uzs, balance_usd FROM users WHERE id = ?", (session["user_id"],)).fetchone()
    if from_currency == "UZS" and int(user["balance_uzs"] or 0) < int(round(amount)):
        flash("UZS balans yetarli emas.", "warning")
        return redirect(url_for("wallet"))
    if from_currency == "USD" and float(user["balance_usd"] or 0) < amount:
        flash("USD balans yetarli emas.", "warning")
        return redirect(url_for("wallet"))

    converted = convert_currency(amount, from_currency, to_currency)

    if from_currency == "UZS":
        db.execute("UPDATE users SET balance_uzs = balance_uzs - ?, balance_usd = balance_usd + ? WHERE id = ?", (int(round(amount)), converted, session["user_id"]))
    else:
        db.execute("UPDATE users SET balance_usd = balance_usd - ?, balance_uzs = balance_uzs + ? WHERE id = ?", (amount, int(round(converted)), session["user_id"]))

    db.commit()
    flash(f"Konvertatsiya bajarildi: {amount:.2f} {from_currency} -> {converted:.2f} {to_currency}", "success")
    return redirect(url_for("wallet"))


@app.route("/profile", methods=["GET", "POST"])
@login_required
def profile():
    db = get_db()
    uid = session["user_id"]
    next_url = (request.form.get("next") or request.args.get("next") or "").strip()
    active_tab = (request.form.get("tab") or request.args.get("tab") or "profile").strip().lower()
    if active_tab not in {"profile", "security", "bookings", "devices"}:
        active_tab = "profile"

    user_row = db.execute("SELECT full_name, email, phone, password_hash FROM users WHERE id = ?", (uid,)).fetchone()
    current_phone = (user_row["phone"] if user_row and user_row["phone"] else "").strip()
    current_auth_session_id = str(session.get("_auth_session_id") or "").strip()

    bookings = db.execute(
        """
        SELECT b.*, f.from_city, f.to_city, f.departure_date, f.departure_time, f.route_code,
               f.travel_class, f.price_uzs, f.price_value, f.price_currency,
               p.status AS payment_status,
               t.ticket_no
        FROM bookings b
        JOIN flights f ON f.id = b.flight_id
        LEFT JOIN payments p ON p.booking_id = b.id
        LEFT JOIN tickets t ON t.booking_id = b.id
        WHERE b.user_id = ?
        ORDER BY b.id DESC
        LIMIT 100
        """,
        (uid,),
    ).fetchall()

    device_sessions = db.execute(
        """
        SELECT session_id, device_label, ip_address, created_at, last_seen_at
        FROM user_device_sessions
        WHERE user_id = ? AND revoked_at IS NULL
        ORDER BY CASE WHEN session_id = ? THEN 0 ELSE 1 END, last_seen_at DESC
        """,
        (uid, current_auth_session_id),
    ).fetchall()

    today = datetime.now(timezone.utc).date()
    ten_years_ago = today - timedelta(days=PASSPORT_ISSUE_MAX_AGE_YEARS * 366)
    base_exp_min = today + timedelta(days=PASSPORT_MIN_VALIDITY_DAYS)
    user_profile = db.execute("SELECT * FROM user_profiles WHERE user_id = ?", (uid,)).fetchone()
    issue_source = (request.form.get("passport_issue_date") or "").strip() if request.method == "POST" else (user_profile["passport_issue_date"] if user_profile else "")
    issue_dt = parse_date_ymd(issue_source or "")
    if issue_dt:
        base_exp_min = max(base_exp_min, issue_dt + timedelta(days=1))
    passport_expiration_max_date = (issue_dt + timedelta(days=PASSPORT_EXPIRY_MAX_YEARS * 366)).strftime("%Y-%m-%d") if issue_dt else ""

    if request.method == "POST":
        action = (request.form.get("action") or "save_profile").strip().lower()

        if action == "send_password_code":
            active_tab = "security"
            current_password = request.form.get("current_password") or ""
            if not user_row or not check_password_hash(user_row["password_hash"], current_password):
                flash("Joriy parol noto'g'ri.", "danger")
                return redirect(url_for("profile", next=next_url, tab="security"))

            code = issue_security_code(db, uid, "password_change", current_request_ip())
            sent, note = send_password_change_code_email(str(user_row["email"]), str(user_row["full_name"]), code)
            if not sent:
                db.rollback()
                flash("Tasdiqlash kodini yuborib bo'lmadi. Email sozlamasini tekshiring.", "danger")
                if note:
                    flash(note, "warning")
                return redirect(url_for("profile", next=next_url, tab="security"))

            db.commit()
            flash("Tasdiqlash kodi emailingizga yuborildi.", "success")
            return redirect(url_for("profile", next=next_url, tab="security"))

        if action == "change_password":
            active_tab = "security"
            current_password = request.form.get("current_password") or ""
            new_password = request.form.get("new_password") or ""
            confirm_new_password = request.form.get("confirm_new_password") or ""
            verification_code = (request.form.get("verification_code") or "").strip()

            if not user_row or not check_password_hash(user_row["password_hash"], current_password):
                flash("Joriy parol noto'g'ri.", "danger")
                return redirect(url_for("profile", next=next_url, tab="security"))

            password_errors = validate_password_strength(new_password, user_row["password_hash"])
            if password_errors:
                flash(password_errors[0], "danger")
                return redirect(url_for("profile", next=next_url, tab="security"))

            if new_password != confirm_new_password:
                flash("Yangi parollar mos emas.", "danger")
                return redirect(url_for("profile", next=next_url, tab="security"))

            if not verification_code:
                flash("Tasdiqlash kodini kiriting.", "danger")
                return redirect(url_for("profile", next=next_url, tab="security"))

            code_row = consume_security_code(db, uid, "password_change", verification_code)
            if not code_row:
                flash("Tasdiqlash kodi noto'g'ri yoki muddati tugagan.", "danger")
                return redirect(url_for("profile", next=next_url, tab="security"))

            db.execute("UPDATE users SET password_hash = ? WHERE id = ?", (generate_password_hash(new_password), uid))
            revoked_count = revoke_other_user_device_sessions(db, uid, current_auth_session_id)
            db.commit()
            flash("Parol muvaffaqiyatli yangilandi.", "success")
            if revoked_count > 0:
                flash(f"Xavfsizlik uchun {revoked_count} ta boshqa qurilma sessiyasi yopildi.", "info")
            return redirect(url_for("profile", next=next_url, tab="security"))

        phone = normalize_phone_input(request.form.get("phone"))
        passport_number = (request.form.get("passport_number") or "").strip().upper()
        passport_series = (request.form.get("passport_series") or "").strip().upper()
        birth_date = normalize_date_ymd((request.form.get("birth_date") or "").strip())
        nationality = (request.form.get("nationality") or "").strip()
        passport_issue_date = normalize_date_ymd((request.form.get("passport_issue_date") or "").strip())
        passport_expiration_date = normalize_date_ymd((request.form.get("passport_expiration_date") or "").strip())
        gender = (request.form.get("gender") or "").strip().lower()
        notification_email = (request.form.get("notification_email") or "").strip().lower()

        errors = validate_profile_payload(
            phone,
            passport_number,
            passport_series,
            birth_date,
            nationality,
            passport_issue_date,
            passport_expiration_date,
            gender,
            notification_email,
        )
        if errors:
            for err in errors:
                flash(err, "danger")
            return render_template(
                "profile.html",
                user_profile=user_profile,
                user_phone=phone,
                next_url=next_url,
                active_tab="profile",
                bookings=bookings,
                device_sessions=device_sessions,
                current_auth_session_id=current_auth_session_id,
                today_date=today.strftime("%Y-%m-%d"),
                ten_years_ago_date=ten_years_ago.strftime("%Y-%m-%d"),
                passport_expiration_min_date=base_exp_min.strftime("%Y-%m-%d"),
                passport_expiration_max_date=passport_expiration_max_date,
            )

        db.execute("UPDATE users SET phone = ? WHERE id = ?", (phone, uid))

        existing = db.execute("SELECT id FROM user_profiles WHERE user_id = ?", (uid,)).fetchone()
        if existing:
            db.execute(
                """
                UPDATE user_profiles
                SET passport_number = ?, passport_series = ?, birth_date = ?, nationality = ?,
                    passport_issue_date = ?, passport_expiration_date = ?, gender = ?, notification_email = ?
                WHERE user_id = ?
                """,
                (passport_number, passport_series, birth_date, nationality, 
                 passport_issue_date, passport_expiration_date, gender, notification_email, uid),
            )
        else:
            db.execute(
                """
                INSERT INTO user_profiles (user_id, passport_number, passport_series, birth_date, nationality,
                                          passport_issue_date, passport_expiration_date, gender, notification_email)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (uid, passport_number, passport_series, birth_date, nationality,
                 passport_issue_date, passport_expiration_date, gender, notification_email),
            )
        db.commit()
        flash("Profil ma'lumotlari saqlandi.", "success")
        if is_safe_internal_path(next_url):
            return redirect(next_url)
        return redirect(url_for("dashboard"))

    return render_template(
        "profile.html",
        user_profile=user_profile,
        user_phone=current_phone,
        next_url=next_url,
        active_tab=active_tab,
        bookings=bookings,
        device_sessions=device_sessions,
        current_auth_session_id=current_auth_session_id,
        today_date=today.strftime("%Y-%m-%d"),
        ten_years_ago_date=ten_years_ago.strftime("%Y-%m-%d"),
        passport_expiration_min_date=base_exp_min.strftime("%Y-%m-%d"),
        passport_expiration_max_date=passport_expiration_max_date,
    )


@app.route("/profile/devices/<session_id>/revoke", methods=["POST"])
@login_required
def profile_revoke_device(session_id: str):
    uid = int(session.get("user_id") or 0)
    current_sid = str(session.get("_auth_session_id") or "").strip()
    next_url = (request.form.get("next") or "").strip()

    if not session_id:
        flash("Sessiya topilmadi.", "warning")
        return redirect(url_for("profile", tab="devices"))

    if session_id == current_sid:
        flash("Joriy qurilmani bu yerda chiqarib bo'lmaydi.", "warning")
        return redirect(url_for("profile", tab="devices"))

    db = get_db()
    cursor = db.execute(
        """
        UPDATE user_device_sessions
        SET revoked_at = ?
        WHERE user_id = ?
          AND session_id = ?
          AND revoked_at IS NULL
        """,
        (now_iso(), uid, session_id),
    )
    db.commit()

    if int(cursor.rowcount or 0) > 0:
        flash("Qurilma sessiyasi muvaffaqiyatli chiqarib yuborildi.", "success")
    else:
        flash("Qurilma sessiyasi topilmadi yoki allaqachon yopilgan.", "warning")

    if is_safe_internal_path(next_url):
        return redirect(next_url)
    return redirect(url_for("profile", tab="devices"))


# ----------------------------
# Booking / payment / ticket
# ----------------------------
@app.route("/book/<int:flight_id>", methods=["GET", "POST"])
@login_required
def book_flight(flight_id: int):
    db = get_db()
    flight = db.execute(
        "SELECT * FROM flights WHERE id = ? AND status = 'active' AND seats > 0", (flight_id,)
    ).fetchone()
    if not flight:
        flash("Ushbu reys mavjud emas yoki joylar tugagan.", "danger")
        return redirect(url_for("search"))

    default_adult_count, default_child_count, default_infant_count = normalize_booking_mix_for_ui(
        parse_positive_int(request.args.get("adult_count"), default=1, minimum=1),
        parse_positive_int(request.args.get("child_count"), default=0, minimum=0),
        parse_positive_int(request.args.get("infant_count"), default=0, minimum=0),
    )

    booking_defaults = {
        "adult_count": default_adult_count,
        "child_count": default_child_count,
        "infant_count": default_infant_count,
    }

    if request.method == "POST":
        baggage_option = (request.form.get("baggage_option") or "standard").strip()
        if baggage_option not in {"cabin_only", "standard", "premium"}:
            flash("Bagaj turi noto'g'ri.", "danger")
            return render_template("book_flight.html", flight=flight, booking_defaults=booking_defaults)

        adult_count = parse_positive_int(request.form.get("adult_count"), default=1, minimum=1)
        child_count = parse_positive_int(request.form.get("child_count"), default=0, minimum=0)
        infant_count = parse_positive_int(request.form.get("infant_count"), default=0, minimum=0)
        passenger_count = adult_count + child_count + infant_count
        seat_count = adult_count + child_count

        booking_defaults = {
            "adult_count": adult_count,
            "child_count": child_count,
            "infant_count": infant_count,
        }

        if passenger_count < 1:
            flash("Kamida bitta yo'lovchi bo'lishi kerak.", "danger")
            return render_template("book_flight.html", flight=flight, booking_defaults=booking_defaults)

        if passenger_count > MAX_BOOKING_TRAVELERS:
            flash(f"Bitta bron uchun maksimal {MAX_BOOKING_TRAVELERS} yo'lovchi ruxsat etiladi.", "danger")
            return render_template("book_flight.html", flight=flight, booking_defaults=booking_defaults)

        if infant_count > adult_count:
            flash("Har bir INF (0-2) yo'lovchi uchun kamida bitta ADT (12+) bo'lishi kerak.", "danger")
            return render_template("book_flight.html", flight=flight, booking_defaults=booking_defaults)

        if seat_count < 1:
            flash("Kamida bitta o'rindiq band qilinishi kerak.", "danger")
            return render_template("book_flight.html", flight=flight, booking_defaults=booking_defaults)

        if seat_count > flight["seats"]:
            flash(f"Faqat {flight['seats']} ta o'rindiq qoldi (INF o'rindiqsiz hisoblanadi).", "danger")
            return render_template("book_flight.html", flight=flight, booking_defaults=booking_defaults)

        reservation_expires_at = (
            datetime.now(timezone.utc) + timedelta(minutes=BOOKING_HOLD_MINUTES)
        ).replace(microsecond=0).isoformat()

        seat_update = db.execute(
            "UPDATE flights SET seats = seats - ? WHERE id = ? AND seats >= ?",
            (seat_count, flight_id, seat_count),
        )
        if seat_update.rowcount == 0:
            db.rollback()
            flash("Tanlangan joylar allaqachon band bo'lib qolgan.", "danger")
            flight = db.execute(
                "SELECT * FROM flights WHERE id = ? AND status = 'active'",
                (flight_id,),
            ).fetchone() or flight
            return render_template("book_flight.html", flight=flight, booking_defaults=booking_defaults)

        booking_cur = db.execute(
            """
            INSERT INTO bookings (
                user_id, flight_id, baggage_option, passenger_count,
                adult_count, child_count, infant_count, seat_count,
                status, reservation_expires_at, created_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'pending_payment', ?, ?)
            """,
            (
                session["user_id"],
                flight_id,
                baggage_option,
                passenger_count,
                adult_count,
                child_count,
                infant_count,
                seat_count,
                reservation_expires_at,
                now_iso(),
            ),
        )
        booking_id = booking_cur.lastrowid

        flight_currency = (flight["price_currency"] or "UZS").upper()
        base_price_value = float(
            flight["price_value"]
            if flight["price_value"] is not None
            else (flight["price_uzs"] / USD_TO_UZS_RATE if flight_currency == "USD" else flight["price_uzs"])
        )

        fare_components = calculate_booking_fare_components(
            base_price_value,
            flight_currency,
            baggage_option,
            adult_count,
            child_count,
            infant_count,
        )
        payment_totals = calculate_payment_breakdown(
            float(fare_components["base_total_value"]),
            flight_currency,
            "wallet",
        )

        db.execute(
            """
            INSERT INTO payments (
                booking_id,
                amount_uzs,
                amount_value,
                currency,
                base_amount_uzs,
                base_amount_value,
                method_fee_pct,
                method_fee_uzs,
                method_fee_value,
                proof_image,
                status,
                created_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, '', 'not_submitted', ?)
            """,
            (
                booking_id,
                int(payment_totals["amount_uzs"]),
                float(payment_totals["amount_value"]),
                str(payment_totals["currency"]),
                int(payment_totals["base_amount_uzs"]),
                float(payment_totals["base_amount_value"]),
                float(payment_totals["method_fee_pct"]),
                int(payment_totals["method_fee_uzs"]),
                float(payment_totals["method_fee_value"]),
                now_iso(),
            ),
        )
        db.commit()

        flash("Bron yaratildi. Yo'lovchilar ma'lumotini kiriting.", "success")
        return redirect(url_for("add_passengers", booking_id=booking_id))

    return render_template("book_flight.html", flight=flight, booking_defaults=booking_defaults)


@app.route("/booking/<int:booking_id>/add-passengers", methods=["GET", "POST"])
@login_required
def add_passengers(booking_id: int):
    db = get_db()
    booking = db.execute(
        "SELECT b.*, f.from_city, f.to_city, f.departure_date FROM bookings b JOIN flights f ON f.id = b.flight_id WHERE b.id = ? AND b.user_id = ?",
        (booking_id, session["user_id"]),
    ).fetchone()
    if not booking:
        abort(404)
    
    if expire_booking_hold_if_needed(db, booking):
        db.commit()
        flash("Bron vaqti tugadi. Joylar qaytarildi, qaytadan bron qiling.", "warning")
        return redirect(url_for("dashboard"))

    existing_rows = db.execute(
        """
        SELECT id, full_name, first_name, last_name, passenger_type, passport_number, passport_series, birth_date, nationality, gender,
               passport_issue_date, passport_expiration_date, phone, notification_email
        FROM passengers
        WHERE booking_id = ?
        ORDER BY id ASC
        """,
        (booking_id,),
    ).fetchall()
    user_row = db.execute("SELECT full_name, phone, email FROM users WHERE id = ?", (session["user_id"],)).fetchone()
    profile = db.execute("SELECT * FROM user_profiles WHERE user_id = ?", (session["user_id"],)).fetchone()
    today = datetime.now(timezone.utc).date()
    ten_years_ago = today - timedelta(days=PASSPORT_ISSUE_MAX_AGE_YEARS * 366)
    passport_expiration_min = today + timedelta(days=PASSPORT_MIN_VALIDITY_DAYS)
    primary_existing = existing_rows[0] if existing_rows else None
    group_contact = {
        "phone": str((primary_existing["phone"] if primary_existing else (user_row["phone"] if user_row else "")) or ""),
        "notification_email": str(
            (
                primary_existing["notification_email"]
                if primary_existing
                else (
                    (profile["notification_email"] if profile and profile["notification_email"] else "")
                    or (user_row["email"] if user_row else "")
                )
            )
            or ""
        ),
    }

    existing_passengers: list[dict[str, str]] = []
    for idx in range(booking["passenger_count"]):
        row = existing_rows[idx] if idx < len(existing_rows) else None
        is_primary = idx == 0
        raw_full_name = str((row["full_name"] if row else (user_row["full_name"] if is_primary and user_row else "")) or "")
        first_name, last_name = passenger_name_parts(row) if row else split_name_parts(raw_full_name)
        birth_date = str((row["birth_date"] if row else (profile["birth_date"] if is_primary and profile else "")) or "")
        passenger_type_code = passenger_type_from_birth_date(birth_date, booking["departure_date"]) or normalize_passenger_type(row["passenger_type"] if row else "")
        existing_passengers.append(
            {
                "full_name": raw_full_name,
                "first_name": first_name,
                "last_name": last_name,
                "passenger_type": passenger_type_code,
                "passenger_type_label": passenger_type_label(passenger_type_code),
                "passport_number": str((row["passport_number"] if row else (profile["passport_number"] if is_primary and profile else "")) or ""),
                "passport_series": str((row["passport_series"] if row else (profile["passport_series"] if is_primary and profile else "")) or ""),
                "birth_date": birth_date,
                "nationality": str((row["nationality"] if row else (profile["nationality"] if is_primary and profile else "")) or ""),
                "gender": str((row["gender"] if row else (profile["gender"] if is_primary and profile else "")) or ""),
                "passport_issue_date": str((row["passport_issue_date"] if row else (profile["passport_issue_date"] if is_primary and profile else "")) or ""),
                "passport_expiration_date": str((row["passport_expiration_date"] if row else (profile["passport_expiration_date"] if is_primary and profile else "")) or ""),
            }
        )

    if request.method == "POST":
        group_phone = (request.form.get("group_phone") or request.form.get("passenger_1_phone") or group_contact["phone"]).strip()
        group_notification_email = (
            request.form.get("group_notification_email")
            or request.form.get("passenger_1_notification_email")
            or group_contact["notification_email"]
            or ""
        ).strip().lower()
        group_contact = {
            "phone": group_phone,
            "notification_email": group_notification_email,
        }
        passengers_payload: list[dict[str, str]] = []
        for i in range(1, booking["passenger_count"] + 1):
            first_name = (request.form.get(f"passenger_{i}_first_name") or "").strip()
            last_name = (request.form.get(f"passenger_{i}_last_name") or "").strip()
            if not first_name and not last_name and request.form.get(f"passenger_{i}_name"):
                first_name, last_name = split_name_parts(str(request.form.get(f"passenger_{i}_name") or ""))
            birth_date = (request.form.get(f"passenger_{i}_birth_date") or "").strip()
            passenger_type_code = passenger_type_from_birth_date(birth_date, booking["departure_date"])
            passenger = {
                "first_name": first_name,
                "last_name": last_name,
                "full_name": build_full_name(first_name, last_name),
                "passenger_type": passenger_type_code,
                "passport_number": (request.form.get(f"passenger_{i}_passport_number") or "").strip().upper(),
                "passport_series": (request.form.get(f"passenger_{i}_passport_series") or "").strip().upper(),
                "birth_date": birth_date,
                "nationality": (request.form.get(f"passenger_{i}_nationality") or "").strip(),
                "gender": (request.form.get(f"passenger_{i}_gender") or "").strip().lower(),
                "passport_issue_date": (request.form.get(f"passenger_{i}_passport_issue_date") or "").strip(),
                "passport_expiration_date": (request.form.get(f"passenger_{i}_passport_expiration_date") or "").strip(),
                "phone": group_phone,
                "notification_email": group_notification_email,
                "passenger_type_label": passenger_type_label(passenger_type_code),
            }
            name_errors = validate_name_parts(first_name, last_name)
            if name_errors:
                flash(f"Yo'lovchi {i}: {name_errors[0]}", "danger")
                return render_template(
                    "add_passengers.html",
                    booking=booking,
                    existing_passengers=passengers_payload + [passenger],
                    group_contact=group_contact,
                    today_date=today.strftime("%Y-%m-%d"),
                    ten_years_ago_date=ten_years_ago.strftime("%Y-%m-%d"),
                    passport_expiration_min_date=passport_expiration_min.strftime("%Y-%m-%d"),
                )
            errors = validate_passenger_payload(
                passenger["full_name"],
                passenger["phone"],
                passenger["passport_number"],
                passenger["passport_series"],
                passenger["birth_date"],
                passenger["nationality"],
                passenger["passport_issue_date"],
                passenger["passport_expiration_date"],
                passenger["gender"],
                passenger["notification_email"],
                passenger["passenger_type"],
                booking["departure_date"],
            )
            if errors:
                flash(f"Yo'lovchi {i}: {errors[0]}", "danger")
                return render_template(
                    "add_passengers.html",
                    booking=booking,
                    existing_passengers=passengers_payload + [passenger],
                    group_contact=group_contact,
                    today_date=today.strftime("%Y-%m-%d"),
                    ten_years_ago_date=ten_years_ago.strftime("%Y-%m-%d"),
                    passport_expiration_min_date=passport_expiration_min.strftime("%Y-%m-%d"),
                )
            passengers_payload.append(passenger)

        entered_mix = {"ADT": 0, "CHD": 0, "INF": 0}
        for passenger in passengers_payload:
            code = normalize_passenger_type(passenger["passenger_type"]) or "ADT"
            entered_mix[code] += 1

        expected_mix = {
            "ADT": int(booking["adult_count"] or 0),
            "CHD": int(booking["child_count"] or 0),
            "INF": int(booking["infant_count"] or 0),
        }

        if entered_mix != expected_mix:
            expected_label = passenger_mix_label(expected_mix["ADT"], expected_mix["CHD"], expected_mix["INF"])
            entered_label = passenger_mix_label(entered_mix["ADT"], entered_mix["CHD"], entered_mix["INF"])
            flash(
                f"Yosh kategoriyasi bron bilan mos emas. Kutilgan: {expected_label}. Kiritilgan: {entered_label}.",
                "danger",
            )
            return render_template(
                "add_passengers.html",
                booking=booking,
                existing_passengers=passengers_payload,
                group_contact=group_contact,
                today_date=today.strftime("%Y-%m-%d"),
                ten_years_ago_date=ten_years_ago.strftime("%Y-%m-%d"),
                passport_expiration_min_date=passport_expiration_min.strftime("%Y-%m-%d"),
            )

        db.execute("DELETE FROM passengers WHERE booking_id = ?", (booking_id,))
        for passenger in passengers_payload:
            db.execute(
                """
                INSERT INTO passengers (
                    booking_id, full_name, first_name, last_name, passenger_type, passport_number, passport_series, birth_date, nationality, gender,
                    passport_issue_date, passport_expiration_date, phone, notification_email, created_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    booking_id,
                    passenger["full_name"],
                    passenger["first_name"],
                    passenger["last_name"],
                    passenger["passenger_type"] or "ADT",
                    passenger["passport_number"],
                    passenger["passport_series"],
                    passenger["birth_date"],
                    passenger["nationality"],
                    passenger["gender"],
                    passenger["passport_issue_date"],
                    passenger["passport_expiration_date"],
                    passenger["phone"],
                    passenger["notification_email"],
                    now_iso(),
                ),
            )
        db.commit()
        flash("Yo'lovchilar ma'lumotlari saqlandi.", "success")
        return redirect(url_for("payment_upload", booking_id=booking_id))

    return render_template(
        "add_passengers.html",
        booking=booking,
        existing_passengers=existing_passengers,
        group_contact=group_contact,
        today_date=today.strftime("%Y-%m-%d"),
        ten_years_ago_date=ten_years_ago.strftime("%Y-%m-%d"),
        passport_expiration_min_date=passport_expiration_min.strftime("%Y-%m-%d"),
    )


@app.route("/booking/<int:booking_id>/cancel", methods=["POST"])
@login_required
def cancel_booking(booking_id: int):
    db = get_db()
    booking = db.execute(
        "SELECT * FROM bookings WHERE id = ? AND user_id = ?", (booking_id, session["user_id"])
    ).fetchone()
    if not booking:
        abort(404)
    
    if expire_booking_hold_if_needed(db, booking):
        db.commit()
        flash("Bron vaqti tugadi. Joylar qaytarildi.", "warning")
        return redirect(url_for("dashboard"))

    if booking["status"] not in {"pending_payment", "payment_review", "payment_rejected"}:
        flash("Bu bron bekor qilinishi mumkin emas.", "danger")
        return redirect(url_for("dashboard"))

    reason = (request.form.get("reason") or "").strip()
    release_booking_seats(db, booking, "cancelled", "cancelled_at", reason)
    db.commit()
    flash("Bron bekor qilindi.", "success")
    return redirect(url_for("dashboard"))


@app.route("/payment/<int:booking_id>", methods=["GET", "POST"])
@login_required
def payment_upload(booking_id: int):
    db = get_db()
    booking = db.execute(
        """
        SELECT b.*, f.from_city, f.to_city, f.departure_date, f.airline, f.price_uzs
               ,f.price_value, f.price_currency, f.route_code, f.from_airport_code, f.to_airport_code, f.departure_time, f.travel_class
        FROM bookings b
        JOIN flights f ON f.id = b.flight_id
        WHERE b.id = ? AND b.user_id = ?
        """,
        (booking_id, session["user_id"]),
    ).fetchone()

    if not booking:
        abort(404)
    
    if expire_booking_hold_if_needed(db, booking):
        db.commit()
        flash("Bron vaqti tugadi. Joylar qaytarildi, qaytadan bron qiling.", "warning")
        return redirect(url_for("dashboard"))

    payment = db.execute("SELECT * FROM payments WHERE booking_id = ?", (booking_id,)).fetchone()
    pax_count = db.execute("SELECT COUNT(1) AS c FROM passengers WHERE booking_id = ?", (booking_id,)).fetchone()
    profile = db.execute("SELECT * FROM user_profiles WHERE user_id = ?", (session["user_id"],)).fetchone()
    passenger_rows = booking_passenger_rows(db, booking_id)
    entered_passengers = int(pax_count["c"] if pax_count else 0)
    user_balance_row = db.execute("SELECT balance_uzs, balance_usd FROM users WHERE id = ?", (session["user_id"],)).fetchone()
    user_balance_uzs = int(user_balance_row["balance_uzs"] if user_balance_row else 0)
    user_balance_usd = float(user_balance_row["balance_usd"] if user_balance_row else 0)
    base_amount_value = float(
        payment["base_amount_value"]
        if payment["base_amount_value"] is not None
        else (
            payment["amount_value"]
            if payment["amount_value"] is not None
            else float(payment["amount_uzs"] or 0)
        )
    )
    payment_currency = normalize_currency_code(str(payment["currency"] or "UZS"))
    method_preview = (request.form.get("payment_method") or payment["payment_method"] or "wallet").strip().lower()
    if method_preview not in {"wallet", "click", "payme"}:
        method_preview = "wallet"
    payment_preview = dict(payment)
    payment_preview.update(calculate_payment_breakdown(base_amount_value, payment_currency, method_preview))
    payment_preview["payment_method"] = method_preview
    payment_fee_options = {method: payment_fee_percent(method) for method in {"wallet", "click", "payme"}}
    notification_recipients = notification_recipients_from_passengers(passenger_rows, g.user, profile)
    notification_email = ", ".join(notification_recipients) if notification_recipients else resolve_notification_email(g.user, profile)

    if entered_passengers < int(booking["passenger_count"]):
        flash("To'lovdan oldin barcha yo'lovchi ma'lumotlarini to'ldiring.", "warning")
        return redirect(url_for("add_passengers", booking_id=booking_id))

    if request.method == "POST":
        payment_method = (request.form.get("payment_method") or "wallet").strip().lower()
        confirm_details = request.form.get("confirm_details")

        if not confirm_details:
            flash("To'lovdan oldin ma'lumotlarni tekshirdim degan belgini qo'ying.", "warning")
            return render_template(
                "payment.html",
                booking=booking,
                payment=payment_preview,
                user_balance_uzs=user_balance_uzs,
                user_balance_usd=user_balance_usd,
                notification_email=notification_email,
                payment_fee_options=payment_fee_options,
            )

        if payment_method not in {"wallet", "click", "payme"}:
            flash("To'lov usuli noto'g'ri.", "danger")
            return render_template(
                "payment.html",
                booking=booking,
                payment=payment_preview,
                user_balance_uzs=user_balance_uzs,
                user_balance_usd=user_balance_usd,
                notification_email=notification_email,
                payment_fee_options=payment_fee_options,
            )

        calculated = calculate_payment_breakdown(base_amount_value, payment_currency, payment_method)
        payment_preview.update(calculated)
        payment_preview["payment_method"] = payment_method

        db.execute(
            """
            UPDATE payments
            SET payment_method = ?,
                currency = ?,
                base_amount_uzs = ?,
                base_amount_value = ?,
                method_fee_pct = ?,
                method_fee_uzs = ?,
                method_fee_value = ?,
                amount_uzs = ?,
                amount_value = ?
            WHERE id = ?
            """,
            (
                payment_method,
                str(calculated["currency"]),
                int(calculated["base_amount_uzs"]),
                float(calculated["base_amount_value"]),
                float(calculated["method_fee_pct"]),
                int(calculated["method_fee_uzs"]),
                float(calculated["method_fee_value"]),
                int(calculated["amount_uzs"]),
                float(calculated["amount_value"]),
                payment["id"],
            ),
        )

        if payment_method == "wallet":
            amount_currency = str(calculated["currency"])
            amount_value = float(calculated["amount_value"])
            if amount_currency == "USD":
                deduct = db.execute(
                    "UPDATE users SET balance_usd = balance_usd - ? WHERE id = ? AND balance_usd >= ?",
                    (amount_value, session["user_id"], amount_value),
                )
            else:
                amount_value = float(int(round(amount_value)))
                deduct = db.execute(
                    "UPDATE users SET balance_uzs = balance_uzs - ? WHERE id = ? AND balance_uzs >= ?",
                    (int(amount_value), session["user_id"], int(amount_value)),
                )
            if deduct.rowcount == 0:
                db.rollback()
                flash("Balans yetarli emas. Hisobni to'ldiring yoki boshqa to'lov usulini tanlang.", "danger")
                return render_template(
                    "payment.html",
                    booking=booking,
                    payment=payment_preview,
                    user_balance_uzs=user_balance_uzs,
                    user_balance_usd=user_balance_usd,
                    notification_email=notification_email,
                    payment_fee_options=payment_fee_options,
                )

            try:
                ticket_no, email_sent, email_note = approve_payment_and_issue_ticket(
                    db,
                    payment["id"],
                    "Auto-approved by wallet balance",
                    "wallet",
                    "",
                )
                db.commit()
                email_tail = " Email yuborilmadi." if not email_sent else " Email yuborildi."
                flash(f"To'lov walletdan muvaffaqiyatli yechildi. Ticket: {ticket_no}.{email_tail}", "success")
                if email_note:
                    flash(email_note, "info" if email_sent else "warning")
                return redirect(url_for("dashboard"))
            except ValueError as ex:
                db.rollback()
                flash(str(ex), "danger")
                return render_template(
                    "payment.html",
                    booking=booking,
                    payment=payment_preview,
                    user_balance_uzs=user_balance_uzs,
                    user_balance_usd=user_balance_usd,
                    notification_email=notification_email,
                    payment_fee_options=payment_fee_options,
                )

        if payment_method in {"click", "payme"}:
            db.execute(
                """
                UPDATE payments
                SET payment_method = ?, payment_reference = ?, status = 'pending_online', admin_note = ?
                WHERE booking_id = ?
                """,
                (
                    payment_method,
                    "",
                    "Online gateway tez kunda. Hozircha admin tekshiruvi kutiladi.",
                    booking_id,
                ),
            )
            db.execute(
                """
                INSERT INTO wallet_requests (user_id, amount_uzs, amount_value, currency, payment_method, note, status, created_at)
                VALUES (?, ?, ?, ?, ?, ?, 'pending', ?)
                """,
                (
                    session["user_id"],
                    int(calculated["amount_uzs"]),
                    float(calculated["amount_value"]),
                    str(calculated["currency"]),
                    payment_method,
                    f"Booking #{booking_id} uchun online to'lov niyati.",
                    now_iso(),
                ),
            )
            db.execute("UPDATE bookings SET status = 'payment_review' WHERE id = ?", (booking_id,))
            db.commit()
            flash("Click/Payme integratsiyasi tez kunda. So'rovingiz admin tekshiruviga yuborildi.", "warning")
            return redirect(url_for("dashboard"))

    return render_template(
        "payment.html",
        booking=booking,
        payment=payment_preview,
        user_balance_uzs=user_balance_uzs,
        user_balance_usd=user_balance_usd,
        notification_email=notification_email,
        payment_fee_options=payment_fee_options,
    )


@app.route("/my-tickets")
@login_required
def my_tickets():
    db = get_db()
    rows = db.execute(
        """
        SELECT t.ticket_no, t.pdf_path, t.created_at,
             f.from_city, f.from_airport_code, f.to_city, f.to_airport_code, f.departure_date, f.departure_time, f.airline,
             f.route_code, f.travel_class,
               b.id AS booking_id, b.passenger_count, b.adult_count, b.child_count, b.infant_count, b.seat_count, b.baggage_option,
               p.payment_method, p.currency, p.amount_value, p.amount_uzs,
               p.base_amount_value, p.base_amount_uzs, p.method_fee_pct, p.method_fee_value, p.method_fee_uzs,
             (SELECT GROUP_CONCAT(px.full_name, ' | ') FROM passengers px WHERE px.booking_id = b.id) AS passenger_names
        FROM tickets t
        JOIN bookings b ON b.id = t.booking_id
        JOIN flights f ON f.id = b.flight_id
        LEFT JOIN payments p ON p.booking_id = b.id
        WHERE b.user_id = ?
        ORDER BY t.id DESC
        """,
        (session["user_id"],),
    ).fetchall()
    return render_template("my_tickets.html", tickets=rows)


@app.route("/ticket/<ticket_no>")
def ticket_public_view(ticket_no: str):
    db = get_db()
    row = db.execute(
        """
        SELECT t.ticket_no, t.created_at,
               u.full_name,
               f.route_code, f.from_city, f.from_airport_code, f.to_city, f.to_airport_code,
               f.departure_date, f.departure_time, f.airline, f.travel_class,
             b.id AS booking_id, b.passenger_count, b.adult_count, b.child_count, b.infant_count, b.seat_count, b.baggage_option,
             p.payment_method, p.currency, p.amount_value, p.amount_uzs,
             p.base_amount_value, p.base_amount_uzs, p.method_fee_pct, p.method_fee_value, p.method_fee_uzs,
               (SELECT GROUP_CONCAT(px.full_name, ' | ') FROM passengers px WHERE px.booking_id = b.id) AS passenger_names
        FROM tickets t
        JOIN bookings b ON b.id = t.booking_id
        JOIN users u ON u.id = b.user_id
        JOIN flights f ON f.id = b.flight_id
        LEFT JOIN payments p ON p.booking_id = b.id
        WHERE t.ticket_no = ?
        """,
        (ticket_no,),
    ).fetchone()
    if not row:
        abort(404)
    return render_template("ticket_public.html", t=row)


@app.route("/ticket/<ticket_no>/download")
@login_required
def download_ticket(ticket_no: str):
    db = get_db()
    ticket = db.execute(
        """
        SELECT t.*, b.user_id
        FROM tickets t
        JOIN bookings b ON b.id = t.booking_id
        WHERE t.ticket_no = ?
        """,
        (ticket_no,),
    ).fetchone()

    if not ticket:
        abort(404)

    if not session.get("is_admin") and ticket["user_id"] != session["user_id"]:
        abort(403)

    target = BASE_DIR / "static" / ticket["pdf_path"]
    if not target.exists():
        abort(404)

    return send_file(target, as_attachment=True)


# ----------------------------
# Admin pages
# ----------------------------
@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        email = (request.form.get("email") or "").strip().lower()
        password = request.form.get("password") or ""
        bucket = client_bucket(f"admin:{email}")

        if is_login_rate_limited(bucket):
            flash("Juda ko'p urinish bo'ldi. 15 daqiqadan keyin qayta urinib ko'ring.", "danger")
            return render_template("admin_login.html")

        db = get_db()
        admin = db.execute(
            "SELECT * FROM users WHERE email = ? AND is_admin = 1", (email,)
        ).fetchone()

        if not admin or not check_password_hash(admin["password_hash"], password):
            register_failed_login(bucket)
            flash("Admin login muvaffaqiyatsiz.", "danger")
            return render_template("admin_login.html")

        clear_failed_logins(bucket)
        session.clear()
        session["user_id"] = admin["id"]
        session["is_admin"] = True
        session["is_super_admin"] = bool(admin["is_super_admin"])
        session["_csrf_token"] = secrets.token_urlsafe(32)
        session["_auth_session_id"] = create_user_device_session(db, int(admin["id"]))
        db.commit()
        return redirect(url_for("admin_dashboard"))

    return render_template("admin_login.html")


@app.route("/admin/logout")
def admin_logout():
    uid = int(session.get("user_id") or 0)
    sid = str(session.get("_auth_session_id") or "").strip()
    if uid and sid:
        db = get_db()
        db.execute(
            "UPDATE user_device_sessions SET revoked_at = COALESCE(revoked_at, ?) WHERE user_id = ? AND session_id = ?",
            (now_iso(), uid, sid),
        )
        db.commit()
    session.clear()
    return redirect(url_for("index"))


@app.route("/admin/super")
@super_admin_required
def super_admin_panel():
    db = get_db()
    current_super_id = int(session.get("user_id") or 0)
    current_role_granted_by = int(g.user["role_granted_by"]) if g.user and g.user["role_granted_by"] else None
    query = (request.args.get("q") or "").strip().lower()
    db_table = (request.args.get("db_table") or "").strip()
    db_query = (request.args.get("db_q") or "").strip()
    db_page = parse_positive_int(request.args.get("db_page"), default=1, minimum=1)

    where_clause = ""
    params: list[str] = []
    if query:
        like = f"%{query}%"
        where_clause = (
            "WHERE lower(u.full_name) LIKE ? "
            "OR lower(u.email) LIKE ? "
            "OR lower(u.account_id) LIKE ? "
            "OR lower(COALESCE(u.phone, '')) LIKE ?"
        )
        params.extend([like, like, like, like])

    users = db.execute(
        f"""
        SELECT u.id, u.full_name, u.email, u.phone, u.account_id,
             u.is_admin, u.is_super_admin, u.password_hash, u.created_at,
             u.role_granted_by, u.role_granted_at,
             granter.full_name AS role_granter_name,
             granter.email AS role_granter_email,
               (SELECT COUNT(*) FROM bookings b WHERE b.user_id = u.id) AS bookings_total
        FROM users u
         LEFT JOIN users granter ON granter.id = u.role_granted_by
        {where_clause}
        ORDER BY u.is_super_admin DESC, u.is_admin DESC, u.id DESC
        LIMIT 500
        """,
        tuple(params),
    ).fetchall()

    stats = {
        "total_accounts": db.execute("SELECT COUNT(*) AS c FROM users").fetchone()["c"],
        "total_admins": db.execute("SELECT COUNT(*) AS c FROM users WHERE is_admin = 1").fetchone()["c"],
        "total_super_admins": db.execute("SELECT COUNT(*) AS c FROM users WHERE is_super_admin = 1").fetchone()["c"],
    }

    available_tables = list_user_table_names(db)
    table_set = set(available_tables)
    if not db_table and available_tables:
        db_table = available_tables[0]
    if db_table and db_table not in table_set:
        db_table = ""

    table_overview: list[dict[str, object]] = []
    for table_name in available_tables:
        try:
            count_value = db.execute(
                f"SELECT COUNT(*) AS c FROM {quote_sql_identifier(table_name)}"
            ).fetchone()["c"]
        except sqlite3.Error:
            count_value = 0
        table_overview.append({"name": table_name, "rows": int(count_value)})

    db_preview: dict[str, object] = {
        "columns": [],
        "rows": [],
        "page": 1,
        "page_size": SUPER_DB_PAGE_SIZE,
        "total_rows": 0,
        "total_pages": 1,
        "query": db_query,
        "can_delete": False,
        "key_info": {},
        "error": "",
    }
    if db_table:
        try:
            db_preview = fetch_table_preview(db, db_table, db_query, db_page, SUPER_DB_PAGE_SIZE)
            db_page = int(db_preview["page"])
        except (ValueError, sqlite3.Error) as error:
            db_preview["error"] = str(error)

    return render_template(
        "admin_super_panel.html",
        users=users,
        q=query,
        stats=stats,
        current_super_id=current_super_id,
        current_role_granted_by=current_role_granted_by,
        impersonator_id=session.get("impersonator_id"),
        db_tables=table_overview,
        db_table=db_table,
        db_q=db_query,
        db_page=db_page,
        db_preview=db_preview,
        db_export_formats=sorted(SUPER_DB_EXPORT_FORMATS),
    )


@app.route("/admin/super/users/<int:user_id>/role", methods=["POST"])
@super_admin_required
def super_admin_update_role(user_id: int):
    role = (request.form.get("role") or "").strip().lower()
    query = (request.form.get("q") or "").strip()
    if role not in {"user", "admin", "super_admin"}:
        flash("Role noto'g'ri tanlandi.", "warning")
        return redirect(url_for("super_admin_panel", q=query))

    db = get_db()
    current_super_id = int(session.get("user_id") or 0)
    current_role_granted_by = int(g.user["role_granted_by"]) if g.user and g.user["role_granted_by"] else None

    if current_role_granted_by and int(user_id) == current_role_granted_by:
        flash("Sizga role bergan admin/super admin lavozimini o'zgartira olmaysiz.", "danger")
        return redirect(url_for("super_admin_panel", q=query))

    user = db.execute(
        "SELECT id, full_name, email, is_admin, is_super_admin, role_granted_by, role_granted_at FROM users WHERE id = ?",
        (user_id,),
    ).fetchone()
    if not user:
        abort(404)

    super_admin_count = db.execute("SELECT COUNT(*) AS c FROM users WHERE is_super_admin = 1").fetchone()["c"]
    if int(user["is_super_admin"] or 0) == 1 and role != "super_admin" and int(super_admin_count) <= 1:
        flash("Kamida bitta super admin qolishi shart.", "danger")
        return redirect(url_for("super_admin_panel", q=query))

    is_admin = 1 if role in {"admin", "super_admin"} else 0
    is_super_admin = 1 if role == "super_admin" else 0
    role_granted_by: int | None = None
    role_granted_at: str | None = None

    if role in {"admin", "super_admin"}:
        if current_super_id == int(user_id):
            role_granted_by = int(user["role_granted_by"]) if user["role_granted_by"] else None
            role_granted_at = str(user["role_granted_at"] or "") or None
        else:
            role_granted_by = current_super_id
            role_granted_at = now_iso()

    db.execute(
        "UPDATE users SET is_admin = ?, is_super_admin = ?, role_granted_by = ?, role_granted_at = ? WHERE id = ?",
        (is_admin, is_super_admin, role_granted_by, role_granted_at, user_id),
    )
    db.commit()

    if int(user_id) == current_super_id:
        session["is_admin"] = bool(is_admin)
        session["is_super_admin"] = bool(is_super_admin)

    flash(f"{user['email']} roli yangilandi: {role}", "success")

    if int(user_id) == current_super_id and not session.get("is_super_admin"):
        if session.get("is_admin"):
            return redirect(url_for("admin_dashboard"))
        return redirect(url_for("logout"))

    return redirect(url_for("super_admin_panel", q=query))


@app.route("/admin/super/users/<int:user_id>/password", methods=["POST"])
@super_admin_required
def super_admin_update_password(user_id: int):
    query = (request.form.get("q") or "").strip()
    password = (request.form.get("new_password") or "").strip()
    if not password:
        password = generate_temporary_password(12)

    if len(password) < PASSWORD_MIN_LENGTH:
        flash(f"Parol kamida {PASSWORD_MIN_LENGTH} ta belgidan iborat bo'lsin.", "warning")
        return redirect(url_for("super_admin_panel", q=query))

    db = get_db()
    user = db.execute("SELECT id, email FROM users WHERE id = ?", (user_id,)).fetchone()
    if not user:
        abort(404)

    db.execute("UPDATE users SET password_hash = ? WHERE id = ?", (generate_password_hash(password), user_id))
    db.commit()
    flash(f"{user['email']} uchun yangi parol: {password}", "success")
    return redirect(url_for("super_admin_panel", q=query))


@app.route("/admin/super/db/export", methods=["POST"])
@super_admin_required
def super_admin_export_table():
    table_name = (request.form.get("table") or "").strip()
    export_format = (request.form.get("format") or "csv").strip().lower()
    query = (request.form.get("q") or "").strip()
    db_query = (request.form.get("db_q") or "").strip()
    db_page = parse_positive_int(request.form.get("db_page"), default=1, minimum=1)
    selected_keys = request.form.getlist("record_keys")

    if export_format not in SUPER_DB_EXPORT_FORMATS:
        flash("Eksport formati noto'g'ri.", "warning")
        return redirect_super_admin_with_db_state(query, table_name, db_query, db_page)

    db = get_db()
    if table_name not in set(list_user_table_names(db)):
        flash("Jadval topilmadi.", "warning")
        return redirect_super_admin_with_db_state(query, "", db_query, db_page)

    try:
        export_columns, records, _key_info = fetch_rows_for_export(db, table_name, db_query, selected_keys)
    except (ValueError, sqlite3.Error) as error:
        flash(f"Eksportda xatolik: {error}", "danger")
        return redirect_super_admin_with_db_state(query, table_name, db_query, db_page)

    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename_base = f"{table_name}_{timestamp}"

    if export_format == "json":
        payload = json.dumps(records, ensure_ascii=False, indent=2, default=str).encode("utf-8")
        return send_file(
            BytesIO(payload),
            as_attachment=True,
            download_name=f"{filename_base}.json",
            mimetype="application/json",
        )

    if export_format == "csv":
        csv_buffer = StringIO()
        writer = csv.DictWriter(csv_buffer, fieldnames=export_columns)
        writer.writeheader()
        for row in records:
            writer.writerow({key: row.get(key, "") for key in export_columns})
        payload = csv_buffer.getvalue().encode("utf-8")
        return send_file(
            BytesIO(payload),
            as_attachment=True,
            download_name=f"{filename_base}.csv",
            mimetype="text/csv",
        )

    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = table_name[:31] or "table"
    sheet.append(export_columns)
    for row in records:
        sheet.append([row.get(column, "") for column in export_columns])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=f"{filename_base}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/admin/super/db/delete", methods=["POST"])
@super_admin_required
def super_admin_delete_table_rows():
    table_name = (request.form.get("table") or "").strip()
    query = (request.form.get("q") or "").strip()
    db_query = (request.form.get("db_q") or "").strip()
    db_page = parse_positive_int(request.form.get("db_page"), default=1, minimum=1)
    selected_keys = request.form.getlist("record_keys")

    db = get_db()
    if table_name not in set(list_user_table_names(db)):
        flash("Jadval topilmadi.", "warning")
        return redirect_super_admin_with_db_state(query, "", db_query, db_page)

    try:
        key_info = resolve_table_record_key(db, table_name)
    except (ValueError, sqlite3.Error) as error:
        flash(f"Jadval kalitini aniqlab bo'lmadi: {error}", "danger")
        return redirect_super_admin_with_db_state(query, table_name, db_query, db_page)

    cleaned_keys = sanitize_record_keys(selected_keys, key_info["mode"])
    if not cleaned_keys:
        flash("O'chirish uchun kamida bitta yozuv tanlang.", "warning")
        return redirect_super_admin_with_db_state(query, table_name, db_query, db_page)

    if table_name == "users" and key_info["mode"] == "pk" and key_info["column"] == "id":
        user_ids = {int(value) for value in cleaned_keys if str(value).isdigit()}
        current_user_id = int(session.get("user_id") or 0)
        if current_user_id in user_ids:
            flash("Joriy super admin akkauntini o'chirishga ruxsat yo'q.", "danger")
            return redirect_super_admin_with_db_state(query, table_name, db_query, db_page)

        if user_ids:
            placeholders = ",".join("?" for _ in user_ids)
            selected_super = db.execute(
                f"SELECT COUNT(*) AS c FROM users WHERE is_super_admin = 1 AND id IN ({placeholders})",
                tuple(user_ids),
            ).fetchone()["c"]
            total_super = db.execute("SELECT COUNT(*) AS c FROM users WHERE is_super_admin = 1").fetchone()["c"]
            if int(total_super) - int(selected_super) < 1:
                flash("Kamida bitta super admin qolishi kerak.", "danger")
                return redirect_super_admin_with_db_state(query, table_name, db_query, db_page)

    placeholders = ",".join("?" for _ in cleaned_keys)
    quoted_table = quote_sql_identifier(table_name)
    try:
        cursor = db.execute(
            f"DELETE FROM {quoted_table} WHERE {key_info['where_expr']} IN ({placeholders})",
            tuple(cleaned_keys),
        )
        deleted_count = cursor.rowcount if cursor.rowcount is not None and cursor.rowcount >= 0 else len(cleaned_keys)
        db.commit()
        flash(f"{deleted_count} ta yozuv o'chirildi ({table_name}).", "success")
    except sqlite3.Error as error:
        db.rollback()
        flash(f"O'chirishda xatolik: {error}", "danger")

    return redirect_super_admin_with_db_state(query, table_name, db_query, db_page)


@app.route("/admin/super/users/<int:user_id>/impersonate", methods=["POST"])
@super_admin_required
def super_admin_impersonate(user_id: int):
    query = (request.form.get("q") or "").strip()
    db = get_db()
    target = db.execute(
        "SELECT id, full_name, is_admin, is_super_admin FROM users WHERE id = ?",
        (user_id,),
    ).fetchone()
    if not target:
        abort(404)

    current_user_id = int(session.get("user_id") or 0)
    if int(target["id"]) == current_user_id:
        flash("Siz allaqachon shu akkauntdasiz.", "warning")
        return redirect(url_for("super_admin_panel", q=query))

    impersonator_id = session.get("impersonator_id") or current_user_id
    session["impersonator_id"] = impersonator_id
    session["user_id"] = target["id"]
    session["is_admin"] = bool(target["is_admin"])
    session["is_super_admin"] = bool(target["is_super_admin"])
    session["_csrf_token"] = secrets.token_urlsafe(32)
    session["_auth_session_id"] = create_user_device_session(db, int(target["id"]))
    db.commit()

    flash(f"Endi {target['full_name']} akkauntida turibsiz.", "warning")
    if session.get("is_admin"):
        return redirect(url_for("admin_dashboard"))
    return redirect(url_for("dashboard"))


@app.route("/admin/super/return")
@login_required
def super_admin_return():
    impersonator_id = session.get("impersonator_id")
    if not impersonator_id:
        return redirect(url_for("index"))

    db = get_db()
    super_user = db.execute(
        "SELECT id, is_admin, is_super_admin FROM users WHERE id = ?",
        (impersonator_id,),
    ).fetchone()
    if not super_user or int(super_user["is_super_admin"] or 0) != 1:
        session.clear()
        flash("Super admin sessiyasi topilmadi.", "danger")
        return redirect(url_for("admin_login"))

    session.clear()
    session["user_id"] = super_user["id"]
    session["is_admin"] = bool(super_user["is_admin"])
    session["is_super_admin"] = bool(super_user["is_super_admin"])
    session["_csrf_token"] = secrets.token_urlsafe(32)
    session["_auth_session_id"] = create_user_device_session(db, int(super_user["id"]))
    db.commit()
    flash("Super admin akkauntiga qaytdingiz.", "success")
    return redirect(url_for("super_admin_panel"))


ADMIN_PAGE_META = {
    "overview": {
        "title": "Admin markazi",
        "description": "Asosiy nazorat va bo'limlar bo'yicha tez o'tish.",
        "endpoint": "admin_dashboard",
    },
    "bookings": {
        "title": "Yangi bronlar",
        "description": "Yangi kelgan bronlar va navbatdagi ishlar ro'yxati.",
        "endpoint": "admin_bookings_page",
    },
    "flights": {
        "title": "Reyslar",
        "description": "Faol va eski reyslarni alohida nazorat qilish.",
        "endpoint": "admin_flights_page",
    },
    "new_flight": {
        "title": "Reys qo'shish",
        "description": "Manual qo'shish va CSV/Excel import bir joyda.",
        "endpoint": "admin_new_flight_page",
    },
    "wallet_requests": {
        "title": "So'rovnomalar",
        "description": "Wallet so'rovlari va admin qarorlari.",
        "endpoint": "admin_requests_page",
    },
    "payment_reviews": {
        "title": "To'lovlar",
        "description": "Tasdiqlanishi kerak bo'lgan to'lovlar.",
        "endpoint": "admin_payments_page",
    },
    "ticket_search": {
        "title": "Ticket qidiruv",
        "description": "Ticket, passport va account bo'yicha qidiruv.",
        "endpoint": "admin_tickets_page",
    },
    "users": {
        "title": "Userlar",
        "description": "Mijozlar ro'yxati, parol va profil nazorati.",
        "endpoint": "admin_users_page",
    },
    "accounts": {
        "title": "Akkuntlar",
        "description": "Balans va wallet harakatlarini boshqarish.",
        "endpoint": "admin_accounts_page",
    },
}


def normalize_admin_page(page: str | None, default: str = "overview") -> str:
    candidate = (page or "").strip().lower()
    return candidate if candidate in ADMIN_PAGE_META else default


def admin_page_endpoint(page: str) -> str:
    return ADMIN_PAGE_META[normalize_admin_page(page)]["endpoint"]


def admin_page_redirect(page: str, anchor: str | None = None):
    endpoint = admin_page_endpoint(page)
    return redirect(url_for(endpoint, _anchor=anchor) if anchor else url_for(endpoint))


def admin_return_page(default: str) -> str:
    return normalize_admin_page(request.form.get("next_page") or request.args.get("next_page"), default)


def render_admin_page(page: str):
    db = get_db()
    admin_page = normalize_admin_page(page)
    flight_currency = (request.args.get("flight_currency") or "all").strip().upper()
    if flight_currency not in {"ALL", "UZS", "USD"}:
        flight_currency = "ALL"

    stats = {
        "users": db.execute("SELECT COUNT(*) c FROM users WHERE is_admin = 0").fetchone()["c"],
        "flights": db.execute("SELECT COUNT(*) c FROM flights").fetchone()["c"],
        "payments_pending": db.execute(
            "SELECT COUNT(*) c FROM payments WHERE status IN ('submitted', 'pending_online')"
        ).fetchone()["c"],
        "wallet_pending": db.execute(
            "SELECT COUNT(*) c FROM wallet_requests WHERE status = 'pending'"
        ).fetchone()["c"],
        "tickets": db.execute("SELECT COUNT(*) c FROM tickets").fetchone()["c"],
    }

    pending = db.execute(
        """
         SELECT p.id AS payment_id, p.amount_uzs, p.amount_value, p.currency, p.status AS payment_status, p.proof_image,
             p.payment_method, p.payment_reference,
               b.id AS booking_id, b.status AS booking_status,
               u.full_name, u.email, u.phone, u.account_id,
               up.passport_number, up.nationality,
                             f.route_code, f.from_city, f.from_airport_code, f.to_city, f.to_airport_code,
                             f.departure_date, f.departure_time, f.travel_class, f.airline
        FROM payments p
        JOIN bookings b ON b.id = p.booking_id
        JOIN users u ON u.id = b.user_id
        LEFT JOIN user_profiles up ON up.user_id = u.id
        JOIN flights f ON f.id = b.flight_id
        WHERE p.status IN ('submitted', 'pending_online')
        ORDER BY p.id DESC
        """
    ).fetchall()

    recent_bookings = db.execute(
        """
        SELECT b.id AS booking_id, b.status AS booking_status, b.created_at,
               u.full_name, u.email, u.account_id,
               up.passport_number,
               f.route_code, f.from_city, f.from_airport_code, f.to_city, f.to_airport_code,
             f.departure_date, f.departure_time, f.travel_class, f.airline, f.price_uzs, f.price_value, f.price_currency
        FROM bookings b
        JOIN users u ON u.id = b.user_id
        LEFT JOIN user_profiles up ON up.user_id = u.id
        JOIN flights f ON f.id = b.flight_id
        WHERE b.status IN ('pending_payment', 'payment_review', 'payment_rejected')
        ORDER BY b.id DESC
        LIMIT 30
        """
    ).fetchall()

    wallet_status = (request.args.get("wallet_status") or "pending").strip().lower()
    if wallet_status not in {"pending", "approved", "rejected", "all"}:
        wallet_status = "pending"

    wallet_where = ""
    wallet_params: tuple = ()
    if wallet_status != "all":
        wallet_where = "WHERE wr.status = ?"
        wallet_params = (wallet_status,)

    wallet_requests = db.execute(
        f"""
        SELECT wr.*, u.full_name, u.email, u.account_id
        FROM wallet_requests wr
        JOIN users u ON u.id = wr.user_id
        {wallet_where}
        ORDER BY wr.id DESC
        LIMIT 100
        """,
        wallet_params,
    ).fetchall()

    users_query = (request.args.get("user_q") or "").strip()
    user_where = "WHERE u.is_admin = 0"
    user_params: list[str] = []
    if users_query:
        like = f"%{users_query.lower()}%"
        user_where += """
            AND (
                lower(u.full_name) LIKE ?
                OR lower(u.email) LIKE ?
                OR lower(COALESCE(u.phone, '')) LIKE ?
                OR lower(u.account_id) LIKE ?
            )
        """
        user_params.extend([like, like, like, like])

    account_overview = db.execute(
        f"""
        SELECT u.id, u.full_name, u.email, u.phone, u.account_id, u.balance_uzs, u.balance_usd, u.created_at,
               up.passport_number,
               (SELECT COUNT(*) FROM bookings b WHERE b.user_id = u.id) AS bookings_total,
               (SELECT COUNT(*) FROM bookings b WHERE b.user_id = u.id AND b.status = 'ticketed') AS ticketed_total,
               (SELECT COUNT(*) FROM wallet_requests wr WHERE wr.user_id = u.id) AS wallet_requests_total,
               (SELECT MAX(wr.created_at) FROM wallet_requests wr WHERE wr.user_id = u.id) AS last_wallet_activity
        FROM users u
        LEFT JOIN user_profiles up ON up.user_id = u.id
        {user_where}
        ORDER BY u.id DESC
        LIMIT 120
        """,
        tuple(user_params),
    ).fetchall()

    balance_receipt = None
    receipt_id_raw = (request.args.get("receipt_id") or "").strip()
    if receipt_id_raw.isdigit():
        balance_receipt = db.execute(
            """
            SELECT wr.id, wr.amount_uzs, wr.amount_value, wr.currency, wr.payment_method, wr.note, wr.admin_note,
                   wr.created_at, wr.reviewed_at,
                   u.id AS user_id, u.full_name, u.account_id
            FROM wallet_requests wr
            JOIN users u ON u.id = wr.user_id
            WHERE wr.id = ?
            """,
            (int(receipt_id_raw),),
        ).fetchone()

    admin_query = (request.args.get("q") or request.args.get("ticket_no") or request.args.get("passport_number") or "").strip()
    admin_search_results = []
    if admin_query:
        like = f"%{admin_query.lower()}%"
        admin_search_results = db.execute(
            """
            SELECT t.ticket_no, t.pdf_path, t.created_at,
                   u.full_name, u.email, u.account_id,
                   up.passport_number,
                                         f.route_code, f.from_city, f.from_airport_code, f.to_city, f.to_airport_code,
                                         f.departure_date, f.departure_time, f.travel_class, f.airline
            FROM tickets t
            JOIN bookings b ON b.id = t.booking_id
            JOIN users u ON u.id = b.user_id
            LEFT JOIN user_profiles up ON up.user_id = u.id
            JOIN flights f ON f.id = b.flight_id
            WHERE lower(t.ticket_no) LIKE ?
               OR lower(COALESCE(up.passport_number, '')) LIKE ?
               OR lower(u.full_name) LIKE ?
               OR lower(u.email) LIKE ?
               OR lower(u.account_id) LIKE ?
            ORDER BY t.id DESC
            LIMIT 20
            """,
            (like, like, like, like, like),
        ).fetchall()

    managed_flights = db.execute(
        """
        SELECT f.*,
               (SELECT COUNT(*) FROM bookings b WHERE b.flight_id = f.id) AS bookings_total,
               (SELECT COUNT(*) FROM bookings b WHERE b.flight_id = f.id AND b.status = 'ticketed') AS ticketed_total
        FROM flights f
        ORDER BY f.departure_date ASC, COALESCE(f.departure_time, '') ASC, f.id DESC
        LIMIT 200
        """
    ).fetchall()

    today_date = datetime.now().strftime("%Y-%m-%d")
    admin_nav = [
        {
            "slug": slug,
            "title": meta["title"],
            "description": meta["description"],
            "href": url_for(meta["endpoint"]),
            "count": (
                stats["payments_pending"] if slug == "payment_reviews"
                else stats["wallet_pending"] if slug == "wallet_requests"
                else stats["flights"] if slug in {"flights", "new_flight"}
                else stats["tickets"] if slug == "ticket_search"
                else stats["users"] if slug in {"accounts", "users"}
                else len(recent_bookings) if slug == "bookings"
                else None
            ),
        }
        for slug, meta in ADMIN_PAGE_META.items()
    ]

    return render_template(
        "admin_dashboard.html",
        admin_page=admin_page,
        admin_page_title=ADMIN_PAGE_META[admin_page]["title"],
        admin_page_description=ADMIN_PAGE_META[admin_page]["description"],
        admin_nav=admin_nav,
        stats=stats,
        pending=pending,
        recent_bookings=recent_bookings,
        wallet_requests=wallet_requests,
        account_overview=account_overview,
        users_query=users_query,
        balance_receipt=balance_receipt,
        wallet_status=wallet_status,
        flight_currency=flight_currency,
        admin_query=admin_query,
        admin_search_results=admin_search_results,
        managed_flights=managed_flights,
        today_date=today_date,
        usd_rate=USD_TO_UZS_RATE,
    )


@app.route("/admin")
@admin_required
def admin_dashboard():
    return render_admin_page("overview")


@app.route("/admin/bookings")
@admin_required
def admin_bookings_page():
    return render_admin_page("bookings")


@app.route("/admin/flights")
@admin_required
def admin_flights_page():
    return render_admin_page("flights")


@app.route("/admin/flights/new")
@admin_required
def admin_new_flight_page():
    return render_admin_page("new_flight")


@app.route("/admin/requests")
@admin_required
def admin_requests_page():
    return render_admin_page("wallet_requests")


@app.route("/admin/payments")
@admin_required
def admin_payments_page():
    return render_admin_page("payment_reviews")


@app.route("/admin/tickets")
@admin_required
def admin_tickets_page():
    return render_admin_page("ticket_search")


@app.route("/admin/accounts")
@admin_required
def admin_accounts_page():
    return render_admin_page("accounts")


@app.route("/admin/users")
@admin_required
def admin_users_page():
    return render_admin_page("users")


@app.route("/admin/users/<int:user_id>")
@admin_required
def admin_user_detail(user_id: int):
    db = get_db()
    user = db.execute(
        """
        SELECT u.*, up.passport_number, up.passport_series, up.birth_date, up.nationality,
               up.passport_issue_date, up.passport_expiration_date, up.gender, up.notification_email,
               (SELECT COUNT(*) FROM bookings b WHERE b.user_id = u.id) AS bookings_total,
               (SELECT COUNT(*) FROM bookings b WHERE b.user_id = u.id AND b.status = 'ticketed') AS ticketed_total,
               (SELECT COUNT(*) FROM wallet_requests wr WHERE wr.user_id = u.id) AS wallet_requests_total
        FROM users u
        LEFT JOIN user_profiles up ON up.user_id = u.id
        WHERE u.id = ? AND u.is_admin = 0
        """,
        (user_id,),
    ).fetchone()
    if not user:
        abort(404)

    recent_bookings = db.execute(
        """
        SELECT b.id, b.status, b.created_at, b.passenger_count, b.baggage_option,
               f.route_code, f.from_city, f.from_airport_code, f.to_city, f.to_airport_code,
               f.departure_date, f.departure_time, f.airline,
               t.ticket_no
        FROM bookings b
        JOIN flights f ON f.id = b.flight_id
        LEFT JOIN tickets t ON t.booking_id = b.id
        WHERE b.user_id = ?
        ORDER BY b.id DESC
        LIMIT 10
        """,
        (user_id,),
    ).fetchall()

    recent_wallet_requests = db.execute(
        """
        SELECT id, amount_uzs, amount_value, currency, payment_method, status, note, admin_note, created_at, reviewed_at
        FROM wallet_requests
        WHERE user_id = ?
        ORDER BY id DESC
        LIMIT 12
        """,
        (user_id,),
    ).fetchall()

    recent_payments = db.execute(
        """
        SELECT p.id, p.status, p.amount_uzs, p.amount_value, p.currency, p.payment_method, p.payment_reference, p.created_at,
               b.id AS booking_id
        FROM payments p
        JOIN bookings b ON b.id = p.booking_id
        WHERE b.user_id = ?
        ORDER BY p.id DESC
        LIMIT 12
        """,
        (user_id,),
    ).fetchall()

    balance_receipt = None
    receipt_id_raw = (request.args.get("receipt_id") or "").strip()
    if receipt_id_raw.isdigit():
        balance_receipt = db.execute(
            """
            SELECT wr.id, wr.amount_uzs, wr.amount_value, wr.currency, wr.payment_method, wr.note, wr.admin_note,
                   wr.created_at, wr.reviewed_at,
                   u.id AS user_id, u.full_name, u.account_id
            FROM wallet_requests wr
            JOIN users u ON u.id = wr.user_id
            WHERE wr.id = ? AND u.id = ?
            """,
            (int(receipt_id_raw), user_id),
        ).fetchone()

    return render_template(
        "admin_user_detail.html",
        user=user,
        recent_bookings=recent_bookings,
        recent_wallet_requests=recent_wallet_requests,
        recent_payments=recent_payments,
        balance_receipt=balance_receipt,
        usd_rate=USD_TO_UZS_RATE,
    )


@app.route("/admin/api/account-lookup")
@admin_required
def admin_account_lookup():
    account_id = (request.args.get("account_id") or "").strip().upper()
    if not account_id:
        return {"found": False, "message": "Account ID kiriting"}, 400

    user = get_db().execute(
        """
        SELECT id, full_name, account_id, email, phone, balance_uzs, balance_usd
        FROM users
        WHERE upper(account_id) = ? AND is_admin = 0
        """,
        (account_id,),
    ).fetchone()
    if not user:
        return {"found": False, "message": "Bunday account topilmadi"}, 404

    return {
        "found": True,
        "user": {
            "id": user["id"],
            "full_name": user["full_name"],
            "account_id": user["account_id"],
            "email": user["email"],
            "phone": user["phone"] or "",
            "balance_uzs": int(user["balance_uzs"] or 0),
            "balance_usd": float(user["balance_usd"] or 0),
        },
    }


@app.route("/admin/analytics", methods=["GET"])
@admin_required
def admin_analytics():
    db = get_db()
    from datetime import datetime, timedelta
    
    # Total metrics
    total_revenue = db.execute(
        "SELECT COALESCE(SUM(amount_uzs), 0) as total FROM payments WHERE status = 'approved'"
    ).fetchone()["total"]
    
    total_bookings = db.execute(
        "SELECT COUNT(*) as total FROM bookings WHERE status != 'cancelled'"
    ).fetchone()["total"]
    
    total_passengers = db.execute(
        "SELECT COUNT(*) as total FROM passengers"
    ).fetchone()["total"]
    
    cancelled_bookings = db.execute(
        "SELECT COUNT(*) as total FROM bookings WHERE status = 'cancelled'"
    ).fetchone()["total"]
    
    total_users = db.execute(
        "SELECT COUNT(*) as total FROM users WHERE is_admin = 0"
    ).fetchone()["total"]
    
    # Daily revenue (last 30 days)
    thirty_days_ago = (datetime.now() - timedelta(days=30)).isoformat()
    daily_revenue = db.execute(
        """
        SELECT DATE(p.created_at) as day, SUM(p.amount_uzs) as revenue
        FROM payments p
        WHERE p.status = 'approved' AND p.created_at >= ?
        GROUP BY DATE(p.created_at)
        ORDER BY day DESC
        """,
        (thirty_days_ago,),
    ).fetchall()
    
    # Top flights
    top_flights = db.execute(
        """
        SELECT f.route_code, f.from_city, f.to_city, COUNT(b.id) as bookings, SUM(p.amount_uzs) as revenue
        FROM flights f
        LEFT JOIN bookings b ON b.flight_id = f.id
        LEFT JOIN payments p ON p.booking_id = b.id AND p.status = 'approved'
        GROUP BY f.id
        ORDER BY bookings DESC
        LIMIT 10
        """
    ).fetchall()
    
    # Cancellation rate
    cancellation_rate = 0
    if total_bookings + cancelled_bookings > 0:
        cancellation_rate = round(100 * cancelled_bookings / (total_bookings + cancelled_bookings), 1)
    
    analytics = {
        "total_revenue": total_revenue,
        "total_bookings": total_bookings,
        "total_passengers": total_passengers,
        "total_users": total_users,
        "cancelled_bookings": cancelled_bookings,
        "cancellation_rate": cancellation_rate,
        "daily_revenue": daily_revenue,
        "top_flights": top_flights,
    }
    
    return render_template("admin_analytics.html", analytics=analytics)


@app.route("/admin/add-flight", methods=["POST"])
@admin_required
def admin_add_flight():
    db = get_db()
    next_page = admin_return_page("new_flight")
    route_code = (request.form.get("route_code") or "").strip().upper()
    from_input = (request.form.get("from_airport") or "").strip()
    to_input = (request.form.get("to_airport") or "").strip()
    departure_date = (request.form.get("departure_date") or "").strip()
    departure_time = (request.form.get("departure_time") or "").strip()
    return_date = (request.form.get("return_date") or "").strip()
    airline = (request.form.get("airline") or "").strip()
    travel_class = (request.form.get("travel_class") or "Economy").strip()
    status = (request.form.get("status") or "active").strip().lower()
    price_currency = (request.form.get("price_currency") or "UZS").strip().upper()

    try:
        price_value = float((request.form.get("price_value") or request.form.get("price_uzs") or "0").strip())
        seats = int((request.form.get("seats") or "0").strip())
    except ValueError:
        flash("Narx va joy soni raqam bo'lishi shart.", "danger")
        return admin_page_redirect(next_page)

    if price_currency not in {"UZS", "USD"}:
        flash("Narx valyutasi noto'g'ri.", "danger")
        return admin_page_redirect(next_page)

    if status not in ALLOWED_FLIGHT_STATUSES:
        flash("Status faqat active yoki inactive bo'lishi mumkin.", "danger")
        return admin_page_redirect(next_page)

    if not departure_date:
        flash("Uchish sanasi majburiy.", "danger")
        return admin_page_redirect(next_page)

    from_city = resolve_airport_to_city(from_input)
    to_city = resolve_airport_to_city(to_input)
    from_code = resolve_airport_code(from_input)
    to_code = resolve_airport_code(to_input)

    if not from_city or not to_city:
        flash("Qayerdan va qayerga maydonlarini to'g'ri kiriting.", "danger")
        return admin_page_redirect(next_page)

    if price_value <= 0 or seats < 0:
        flash("Narx musbat, joy soni manfiy bo'lmasligi kerak.", "danger")
        return admin_page_redirect(next_page)

    if price_currency == "USD":
        price_uzs = int(round(convert_currency(price_value, "USD", "UZS")))
        price_value_store = round(price_value, 2)
    else:
        price_uzs = int(round(price_value))
        price_value_store = float(int(round(price_value)))

    is_oneway = 1 if not return_date else 0
    db.execute(
        """
        INSERT INTO flights
        (route_code, from_city, from_airport_code, to_city, to_airport_code, departure_date, departure_time, return_date, is_oneway, airline, travel_class, price_uzs, price_value, price_currency, baggage_kg, include_meal, include_return, seats, status, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            route_code,
            from_city,
            from_code,
            to_city,
            to_code,
            departure_date,
            departure_time,
            return_date,
            is_oneway,
            airline,
            travel_class,
            price_uzs,
            price_value_store,
            price_currency,
            20,
            1,
            1 if not is_oneway else 0,
            seats,
            status,
            now_iso(),
        ),
    )
    db.commit()
    flash("Yangi yo'nalish qo'shildi.", "success")
    return admin_page_redirect(next_page)


@app.route("/admin/review-wallet/<int:request_id>", methods=["POST"])
@admin_required
def admin_review_wallet(request_id: int):
    action = (request.form.get("action") or "").strip().lower()
    note = (request.form.get("admin_note") or "").strip()
    next_page = admin_return_page("wallet_requests")
    db = get_db()

    req = db.execute(
        "SELECT * FROM wallet_requests WHERE id = ?",
        (request_id,),
    ).fetchone()
    if not req or req["status"] != "pending":
        flash("Ushbu wallet so'rov tekshiruv uchun mavjud emas.", "warning")
        return admin_page_redirect(next_page)

    if action == "approve":
        currency = (req["currency"] or "UZS").upper()
        amount_value = float(req["amount_value"] if req["amount_value"] is not None else req["amount_uzs"])
        if currency == "USD":
            db.execute("UPDATE users SET balance_usd = balance_usd + ? WHERE id = ?", (amount_value, req["user_id"]))
        else:
            db.execute("UPDATE users SET balance_uzs = balance_uzs + ? WHERE id = ?", (int(round(amount_value)), req["user_id"]))
        db.execute(
            """
            UPDATE wallet_requests
            SET status = 'approved', admin_note = ?, reviewed_at = ?
            WHERE id = ?
            """,
            (note, now_iso(), request_id),
        )
        db.commit()
        flash("Wallet to'ldirish so'rovi tasdiqlandi.", "success")
    elif action == "reject":
        db.execute(
            """
            UPDATE wallet_requests
            SET status = 'rejected', admin_note = ?, reviewed_at = ?
            WHERE id = ?
            """,
            (note, now_iso(), request_id),
        )
        db.commit()
        flash("Wallet to'ldirish so'rovi rad etildi.", "warning")
    else:
        flash("Noto'g'ri amal.", "danger")

    return admin_page_redirect(next_page)


@app.route("/admin/add-balance", methods=["POST"])
@admin_required
def admin_add_balance():
    db = get_db()
    next_page = admin_return_page("accounts")
    next_user_id = (request.form.get("user_id") or "").strip()
    account_id = (request.form.get("account_id") or "").strip().upper()
    currency = (request.form.get("currency") or "UZS").strip().upper()
    amount_raw = (request.form.get("amount") or "").strip()
    note = (request.form.get("admin_note") or "").strip()
    operation = (request.form.get("operation") or "add").strip().lower()

    if not account_id:
        flash("Account ID kiriting.", "warning")
        return admin_page_redirect(next_page)
    if currency not in {"UZS", "USD"}:
        flash("Valyuta noto'g'ri.", "warning")
        return admin_page_redirect(next_page)
    if operation not in {"add", "subtract"}:
        flash("Amal noto'g'ri.", "warning")
        return admin_page_redirect(next_page)

    try:
        amount = float(amount_raw)
    except ValueError:
        flash("Summa noto'g'ri.", "warning")
        return admin_page_redirect(next_page)
    if amount <= 0:
        flash("Summa musbat bo'lishi kerak.", "warning")
        return admin_page_redirect(next_page)

    user = db.execute(
        "SELECT id, full_name, balance_uzs, balance_usd FROM users WHERE upper(account_id) = ?",
        (account_id,),
    ).fetchone()
    if not user:
        flash("Bunday account topilmadi.", "danger")
        return admin_page_redirect(next_page)

    signed_amount = amount
    signed_amount_uzs = int(round(convert_currency(amount, currency, "UZS")))
    if currency == "USD":
        current_balance = float(user["balance_usd"] or 0)
        if operation == "subtract":
            if current_balance < amount:
                flash("USD balans yetarli emas.", "warning")
                return admin_page_redirect(next_page)
            db.execute("UPDATE users SET balance_usd = balance_usd - ? WHERE id = ?", (amount, user["id"]))
            balance_text = f"{amount:.2f} USD ayirildi"
            signed_amount = -amount
            signed_amount_uzs = -signed_amount_uzs
            log_method = "admin_debit"
            log_title = "Admin manual debit"
        else:
            db.execute("UPDATE users SET balance_usd = balance_usd + ? WHERE id = ?", (amount, user["id"]))
            balance_text = f"{amount:.2f} USD qo'shildi"
            log_method = "admin_topup"
            log_title = "Admin manual topup"
    else:
        current_balance = int(user["balance_uzs"] or 0)
        rounded_amount = int(round(amount))
        if operation == "subtract":
            if current_balance < rounded_amount:
                flash("UZS balans yetarli emas.", "warning")
                return admin_page_redirect(next_page)
            db.execute("UPDATE users SET balance_uzs = balance_uzs - ? WHERE id = ?", (rounded_amount, user["id"]))
            balance_text = f"{rounded_amount:,} UZS ayirildi"
            signed_amount = -rounded_amount
            signed_amount_uzs = -rounded_amount
            log_method = "admin_debit"
            log_title = "Admin manual debit"
        else:
            db.execute("UPDATE users SET balance_uzs = balance_uzs + ? WHERE id = ?", (rounded_amount, user["id"]))
            balance_text = f"{rounded_amount:,} UZS qo'shildi"
            signed_amount = rounded_amount
            signed_amount_uzs = rounded_amount
            log_method = "admin_topup"
            log_title = "Admin manual topup"

    receipt_row = db.execute(
        """
        INSERT INTO wallet_requests (user_id, amount_uzs, amount_value, currency, payment_method, note, status, admin_note, created_at, reviewed_at)
        VALUES (?, ?, ?, ?, ?, ?, 'approved', ?, ?, ?)
        """,
        (
            user["id"],
            signed_amount_uzs,
            signed_amount,
            currency,
            log_method,
            log_title,
            note,
            now_iso(),
            now_iso(),
        ),
    )
    receipt_id = receipt_row.lastrowid

    db.commit()
    flash(f"Balans yangilandi: {user['full_name']} ({account_id}) -> {balance_text}", "success")
    if next_user_id.isdigit():
        return redirect(url_for("admin_user_detail", user_id=int(next_user_id), receipt_id=receipt_id))
    return redirect(url_for(admin_page_endpoint(next_page), receipt_id=receipt_id))


@app.route("/admin/users/<int:user_id>/password", methods=["POST"])
@admin_required
def admin_update_user_password(user_id: int):
    password = request.form.get("new_password") or ""
    if len(password) < PASSWORD_MIN_LENGTH:
        flash(f"Parol kamida {PASSWORD_MIN_LENGTH} ta belgidan iborat bo'lsin.", "warning")
        return redirect(url_for("admin_user_detail", user_id=user_id))

    db = get_db()
    user = db.execute("SELECT id, full_name FROM users WHERE id = ? AND is_admin = 0", (user_id,)).fetchone()
    if not user:
        abort(404)

    db.execute("UPDATE users SET password_hash = ? WHERE id = ?", (generate_password_hash(password), user_id))
    db.commit()
    flash(f"{user['full_name']} uchun parol yangilandi.", "success")
    return redirect(url_for("admin_user_detail", user_id=user_id))


@app.route("/admin/update-flight/<int:flight_id>", methods=["POST"])
@admin_required
def admin_update_flight(flight_id: int):
    db = get_db()
    next_page = admin_return_page("flights")
    flight = db.execute("SELECT * FROM flights WHERE id = ?", (flight_id,)).fetchone()
    if not flight:
        flash("Reys topilmadi.", "warning")
        return admin_page_redirect(next_page)

    route_code = (request.form.get("route_code") or "").strip().upper()
    from_input = (request.form.get("from_airport") or "").strip()
    to_input = (request.form.get("to_airport") or "").strip()
    departure_date = (request.form.get("departure_date") or "").strip()
    departure_time = (request.form.get("departure_time") or "").strip()
    return_date = (request.form.get("return_date") or "").strip()
    airline = (request.form.get("airline") or "").strip()
    travel_class = (request.form.get("travel_class") or "Economy").strip()
    status = (request.form.get("status") or "active").strip().lower()
    price_currency = (request.form.get("price_currency") or "UZS").strip().upper()

    try:
        price_value = float((request.form.get("price_value") or request.form.get("price_uzs") or "0").strip())
        seats = int((request.form.get("seats") or "0").strip())
    except ValueError:
        flash("Narx va joy soni raqam bo'lishi shart.", "danger")
        return admin_page_redirect(next_page)

    if price_currency not in {"UZS", "USD"}:
        flash("Narx valyutasi noto'g'ri.", "danger")
        return admin_page_redirect(next_page)

    if status not in ALLOWED_FLIGHT_STATUSES:
        flash("Status faqat active yoki inactive bo'lishi mumkin.", "danger")
        return admin_page_redirect(next_page)

    if not departure_date or not airline:
        flash("Sana va aviakompaniya majburiy.", "danger")
        return admin_page_redirect(next_page)

    from_city = resolve_airport_to_city(from_input)
    to_city = resolve_airport_to_city(to_input)
    from_code = resolve_airport_code(from_input)
    to_code = resolve_airport_code(to_input)

    if not from_city or not to_city or price_value <= 0 or seats < 0:
        flash("Kiritilgan ma'lumotlarni tekshiring.", "danger")
        return admin_page_redirect(next_page)

    if price_currency == "USD":
        price_uzs = int(round(convert_currency(price_value, "USD", "UZS")))
        price_value_store = round(price_value, 2)
    else:
        price_uzs = int(round(price_value))
        price_value_store = float(int(round(price_value)))

    is_oneway = 1 if not return_date else 0
    db.execute(
        """
        UPDATE flights
        SET route_code = ?,
            from_city = ?,
            from_airport_code = ?,
            to_city = ?,
            to_airport_code = ?,
            departure_date = ?,
            departure_time = ?,
            return_date = ?,
            is_oneway = ?,
            airline = ?,
            travel_class = ?,
            price_uzs = ?,
            price_value = ?,
            price_currency = ?,
            baggage_kg = ?,
            include_meal = ?,
            include_return = ?,
            seats = ?,
            status = ?
        WHERE id = ?
        """,
        (
            route_code,
            from_city,
            from_code,
            to_city,
            to_code,
            departure_date,
            departure_time,
            return_date,
            is_oneway,
            airline,
            travel_class,
            price_uzs,
            price_value_store,
            price_currency,
            int((request.form.get("baggage_kg") or "20").strip()),
            1 if request.form.get("include_meal") else 0,
            1 if request.form.get("include_return") else 0,
            seats,
            status,
            flight_id,
        ),
    )
    db.commit()
    flash(f"Reys #{flight_id} yangilandi.", "success")
    return admin_page_redirect(next_page)


@app.route("/admin/flight/<int:flight_id>")
@admin_required
def admin_flight_detail(flight_id: int):
    db = get_db()
    flight = db.execute("SELECT * FROM flights WHERE id = ?", (flight_id,)).fetchone()
    if not flight:
        abort(404)

    passengers = db.execute(
        """
        SELECT b.id AS booking_id, b.status AS booking_status, b.created_at,
               px.full_name, px.notification_email, px.phone, u.email, u.account_id,
               px.passport_number, px.passport_series, px.birth_date, px.nationality,
               px.gender, px.passport_issue_date, px.passport_expiration_date,
               p.status AS payment_status, p.payment_method, p.payment_reference,
               t.ticket_no
        FROM bookings b
        JOIN passengers px ON px.booking_id = b.id
        JOIN users u ON u.id = b.user_id
        LEFT JOIN payments p ON p.booking_id = b.id
        LEFT JOIN tickets t ON t.booking_id = b.id
        WHERE b.flight_id = ?
        ORDER BY b.id DESC, px.id ASC
        """,
        (flight_id,),
    ).fetchall()

    return render_template("admin_flight_detail.html", flight=flight, passengers=passengers)


@app.route("/admin/flight/<int:flight_id>/export-passengers", methods=["GET"])
@admin_required
def admin_export_passengers(flight_id: int):
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    db = get_db()
    flight = db.execute("SELECT * FROM flights WHERE id = ?", (flight_id,)).fetchone()
    if not flight:
        abort(404)

    export_format = (request.args.get("export_format") or "standard").strip().lower()
    if export_format not in {"standard", "roster", "manifest"}:
        export_format = "standard"
    is_roster_export = export_format in {"roster", "manifest"}

    passengers = db.execute(
        """
        SELECT b.id AS booking_id, b.status AS booking_status, b.created_at,
               px.full_name, px.first_name, px.last_name, px.passenger_type,
               px.notification_email, px.phone, u.email, u.account_id,
               px.passport_number, px.passport_series, px.birth_date, px.nationality,
               px.gender, px.passport_issue_date, px.passport_expiration_date,
               p.status AS payment_status, p.payment_method,
               t.ticket_no
        FROM bookings b
        JOIN passengers px ON px.booking_id = b.id
        JOIN users u ON u.id = b.user_id
        LEFT JOIN payments p ON p.booking_id = b.id
        LEFT JOIN tickets t ON t.booking_id = b.id
        WHERE b.flight_id = ?
        ORDER BY b.id ASC, px.id ASC
        """,
        (flight_id,),
    ).fetchall()
    
    if not passengers:
        flash("Bu reysda yo'lovchilar yo'q.", "warning")
        return redirect(url_for("admin_flight_detail", flight_id=flight_id))

    safe_route_code = re.sub(r"[^A-Za-z0-9]+", "_", str(flight["route_code"] or "REYS")).strip("_") or "REYS"
    route_short_code = export_route_short_code(flight)
    passenger_count = len(passengers)

    if not is_roster_export:
        wb = build_centrum_export_workbook(flight, passengers)
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"{route_short_code}_HK{passenger_count}.xlsx"
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )

    wb = Workbook()
    ws = wb.active
    ws.title = f"{route_short_code[:20]}-{'RST' if is_roster_export else 'STD'}"

    header_fill = PatternFill(start_color="07111F", end_color="07111F", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    if is_roster_export:
        headers = [
            "Sequence",
            "Pax Type",
            "Title",
            "Last Name",
            "First Name",
            "Gender",
            "DOB (dd/mm/yyyy)",
            "Nationality",
            "Passport Number",
            "Passport Expiry (dd/mm/yyyy)",
            "Passport Issued Country",
            "Passport Nationality",
        ]
        title_text = "SAFAR24 PASSENGER MANIFEST"
        centered_columns = {1, 2, 3, 6, 7, 9, 10, 11, 12}
        column_widths = [4, 10, 12, 18, 18, 12, 12, 14, 18, 16, 22, 22]
    else:
        headers = [
            "#",
            "First Name",
            "Last Name",
            "Pax Type",
            "Passenger Gmail",
            "Telefon",
            "Account Email",
            "Account ID",
            "Passport Number",
            "DOB",
            "Nationality",
            "Gender",
            "Passport berilgan",
            "Passport tugash",
            "Bron ID",
            "Status",
            "To'lov",
            "Ticket",
        ]
        title_text = "SAFAR24 - REYS YO'LOVCHILARI RO'YXATI"
        centered_columns = {1, 4, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18}
        column_widths = [4, 18, 18, 10, 24, 18, 22, 16, 16, 12, 14, 12, 14, 14, 10, 12, 12, 14]

    ws["A1"] = title_text
    ws["A1"].font = Font(bold=True, size=14, color="07111F")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws["A2"] = f"Reys: {route_short_code} | {flight['from_city']} → {flight['to_city']}"
    ws["A2"].font = Font(size=10, italic=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))

    ws["A3"] = f"Sana: {format_date_dot(flight['departure_date'])} | Vaqt: {flight['departure_time'] or '--:--'} | Klass: {flight['travel_class']}"
    ws["A3"].font = Font(size=10, italic=True)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(headers))

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

    data_fill = PatternFill(start_color="F6F4ED", end_color="F6F4ED", fill_type="solid")
    alt_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    left_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    travel_date = str(flight["departure_date"] or "")
    for row_num, pax in enumerate(passengers, 6):
        is_even = (row_num - 6) % 2 == 0
        current_fill = data_fill if is_even else alt_fill

        first_name, last_name = passenger_name_parts(pax)
        passenger_type_code = passenger_type_from_birth_date(str(pax["birth_date"] or ""), travel_date) or normalize_passenger_type(pax["passenger_type"])
        if not passenger_type_code:
            passenger_type_code = "ADT"

        gender_value = str(pax["gender"] or "").strip().lower()
        gender_label = "MALE" if gender_value == "male" else ("FEMALE" if gender_value == "female" else "-")
        nationality = str(pax["nationality"] or "").strip() or "-"
        nationality_display = nationality.upper() if nationality != "-" else "-"
        passport_number = str(pax["passport_number"] or "").strip().upper()
        passport_series = str(pax["passport_series"] or "").strip().upper()
        display_passport_number = passport_number
        if passport_series and passport_number and not passport_number.startswith(passport_series):
            display_passport_number = f"{passport_series}{passport_number}"

        if is_roster_export:
            data = [
                row_num - 5,
                passenger_type_code,
                gender_label,
                (last_name.upper() if last_name and last_name != "-" else "-"),
                (first_name.upper() if first_name and first_name != "-" else "-"),
                gender_label,
                format_date_dot(pax["birth_date"]),
                nationality_display,
                display_passport_number or "-",
                format_date_dot(pax["passport_expiration_date"]),
                nationality_display,
                nationality_display,
            ]
        else:
            data = [
                row_num - 5,
                first_name or "-",
                last_name or "-",
                passenger_type_code,
                pax["notification_email"] or "-",
                pax["phone"] or "-",
                pax["email"] or "-",
                pax["account_id"] or "-",
                display_passport_number or "-",
                format_date_dot(pax["birth_date"]),
                nationality,
                gender_label,
                format_date_dot(pax["passport_issue_date"]),
                format_date_dot(pax["passport_expiration_date"]),
                pax["booking_id"] or "-",
                (pax["booking_status"] or "-").upper(),
                (pax["payment_status"] or "pending").upper(),
                pax["ticket_no"] or "TBD",
            ]

        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.fill = current_fill
            cell.border = border
            cell.alignment = center_alignment if col_num in centered_columns else left_alignment

    for idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width

    summary_row = len(passengers) + 8
    ws[f"A{summary_row}"] = "Jami yo'lovchilar:"
    ws[f"A{summary_row}"].font = Font(bold=True)
    ws[f"B{summary_row}"] = len(passengers)
    ws[f"B{summary_row}"].font = Font(bold=True)
    
    ticketed = sum(1 for p in passengers if p["ticket_no"])
    ws[f"A{summary_row + 1}"] = "Ticketlangan:"
    ws[f"B{summary_row + 1}"] = ticketed
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{'Roster' if is_roster_export else 'HK'}_{route_short_code}_HK{passenger_count}.xlsx"
    
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )


@app.route("/admin/import-flights", methods=["POST"])
@admin_required
def admin_import_flights():
    next_page = admin_return_page("new_flight")
    file = request.files.get("file")
    if not file or not file.filename:
        flash("Fayl tanlanmadi.", "danger")
        return admin_page_redirect(next_page)

    filename = secure_filename(file.filename)
    temp_path = BASE_DIR / f"_import_{int(datetime.now(timezone.utc).timestamp())}_{filename}"
    file.save(temp_path)

    db = get_db()
    inserted = 0
    try:
        df = parse_uploaded_flights(temp_path)
        for _, row in df.iterrows():
            db.execute(
                """
                INSERT INTO flights
                (route_code, from_city, from_airport_code, to_city, to_airport_code, departure_date, departure_time, return_date, airline, travel_class, price_uzs, price_value, price_currency, seats, status, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    str(row["route_code"]),
                    resolve_airport_to_city(str(row["from_city"])),
                    str(row.get("from_airport_code", "")).upper() or resolve_airport_code(str(row["from_city"])),
                    resolve_airport_to_city(str(row["to_city"])),
                    str(row.get("to_airport_code", "")).upper() or resolve_airport_code(str(row["to_city"])),
                    str(row["departure_date"]),
                    str(row.get("departure_time", "")),
                    str(row["return_date"]),
                    str(row["airline"]),
                    str(row.get("travel_class", "Economy")) or "Economy",
                    int(row["price_uzs"]),
                    float(int(row["price_uzs"])),
                    "UZS",
                    int(row["seats"]),
                    str(row["status"]).lower(),
                    now_iso(),
                ),
            )
            inserted += 1
        db.commit()
        flash(f"Import muvaffaqiyatli: {inserted} ta reys qo'shildi.", "success")
    except Exception as ex:
        flash(f"Import xatosi: {ex}", "danger")
    finally:
        if temp_path.exists():
            temp_path.unlink()

    return admin_page_redirect(next_page)


@app.route("/admin/review-payment/<int:payment_id>", methods=["POST"])
@admin_required
def admin_review_payment(payment_id: int):
    action = (request.form.get("action") or "").strip().lower()
    note = (request.form.get("admin_note") or "").strip()
    next_page = admin_return_page("payment_reviews")

    db = get_db()
    payment = db.execute(
        """
        SELECT p.*, b.id AS booking_id, b.user_id, b.flight_id, b.status AS booking_status,
               f.seats
        FROM payments p
        JOIN bookings b ON b.id = p.booking_id
        JOIN flights f ON f.id = b.flight_id
        WHERE p.id = ?
        """,
        (payment_id,),
    ).fetchone()

    if not payment or payment["status"] not in {"submitted", "pending_online"}:
        flash("To'lov holati tekshiruv uchun yaroqsiz.", "warning")
        return admin_page_redirect(next_page)
    
    booking_state = db.execute("SELECT * FROM bookings WHERE id = ?", (payment["booking_id"],)).fetchone()
    if expire_booking_hold_if_needed(db, booking_state):
        db.commit()
        flash("Bron vaqti tugagan. Joylar qaytarildi va to'lov tasdiqlanmadi.", "warning")
        return admin_page_redirect(next_page)

    if action == "approve":
        try:
            method = (payment["payment_method"] or "manual").strip().lower()
            ticket_no, email_sent, email_note = approve_payment_and_issue_ticket(db, payment_id, note, method, payment["payment_reference"])
            db.commit()
            email_tail = " Email yuborilmadi." if not email_sent else " Email yuborildi."
            flash(f"To'lov tasdiqlandi va ticket yaratildi: {ticket_no}.{email_tail}", "success")
            if email_note:
                flash(email_note, "info" if email_sent else "warning")
        except ValueError as ex:
            db.rollback()
            flash(str(ex), "danger")
    elif action == "reject":
        db.execute(
            """
            UPDATE payments
            SET status = 'rejected', admin_note = ?, reviewed_at = ?
            WHERE id = ?
            """,
            (note, now_iso(), payment_id),
        )
        db.execute("UPDATE bookings SET status = 'payment_rejected' WHERE id = ?", (payment["booking_id"],))
        db.commit()
        flash("To'lov rad etildi.", "warning")
    else:
        flash("Noto'g'ri amal.", "danger")

    return admin_page_redirect(next_page)


@app.cli.command("init-db")
def init_db_command():
    init_db()
    print("Database initialized.")


@app.cli.command("send-reminders")
def send_reminders_command():
    db = get_db()
    now = datetime.now(timezone.utc)
    targets = [24, 3]
    total_sent = 0

    rows = db.execute(
        """
        SELECT t.ticket_no, u.full_name, u.email,
               up.notification_email,
               f.from_city, f.to_city, f.departure_date, f.departure_time
        FROM tickets t
        JOIN bookings b ON b.id = t.booking_id
        JOIN users u ON u.id = b.user_id
        LEFT JOIN user_profiles up ON up.user_id = u.id
        JOIN flights f ON f.id = b.flight_id
        WHERE b.status = 'ticketed'
        """
    ).fetchall()

    for row in rows:
        dep_time = row["departure_time"] or "00:00"
        try:
            dep_dt = datetime.strptime(f"{row['departure_date']} {dep_time}", "%Y-%m-%d %H:%M")
        except ValueError:
            continue

        hours_left = int((dep_dt - now).total_seconds() // 3600)
        for target_hour in targets:
            if hours_left < target_hour or hours_left > target_hour + 1:
                continue

            reminder_type = f"{target_hour}h"
            exists = db.execute(
                "SELECT id FROM reminder_logs WHERE ticket_no = ? AND reminder_type = ?",
                (row["ticket_no"], reminder_type),
            ).fetchone()
            if exists:
                continue

            to_email = (row["notification_email"] or row["email"] or "").strip().lower()
            ok, note = send_flight_reminder_email(to_email, row["full_name"], row["ticket_no"], row, target_hour)
            if ok:
                db.execute(
                    "INSERT INTO reminder_logs (ticket_no, reminder_type, sent_to_email, created_at) VALUES (?, ?, ?, ?)",
                    (row["ticket_no"], reminder_type, to_email, now_iso()),
                )
                total_sent += 1
            else:
                print(f"[WARN] {row['ticket_no']}: {note}")

    db.commit()
    print(f"Eslatma yuborish tugadi. Yuborilgan xatlar: {total_sent}")


def ensure_dirs() -> None:
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    TICKETS_DIR.mkdir(parents=True, exist_ok=True)


def start_booking_hold_sweeper(debug_mode: bool = False) -> None:
    if os.getenv("BOOKING_HOLD_SWEEPER", "1") != "1":
        return
    if debug_mode and os.environ.get("WERKZEUG_RUN_MAIN") != "true":
        return

    def sweep_loop() -> None:
        while True:
            try:
                with app.app_context():
                    db = get_db()
                    released_count = cleanup_expired_booking_holds(db)
                    if released_count:
                        db.commit()
            except Exception:
                pass
            time.sleep(60)

    threading.Thread(target=sweep_loop, name="booking-hold-sweeper", daemon=True).start()


with app.app_context():
    ensure_dirs()
    init_db()
    ensure_schema_updates()


if __name__ == "__main__":
    debug = os.getenv("FLASK_ENV", "production") == "development"
    port = int(os.getenv("PORT", "5000"))
    start_booking_hold_sweeper(debug)
    app.run(host="0.0.0.0", port=port, debug=debug)
